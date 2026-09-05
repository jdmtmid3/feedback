import os
import csv
import base64
import io
import urllib.parse
import logging
import sys
import time
import traceback
import socket
import json
import re
import secrets
import mysql.connector
from mysql.connector import MySQLConnection, Error
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_mail import Mail, Message
import qrcode
from email_config import EmailConfig
from collections import defaultdict
from typing import List, Dict, Any, Optional
from fpdf import FPDF
from dotenv import load_dotenv
from datetime import date, datetime, timedelta, timezone
import bcrypt
from functools import wraps


load_dotenv()


# Default licensing portal URL when not configured in DB or env
DEFAULT_PORTAL_URL = os.getenv(
    "LICENSING_PORTAL_URL",
    "https://feedbacklicensing-production-c938.up.railway.app",
).rstrip("/")
LEGACY_PORTAL_URLS = {
    "https://feedbacklicensing-production.up.railway.app",
    "https://feedbacklicensing-production.up.railway.app/",
}


def create_app() -> Flask:
    app = Flask(__name__)

    # --- LOGGING SEUP ---
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        handlers=[logging.StreamHandler(sys.stdout)]
    )
    logger = logging.getLogger(__name__)

    logger.info(f"AVAILABLE ENV VARS: {list(os.environ.keys())}")

    # --- ENVIRONMENT CONFIG ---
    app.config["SECRET_KEY"] = os.getenv("SECRET_KEY") or secrets.token_hex(32)
    app.config.update(
        SESSION_COOKIE_HTTPONLY=True,
        SESSION_COOKIE_SAMESITE="Lax",
        SESSION_COOKIE_SECURE=os.getenv("RAILWAY_ENVIRONMENT") is not None,
    )

    # Database configuration handling
    # PRIORITY: 1. Railway individual variables (Most reliable)
    #           2. MYSQL_URL connection string
    #           3. Local environment / Defaults
    
    if os.getenv("MYSQLHOST"):
        logger.info("Railway individual variables detected, using them for DB config.")
        app.config["DB_CONFIG"] = {
            "host": os.getenv("MYSQLHOST"),
            "user": os.getenv("MYSQLUSER"),
            "password": os.getenv("MYSQLPASSWORD"),
            "database": os.getenv("MYSQLDATABASE"),
            "port": int(os.getenv("MYSQLPORT", 3306)),
        }
    elif os.getenv("MYSQL_URL"):
        mysql_url = os.getenv("MYSQL_URL")
        logger.info("MYSQL_URL detected, parsing connection string...")
        try:
            # Clean up the URL
            mysql_url = mysql_url.strip()
            parsed = urllib.parse.urlparse(mysql_url)
            app.config["DB_CONFIG"] = {
                "host": parsed.hostname,
                "user": parsed.username,
                "password": parsed.password,
                "database": parsed.path.lstrip('/'),
                "port": parsed.port or 3306,
            }
        except Exception as e:
            logger.error(f"CRITICAL: Failed to parse MYSQL_URL: {e}")
            app.config["DB_CONFIG"] = {"host": "localhost", "port": 3306}
    else:
        logger.info("No production variables found, falling back to local .env or defaults.")
        app.config["DB_CONFIG"] = {
            "host": os.getenv("DB_HOST", "localhost"),
            "user": os.getenv("DB_USER", "root"),
            "password": os.getenv("DB_PASSWORD", ""),
            "database": os.getenv("DB_NAME", "feedback_system"),
            "port": int(os.getenv("DB_PORT", "3306")),
        }

    db_host = app.config["DB_CONFIG"].get("host")
    db_port = app.config["DB_CONFIG"].get("port")
    db_name = app.config["DB_CONFIG"].get("database")
    logger.info(f"DB CONFIG FINALIZED: host={db_host}, port={db_port}, database={db_name}")

    # FORCE FAIL if host is still localhost on Railway
    if os.getenv("RAILWAY_ENVIRONMENT") and db_host == "localhost":
        logger.critical("FATAL: App is running on Railway but host is still 'localhost'. Check variables!")

    def get_db_connection() -> MySQLConnection:
        try:
            return mysql.connector.connect(**app.config["DB_CONFIG"])
        except Exception as e:
            logger.error(f"Failed to connect to database: {e}")
            raise

    def normalize_portal_url(value: Optional[str]) -> str:
        candidate = (value or "").strip()
        if not candidate or candidate in LEGACY_PORTAL_URLS:
            candidate = DEFAULT_PORTAL_URL
        return candidate.rstrip("/")

    def licensing_api_headers() -> Dict[str, str]:
        return {"X-Licensing-API-Key": os.getenv("LICENSING_API_KEY", "")}

    from contextlib import contextmanager

    @contextmanager
    def get_db_connection_with_transaction():
        """Context manager for database connections with automatic rollback on error."""
        conn = get_db_connection()
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def log_audit(entity_type: str, entity_id: int, action: str, old_values: str = None, new_values: str = None, user_id: str = None) -> None:
        """Log an audit entry for tracking changes"""
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                INSERT INTO audit_logs (entity_type, entity_id, action, old_values, new_values, user_id)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (entity_type, entity_id, action, old_values, new_values, user_id),
            )

    def prune_audit_logs(days: int = 90) -> int:
        """Delete audit logs older than specified days"""
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                """
                DELETE FROM audit_logs
                WHERE created_at < DATE_SUB(NOW(), INTERVAL %s DAY)
                """,
                (days,),
            )
            deleted_count = cursor.rowcount
            conn.commit()
            return deleted_count
        finally:
            conn.close()

    # License validation functions
    def get_license_config() -> Optional[Dict[str, Any]]:
        """Get license configuration from database"""
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            # Create table if it doesn't exist
            cursor.execute("CREATE TABLE IF NOT EXISTS license_config (id INT AUTO_INCREMENT PRIMARY KEY, license_key VARCHAR(255) NOT NULL, api_key VARCHAR(255) NOT NULL, licensing_portal_url VARCHAR(255) DEFAULT 'https://feedbacklicensing-production-c938.up.railway.app', main_system_url VARCHAR(255) NULL, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP)")
            # Check if main_system_url column exists
            cursor.execute("SHOW COLUMNS FROM license_config LIKE 'main_system_url'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE license_config ADD COLUMN main_system_url VARCHAR(255) NULL AFTER licensing_portal_url")
                conn.commit()
            cursor.execute("SELECT * FROM license_config ORDER BY id DESC LIMIT 1")
            return cursor.fetchone()
        finally:
            conn.close()

    def validate_license_from_portal() -> Dict[str, Any]:
        """Validate license by calling the licensing portal API"""
        import requests
        from requests.exceptions import RequestException, Timeout
        
        config = get_license_config()
        if not config:
            return {"valid": False, "error": "No license configured"}
        
        try:
            portal_url = normalize_portal_url(config.get("licensing_portal_url"))
            response = requests.post(
                f"{portal_url}/api/validate/{config['license_key']}",
                headers=licensing_api_headers(), timeout=10
            )
            
            if response.status_code == 200:
                return response.json()
            else:
                logger.error(f"License validation failed: {response.status_code}")
                return {"valid": False, "error": "License validation failed"}
        except Timeout:
            logger.error("License validation request timed out")
            return {"valid": False, "error": "Request timed out"}
        except RequestException as e:
            logger.error(f"License validation network error: {e}")
            return {"valid": False, "error": "Network error"}
        except Exception as e:
            logger.error(f"Unexpected error validating license: {e}")
            return {"valid": False, "error": str(e)}

    license_status_cache: Dict[str, Dict[str, Any]] = {}

    def _parse_license_expiry(data: Dict[str, Any]) -> datetime | None:
        raw = (data.get("expiry_date") or data.get("expires_at")
               or data.get("expiration_date") or data.get("expiry"))
        if not raw:
            return None
        try:
            value = str(raw).strip()
            date_only = len(value) == 10
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=timezone.utc)
            if date_only:
                parsed = parsed.replace(hour=23, minute=59, second=59)
            return parsed.astimezone(timezone.utc)
        except (TypeError, ValueError):
            logger.warning("Unable to parse license expiry value: %r", raw)
            return None

    def _license_is_expired(data: Dict[str, Any]) -> bool:
        expiry = _parse_license_expiry(data)
        if expiry and expiry <= datetime.now(timezone.utc):
            return True
        message = str(data.get("message") or data.get("error") or "").lower()
        return bool(data.get("expired") or "expired" in message)

    def validate_tenant_license(license_key: str, force: bool = False) -> Dict[str, Any]:
        """Validate one client license at most once per hour per web worker."""
        if not license_key:
            return {"valid": False, "error": "No license configured"}
        cached = license_status_cache.get(license_key)
        if cached and not force and time.time() - cached["checked_at"] < 3600:
            return cached["data"]
        config = get_license_config()
        portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
        try:
            import requests as http_requests
            response = http_requests.post(
                f"{portal_url}/api/validate/{license_key}",
                headers=licensing_api_headers(), timeout=10,
            )
            data = response.json() if response.content else {}
            if response.status_code != 200 and not data:
                data = {"valid": False, "error": f"License API error: {response.status_code}"}
            license_status_cache[license_key] = {"checked_at": time.time(), "data": data}
            return data
        except Exception as exc:
            logger.error("Unable to validate tenant license %s: %s", license_key[:8], exc)
            # A temporary portal/network failure must not lock out a client.
            return cached["data"] if cached else {"valid": None, "error": "License service unavailable"}

    def _user_license_keys_for_access(user: Dict[str, Any]) -> List[str]:
        if not user or user.get("role") == "superadmin":
            return []
        if user.get("role") == "admin":
            return [user["license_key"]] if user.get("license_key") else []
        if user.get("role") != "user":
            return []
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                """SELECT DISTINCT owner.license_key
                   FROM user_stores us
                   JOIN stores s ON s.id = us.store_id
                   JOIN users owner ON owner.id = s.user_id
                   WHERE us.user_id = %s AND owner.license_key IS NOT NULL""",
                (int(user["id"]),),
            )
            return [row[0] for row in cursor.fetchall() if row[0]]
        finally:
            conn.close()

    def _expired_license_for_user(user: Dict[str, Any], force: bool = False) -> str | None:
        for key in _user_license_keys_for_access(user):
            if _license_is_expired(validate_tenant_license(key, force=force)):
                return key
        return None

    @app.context_processor
    def inject_license_expiry_warning():
        if session.get("role") != "admin" or "user_id" not in session:
            return {}
        user = get_user_by_id(session["user_id"])
        if not user or not user.get("license_key"):
            return {}
        status = validate_tenant_license(user["license_key"])
        expiry = _parse_license_expiry(status)
        if not expiry:
            return {}
        seconds_left = int((expiry - datetime.now(timezone.utc)).total_seconds())
        if seconds_left <= 0 or seconds_left > 30 * 24 * 60 * 60:
            return {}
        return {"license_expiry_warning": {
            "expires_at": expiry.isoformat(),
            "expiry_date": expiry.strftime("%B %d, %Y"),
        }}

    def check_store_limit() -> bool:
        """Check if the current store count is within the license limit"""
        config = get_license_config()
        
        # If no license is configured, block store creation
        if not config or not config.get("license_key"):
            return False
        
        license_status = validate_license_from_portal()
        
        # If license validation fails (invalid key, etc.), block store creation
        if not license_status.get("valid"):
            return False
        
        max_stores = license_status.get("max_stores", 0)
        if max_stores == 0:
            return True  # 0 means unlimited
        
        # Count current stores
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM stores")
            current_count = cursor.fetchone()[0]
            return current_count < max_stores
        finally:
            conn.close()

    # Authentication helper functions
    def get_user_by_username(username: str) -> Optional[Dict[str, Any]]:
        """Get user by username"""
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users WHERE username = %s", (username,))
            return cursor.fetchone()
        finally:
            conn.close()

    def get_user_by_id(user_id: int) -> Optional[Dict[str, Any]]:
        """Get user by ID"""
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users WHERE id = %s", (user_id,))
            return cursor.fetchone()
        finally:
            conn.close()

    def verify_password(password: str, password_hash: str) -> bool:
        """Verify a password against its hash"""
        return bcrypt.checkpw(password.encode('utf-8'), password_hash.encode('utf-8'))

    def hash_password(password: str) -> str:
        """Hash a password"""
        return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    def login_required(f):
        """Decorator to require login"""
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash("Please log in to access this page.", "warning")
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated_function

    def role_required(*allowed_roles):
        """Decorator to require specific role"""
        def decorator(f):
            @wraps(f)
            def decorated_function(*args, **kwargs):
                if 'user_id' not in session:
                    flash("Please log in to access this page.", "warning")
                    return redirect(url_for('login'))
                
                user = get_user_by_id(session['user_id'])
                if not user or not user['is_active']:
                    session.clear()
                    flash("Your account has been deactivated.", "danger")
                    return redirect(url_for('login'))
                
                if user['role'] not in allowed_roles:
                    flash("You don't have permission to access this page.", "danger")
                    return redirect(url_for('admin_dashboard'))
                
                return f(*args, **kwargs)
            return decorated_function
        return decorator

    # ── Global write-block for view-only users ──
    # 'user' role is read-only. Block any non-GET requests, except a small
    # whitelist of safe self-service endpoints (logout, password change).
    READONLY_USER_WHITELIST = {
        'logout',
        'login',
        'account_change_password',
        # Assigned branch viewers act as area managers for staff records only.
        'add_staff',
        'import_staff',
        'edit_staff',
        # Viewers remain read-only for business data, but may participate in
        # their own private support conversation.
        'api_send_client_message',
    }

    @app.before_request
    def _enforce_view_only_user():
        # Central guard for legacy routes that predate per-route decorators.
        # Public survey/store pages remain accessible; all admin and internal
        # JSON endpoints require an active application session.
        protected_path = (
            request.path.startswith("/admin/")
            or request.path.startswith("/api/")
            or request.path == "/dashboard/staff-overall"
        )
        api_exempt = request.path == "/api/licensing/users"  # shared-key auth
        if protected_path and not api_exempt and 'user_id' not in session:
            if request.path.startswith("/api/") or request.is_json:
                return jsonify({"success": False, "error": "Authentication required"}), 401
            flash("Please log in to access this page.", "warning")
            return redirect(url_for('login', next=request.full_path))
        if protected_path and not api_exempt and 'user_id' in session:
            current_user = get_user_by_id(session['user_id'])
            if not current_user or not current_user.get('is_active'):
                session.clear()
                if request.path.startswith("/api/") or request.is_json:
                    return jsonify({"success": False, "error": "Session is no longer active"}), 401
                flash("Your session is no longer active. Please log in again.", "warning")
                return redirect(url_for('login'))
            session['role'] = current_user['role']
            if _expired_license_for_user(current_user):
                session.clear()
                flash("License Expired. Please Renew your license.", "danger")
                return redirect(url_for("login"))

        if request.method in ('GET', 'HEAD', 'OPTIONS'):
            return None
        if session.get('role') != 'user':
            return None
        endpoint = request.endpoint or ''
        if endpoint in READONLY_USER_WHITELIST:
            return None
        # Reject everything else for view-only accounts
        if request.is_json or request.headers.get('Accept', '').startswith('application/json'):
            return jsonify({"success": False, "error": "Read-only account"}), 403
        flash("Your account is read-only and cannot make changes.", "warning")
        # Redirect back to referrer if available, else dashboard
        return redirect(request.referrer or url_for('admin_dashboard'))

    # Initialize SMTP email configuration
    email_config = EmailConfig()
    email_config.init_app(app)

    def init_master_schema() -> None:
        retries = 3
        while retries > 0:
            try:
                logger.info(f"Attempting schema initialization... ({retries} retries left)")
                conn = get_db_connection()
                cursor = conn.cursor()
                
                # --- CREATE TABLES IF NOT EXIST ---
                
                # 1. Stores Table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS stores (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        store_name VARCHAR(255) NOT NULL,
                        address TEXT,
                        city VARCHAR(100),
                        province VARCHAR(100),
                        postal_code VARCHAR(20),
                        contact_number VARCHAR(20),
                        email VARCHAR(255),
                        store_manager_name VARCHAR(255),
                        manager_contact VARCHAR(20),
                        store_type VARCHAR(100),
                        operating_hours VARCHAR(255),
                        status ENUM('active', 'inactive', 'pending') DEFAULT 'active',
                        logo_url VARCHAR(500),
                        access_token VARCHAR(100) UNIQUE,
                        subdomain VARCHAR(100) UNIQUE,
                        google_review_url VARCHAR(1000) NULL,
                        reward_type VARCHAR(255) DEFAULT 'Store Reward or Discount',
                        google_review_mode ENUM('review_only', 'reward') DEFAULT 'reward',
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)

                # 2. Staff Table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS staff (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        store_id INT NOT NULL,
                        first_name VARCHAR(100) NOT NULL,
                        last_name VARCHAR(100) NOT NULL,
                        email VARCHAR(255),
                        phone VARCHAR(20),
                        position VARCHAR(100),
                        photo_url LONGTEXT,
                        role ENUM('staff', 'manager', 'supervisor') DEFAULT 'staff',
                        hire_date DATE,
                        status ENUM('active', 'inactive') DEFAULT 'active',
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (store_id) REFERENCES stores(id) ON DELETE CASCADE
                    )
                """)

                # 3. Staff Commendations Table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS staff_commendations (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        response_id INT NOT NULL,
                        staff_id INT NOT NULL,
                        rating INT DEFAULT 5,
                        commendation_type ENUM('excellent_service', 'friendly_attitude', 'professional', 'helpful', 'knowledgeable') DEFAULT 'excellent_service',
                        comment TEXT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY (response_id) REFERENCES responses(id) ON DELETE CASCADE,
                        FOREIGN KEY (staff_id) REFERENCES staff(id) ON DELETE CASCADE
                    )
                """)

                # 4. Questionnaires Table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS questionnaires (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        store_id INT NULL,
                        title VARCHAR(255) NOT NULL,
                        is_active BOOLEAN DEFAULT TRUE,
                        is_template BOOLEAN DEFAULT FALSE,
                        template_id INT NULL,
                        version INT DEFAULT 1,
                        logo_url VARCHAR(500),
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                """)

                # 3. Questions Table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS questions (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        questionnaire_id INT NOT NULL,
                        question_text TEXT NOT NULL,
                        question_type ENUM('rating', 'text', 'multiple_choice') NOT NULL,
                        target_scope ENUM('overall', 'staff', 'manager') DEFAULT 'overall',
                        min_label VARCHAR(255) DEFAULT 'Poor',
                        max_label VARCHAR(255) DEFAULT 'Excellent',
                        allow_comment BOOLEAN DEFAULT FALSE,
                        is_required BOOLEAN DEFAULT TRUE,
                        question_order INT DEFAULT 0,
                        is_active BOOLEAN DEFAULT TRUE,
                        is_template BOOLEAN DEFAULT FALSE,
                        template_id INT NULL
                    )
                """)

                # 4. Question Options Table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS question_options (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        question_id INT NOT NULL,
                        option_text VARCHAR(255) NOT NULL
                    )
                """)

                # 5. Responses Table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS responses (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        questionnaire_id INT NOT NULL,
                        store_id INT NOT NULL,
                        user_email VARCHAR(255),
                        receipt_number VARCHAR(100),
                        submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        status ENUM('unresolved', 'resolved') DEFAULT 'unresolved'
                    )
                """)

                # Track transaction numbers separately so each receipt can only
                # be used once per store, even across browsers or simultaneous requests.
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS receipt_usages (
                        store_id INT NOT NULL,
                        receipt_number VARCHAR(100) NOT NULL,
                        response_id INT NULL,
                        used_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        PRIMARY KEY (store_id, receipt_number),
                        INDEX idx_receipt_usage_response (response_id)
                    )
                """)

                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS global_receipt_usages (
                        receipt_number VARCHAR(100) NOT NULL PRIMARY KEY,
                        store_id INT NOT NULL,
                        response_id INT NULL,
                        used_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        INDEX idx_global_receipt_response (response_id)
                    )
                """)

                # 6. Answers Table
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS answers (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        response_id INT NOT NULL,
                        question_id INT NOT NULL,
                        staff_id INT NULL,
                        answer_text TEXT,
                        rating_value DECIMAL(3,1)
                    )
                """)

                conn.commit()

                # --- UPDATE EXISTING TABLES (MIGRATIONS) ---
                
                # Ensure question_options table exists (fixing crash in master_questionnaire)
                cursor.execute("SHOW TABLES LIKE 'question_options'")
                if not cursor.fetchone():
                    logger.info("Table 'question_options' missing. Creating it now...")
                    cursor.execute("""
                        CREATE TABLE question_options (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            question_id INT NOT NULL,
                            option_text VARCHAR(255) NOT NULL
                        )
                    """)
                    conn.commit()
                
                # Ensure responses table exists
                cursor.execute("SHOW TABLES LIKE 'responses'")
                if not cursor.fetchone():
                    logger.info("Table 'responses' missing. Creating it now...")
                    cursor.execute("""
                        CREATE TABLE responses (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            questionnaire_id INT NOT NULL,
                            store_id INT NOT NULL,
                            user_email VARCHAR(255),
                            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            status ENUM('unresolved', 'resolved') DEFAULT 'unresolved'
                        )
                    """)
                    conn.commit()

                # Ensure answers table exists
                cursor.execute("SHOW TABLES LIKE 'answers'")
                if not cursor.fetchone():
                    logger.info("Table 'answers' missing. Creating it now...")
                    cursor.execute("""
                        CREATE TABLE answers (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            response_id INT NOT NULL,
                            question_id INT NOT NULL,
                            answer_text TEXT,
                            rating_value DECIMAL(3,1)
                        )
                    """)
                    conn.commit()
                
                # Check for responses table columns
                cursor.execute("SHOW COLUMNS FROM responses LIKE 'user_email'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE responses ADD COLUMN user_email VARCHAR(255) AFTER submitted_at")
                
                cursor.execute("SHOW COLUMNS FROM responses LIKE 'status'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE responses ADD COLUMN status ENUM('unresolved', 'resolved') DEFAULT 'unresolved' AFTER user_email")
                
                cursor.execute("SHOW COLUMNS FROM responses LIKE 'is_read'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE responses ADD COLUMN is_read BOOLEAN DEFAULT FALSE AFTER status")
                
                # Ensure system_notifications table exists
                cursor.execute("SHOW TABLES LIKE 'system_notifications'")
                if not cursor.fetchone():
                    logger.info("Table 'system_notifications' missing. Creating it now...")
                    cursor.execute("""
                        CREATE TABLE system_notifications (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            message TEXT NOT NULL,
                            type VARCHAR(50) DEFAULT 'info',
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            is_read BOOLEAN DEFAULT FALSE
                        )
                    """)
                    conn.commit()
                
                # Create audit log table
                cursor.execute("SHOW TABLES LIKE 'audit_logs'")
                if not cursor.fetchone():
                    logger.info("Creating audit_logs table...")
                    cursor.execute("""
                        CREATE TABLE audit_logs (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            entity_type VARCHAR(50) NOT NULL,
                            entity_id INT NOT NULL,
                            action VARCHAR(50) NOT NULL,
                            old_values TEXT,
                            new_values TEXT,
                            user_id VARCHAR(255),
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                        )
                    """)
                    conn.commit()

                # Create client_conversations table for messaging system
                cursor.execute("SHOW TABLES LIKE 'client_conversations'")
                if not cursor.fetchone():
                    logger.info("Creating client_conversations table...")
                    cursor.execute("""
                        CREATE TABLE client_conversations (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            client_identifier VARCHAR(255) NOT NULL UNIQUE,
                            company_name VARCHAR(255),
                            license_key VARCHAR(255),
                            contact_email VARCHAR(255),
                            portal_conversation_id INT NULL,
                            last_message_at TIMESTAMP NULL,
                            last_message_preview TEXT NULL,
                            unread_count INT DEFAULT 0,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                            INDEX idx_client_identifier (client_identifier),
                            INDEX idx_last_message_at (last_message_at)
                        )
                    """)
                    conn.commit()
                else:
                    # Check if portal_conversation_id column exists, add if missing
                    cursor.execute("SHOW COLUMNS FROM client_conversations LIKE 'portal_conversation_id'")
                    if not cursor.fetchone():
                        logger.info("Adding 'portal_conversation_id' column to client_conversations table...")
                        cursor.execute("ALTER TABLE client_conversations ADD COLUMN portal_conversation_id INT NULL AFTER contact_email")
                        conn.commit()

                # Create messages table for individual messages in conversations
                cursor.execute("SHOW TABLES LIKE 'client_messages'")
                if not cursor.fetchone():
                    logger.info("Creating client_messages table...")
                    cursor.execute("""
                        CREATE TABLE client_messages (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            conversation_id INT NOT NULL,
                            sender_type ENUM('client', 'admin') NOT NULL,
                            sender_name VARCHAR(255),
                            message TEXT NOT NULL,
                            is_read BOOLEAN DEFAULT FALSE,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            FOREIGN KEY (conversation_id) REFERENCES client_conversations(id) ON DELETE CASCADE,
                            INDEX idx_conversation_created (conversation_id, created_at)
                        )
                    """)
                    conn.commit()
                
                # Create users table
                cursor.execute("SHOW TABLES LIKE 'users'")
                if not cursor.fetchone():
                    logger.info("Creating users table...")
                    cursor.execute("""
                        CREATE TABLE users (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            username VARCHAR(100) NOT NULL UNIQUE,
                            email VARCHAR(255) NOT NULL UNIQUE,
                            password_hash VARCHAR(255) NOT NULL,
                            role ENUM('superadmin', 'admin', 'user') DEFAULT 'admin',
                            is_active BOOLEAN DEFAULT TRUE,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                        )
                    """)
                    conn.commit()
                    
                    # Create license_config table
                    cursor.execute("SHOW TABLES LIKE 'license_config'")
                    if not cursor.fetchone():
                        logger.info("Creating license_config table...")
                        cursor.execute("""
                            CREATE TABLE license_config (
                                id INT AUTO_INCREMENT PRIMARY KEY,
                                license_key VARCHAR(255) NOT NULL,
                                api_key VARCHAR(255) NOT NULL,
                                licensing_portal_url VARCHAR(255) DEFAULT 'https://feedbacklicensing-production-c938.up.railway.app',
                                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                            )
                        """)
                        conn.commit()
                    
                    # Optional one-time bootstrap. Never ship a default password.
                    bootstrap_username = os.getenv("BOOTSTRAP_ADMIN_USERNAME")
                    bootstrap_email = os.getenv("BOOTSTRAP_ADMIN_EMAIL")
                    bootstrap_password = os.getenv("BOOTSTRAP_ADMIN_PASSWORD")
                    if bootstrap_username and bootstrap_email and bootstrap_password:
                        password_hash = bcrypt.hashpw(bootstrap_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                        cursor.execute(
                            "INSERT INTO users (username, email, password_hash, role) VALUES (%s, %s, %s, %s)",
                            (bootstrap_username, bootstrap_email, password_hash, "superadmin")
                        )
                        conn.commit()
                        logger.info("Created bootstrap superadmin user from environment")
                    else:
                        logger.warning("Users table created without a bootstrap admin; configure BOOTSTRAP_ADMIN_* variables if this is a new installation")
                
                # ----------------------------------------------------------------
                # Role hierarchy migration (legacy -> new):
                #   dev   -> superadmin    (internal/full access)
                #   admin -> superadmin    (legacy admin had full access too)
                #   user  -> admin         ('admin' is now the client role)
                #   user  (new)            view-only (no rows yet)
                # ----------------------------------------------------------------
                try:
                    cursor.execute("SHOW COLUMNS FROM users LIKE 'role'")
                    role_col = cursor.fetchone()
                    role_type = (role_col[1] if role_col else '') or ''
                    needs_migration = 'dev' in role_type.lower()
                    if needs_migration:
                        logger.info("Migrating user roles to new hierarchy (superadmin/admin/user)...")
                        cursor.execute(
                            "ALTER TABLE users MODIFY COLUMN role "
                            "ENUM('dev','superadmin','admin','user') DEFAULT 'user'"
                        )
                        cursor.execute("UPDATE users SET role='superadmin' WHERE role IN ('dev','admin')")
                        cursor.execute("UPDATE users SET role='admin' WHERE role='user'")
                        cursor.execute(
                            "ALTER TABLE users MODIFY COLUMN role "
                            "ENUM('superadmin','admin','user') DEFAULT 'admin'"
                        )
                        conn.commit()
                        logger.info("Role migration complete.")
                    else:
                        # Defensive: even if ENUM is up-to-date, fix any stragglers with role='dev'
                        cursor.execute("SELECT COUNT(*) FROM users WHERE role='dev'")
                        leftover = cursor.fetchone()[0]
                        if leftover:
                            logger.warning(f"Found {leftover} legacy 'dev' user(s); promoting to superadmin.")
                            cursor.execute("UPDATE users SET role='superadmin' WHERE role='dev'")
                            conn.commit()
                except Exception as e:
                    logger.error(f"Role migration error: {e}")

                # Create user_stores link table for per-store view-only assignments
                cursor.execute("SHOW TABLES LIKE 'user_stores'")
                if not cursor.fetchone():
                    logger.info("Creating user_stores table for per-store view-only access...")
                    cursor.execute("""
                        CREATE TABLE user_stores (
                            id INT AUTO_INCREMENT PRIMARY KEY,
                            user_id INT NOT NULL,
                            store_id INT NOT NULL,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                            UNIQUE KEY uniq_user_store (user_id, store_id),
                            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
                            FOREIGN KEY (store_id) REFERENCES stores(id) ON DELETE CASCADE
                        )
                    """)
                    conn.commit()

                # Check for users table columns - add max_stores and license_key for client licensing
                cursor.execute("SHOW COLUMNS FROM users LIKE 'max_stores'")
                if not cursor.fetchone():
                    logger.info("Adding 'max_stores' column to users table...")
                    cursor.execute("ALTER TABLE users ADD COLUMN max_stores INT DEFAULT 0 AFTER role")
                    conn.commit()
                
                cursor.execute("SHOW COLUMNS FROM users LIKE 'license_key'")
                if not cursor.fetchone():
                    logger.info("Adding 'license_key' column to users table...")
                    cursor.execute("ALTER TABLE users ADD COLUMN license_key VARCHAR(255) NULL AFTER max_stores")
                    conn.commit()

                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS review_rewards (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        response_id INT NOT NULL UNIQUE,
                        store_id INT NOT NULL,
                        owner_user_id INT NOT NULL,
                        license_key VARCHAR(255),
                        customer_email VARCHAR(255) NOT NULL,
                        claim_token VARCHAR(100) NOT NULL UNIQUE,
                        reward_code VARCHAR(40) UNIQUE,
                        reward_type VARCHAR(255) NOT NULL,
                        status ENUM('pending','issued','used') DEFAULT 'pending',
                        email_sent BOOLEAN DEFAULT FALSE,
                        email_error TEXT NULL,
                        google_review_proof MEDIUMTEXT NULL,
                        receipt_proof MEDIUMTEXT NULL,
                        review_ocr_text TEXT NULL,
                        receipt_ocr_text TEXT NULL,
                        proof_verified_at TIMESTAMP NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        issued_at TIMESTAMP NULL,
                        used_at TIMESTAMP NULL,
                        FOREIGN KEY (response_id) REFERENCES responses(id) ON DELETE CASCADE,
                        FOREIGN KEY (store_id) REFERENCES stores(id) ON DELETE CASCADE,
                        FOREIGN KEY (owner_user_id) REFERENCES users(id) ON DELETE CASCADE,
                        INDEX idx_reward_tenant (owner_user_id, license_key, status)
                    )
                """)
                conn.commit()

                for column_name, column_definition in (
                    ("google_review_proof", "MEDIUMTEXT NULL"),
                    ("receipt_proof", "MEDIUMTEXT NULL"),
                    ("review_ocr_text", "TEXT NULL"),
                    ("receipt_ocr_text", "TEXT NULL"),
                    ("proof_verified_at", "TIMESTAMP NULL"),
                ):
                    cursor.execute(f"SHOW COLUMNS FROM review_rewards LIKE '{column_name}'")
                    if not cursor.fetchone():
                        cursor.execute(f"ALTER TABLE review_rewards ADD COLUMN {column_name} {column_definition}")
                conn.commit()
                
                # Check for questionnaires table columns
                cursor.execute("SHOW COLUMNS FROM questionnaires LIKE 'is_template'")
                if not cursor.fetchone():
                    logger.info("Adding 'is_template' column to questionnaires table...")
                    cursor.execute("ALTER TABLE questionnaires ADD COLUMN is_template BOOLEAN DEFAULT FALSE AFTER is_active")
                    conn.commit()
                
                cursor.execute("SHOW COLUMNS FROM questionnaires LIKE 'template_id'")
                if not cursor.fetchone():
                    logger.info("Adding 'template_id' column to questionnaires table...")
                    cursor.execute("ALTER TABLE questionnaires ADD COLUMN template_id INT NULL AFTER is_template")
                    conn.commit()
                
                cursor.execute("SHOW COLUMNS FROM questionnaires LIKE 'version'")
                if not cursor.fetchone():
                    logger.info("Adding 'version' column to questionnaires table...")
                    cursor.execute("ALTER TABLE questionnaires ADD COLUMN version INT DEFAULT 1 AFTER template_id")
                    conn.commit()

                cursor.execute("SHOW COLUMNS FROM questionnaires LIKE 'logo_url'")
                if not cursor.fetchone():
                    logger.info("Adding 'logo_url' column to questionnaires table...")
                    cursor.execute("ALTER TABLE questionnaires ADD COLUMN logo_url LONGTEXT AFTER version")
                    conn.commit()
                else:
                    # Always try to update to LONGTEXT to ensure it can handle base64 data
                    try:
                        logger.info("Ensuring 'logo_url' column is LONGTEXT for base64 storage...")
                        cursor.execute("ALTER TABLE questionnaires MODIFY COLUMN logo_url LONGTEXT")
                        conn.commit()
                        logger.info("'logo_url' column updated to LONGTEXT")
                    except Exception as e:
                        logger.info(f"Column may already be LONGTEXT: {e}")

                cursor.execute("SHOW COLUMNS FROM questionnaires LIKE 'owner_user_id'")
                if not cursor.fetchone():
                    logger.info("Adding tenant owner to questionnaires...")
                    cursor.execute("ALTER TABLE questionnaires ADD COLUMN owner_user_id INT NULL AFTER store_id")
                    conn.commit()

                cursor.execute("SHOW COLUMNS FROM questionnaires LIKE 'license_key'")
                if not cursor.fetchone():
                    logger.info("Adding license scope to questionnaires...")
                    cursor.execute("ALTER TABLE questionnaires ADD COLUMN license_key VARCHAR(255) NULL AFTER owner_user_id")
                    conn.commit()

                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS tenant_branding (
                        id INT AUTO_INCREMENT PRIMARY KEY,
                        owner_user_id INT NOT NULL,
                        scope_key VARCHAR(255) NOT NULL UNIQUE,
                        license_key VARCHAR(255) NULL,
                        primary_color VARCHAR(7) NOT NULL DEFAULT '#FF6B35',
                        secondary_color VARCHAR(7) NOT NULL DEFAULT '#F59E0B',
                        accent_color VARCHAR(7) NOT NULL DEFAULT '#2563EB',
                        text_color VARCHAR(7) NOT NULL DEFAULT '#212529',
                        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                        FOREIGN KEY (owner_user_id) REFERENCES users(id) ON DELETE CASCADE
                    )
                """)
                conn.commit()

                # Check for questions table columns
                cursor.execute("SHOW COLUMNS FROM questions LIKE 'is_active'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE questions ADD COLUMN is_active BOOLEAN DEFAULT TRUE AFTER question_order")
                
                cursor.execute("SHOW COLUMNS FROM questions LIKE 'is_template'")
                if not cursor.fetchone():
                    logger.info("Adding 'is_template' column to questions table...")
                    cursor.execute("ALTER TABLE questions ADD COLUMN is_template BOOLEAN DEFAULT FALSE AFTER is_active")
                    conn.commit()

                cursor.execute("SHOW COLUMNS FROM questions LIKE 'template_id'")
                if not cursor.fetchone():
                    logger.info("Adding 'template_id' column to questions table...")
                    cursor.execute("ALTER TABLE questions ADD COLUMN template_id INT NULL AFTER is_template")
                    conn.commit()
                
                cursor.execute("SHOW COLUMNS FROM questions LIKE 'min_label'")
                if not cursor.fetchone():
                    logger.info("Adding 'min_label' column to questions table...")
                    cursor.execute("ALTER TABLE questions ADD COLUMN min_label VARCHAR(255) DEFAULT 'Poor' AFTER question_type")
                    conn.commit()

                cursor.execute("SHOW COLUMNS FROM questions LIKE 'max_label'")
                if not cursor.fetchone():
                    logger.info("Adding 'max_label' column to questions table...")
                    cursor.execute("ALTER TABLE questions ADD COLUMN max_label VARCHAR(255) DEFAULT 'Excellent' AFTER min_label")
                    conn.commit()

                cursor.execute("SHOW COLUMNS FROM questions LIKE 'allow_comment'")
                if not cursor.fetchone():
                    logger.info("Adding 'allow_comment' column to questions table...")
                    cursor.execute("ALTER TABLE questions ADD COLUMN allow_comment BOOLEAN DEFAULT FALSE AFTER max_label")
                    conn.commit()

                cursor.execute("SHOW COLUMNS FROM questions LIKE 'target_scope'")
                if not cursor.fetchone():
                    logger.info("Adding question target scope...")
                    cursor.execute("ALTER TABLE questions ADD COLUMN target_scope ENUM('overall', 'staff', 'manager') DEFAULT 'overall' AFTER question_type")
                    conn.commit()

                cursor.execute("SHOW COLUMNS FROM staff LIKE 'photo_url'")
                if not cursor.fetchone():
                    logger.info("Adding staff profile photo...")
                    cursor.execute("ALTER TABLE staff ADD COLUMN photo_url LONGTEXT AFTER position")
                    conn.commit()

                cursor.execute("SHOW COLUMNS FROM answers LIKE 'staff_id'")
                if not cursor.fetchone():
                    logger.info("Adding selected staff to questionnaire answers...")
                    cursor.execute("ALTER TABLE answers ADD COLUMN staff_id INT NULL AFTER question_id")
                    conn.commit()
                
                # Check for stores table columns
                store_columns = [
                    ("store_manager_name", "VARCHAR(255)"),
                    ("manager_contact", "VARCHAR(20)"),
                    ("store_type", "VARCHAR(100)"),
                    ("operating_hours", "VARCHAR(255)"),
                    ("status", "ENUM('active', 'inactive', 'pending') DEFAULT 'active'"),
                    ("logo_url", "VARCHAR(500)"),
                    ("access_token", "VARCHAR(100) UNIQUE"),
                    ("subdomain", "VARCHAR(100) UNIQUE"),
                    ("user_id", "INT"),
                    ("license_key", "VARCHAR(255)"),
                    ("google_review_url", "VARCHAR(1000) NULL"),
                    ("reward_type", "VARCHAR(255) DEFAULT 'Store Reward or Discount'"),
                    ("google_review_mode", "ENUM('review_only', 'reward') DEFAULT 'reward'")
                ]
                
                for column_name, column_type in store_columns:
                    cursor.execute(f"SHOW COLUMNS FROM stores LIKE '{column_name}'")
                    if not cursor.fetchone():
                        logger.info(f"Adding column {column_name} to stores table...")
                        cursor.execute(f"ALTER TABLE stores ADD COLUMN {column_name} {column_type}")
                        conn.commit()
                        logger.info(f"Column {column_name} added successfully")

                try:
                    cursor.execute("ALTER TABLE stores MODIFY COLUMN logo_url LONGTEXT")
                    conn.commit()
                except Exception as e:
                    logger.info(f"Store logo column may already be LONGTEXT: {e}")

                # Assign user_id to existing stores that don't have it
                cursor.execute("SELECT id FROM stores WHERE user_id IS NULL")
                stores_without_user = cursor.fetchall()
                if stores_without_user:
                    logger.info(f"Assigning user_id to {len(stores_without_user)} existing stores...")
                    # Get the first admin/dev user to assign as owner
                    cursor.execute("SELECT id FROM users WHERE role = 'superadmin' LIMIT 1")
                    admin_user = cursor.fetchone()
                    if admin_user:
                        admin_id = admin_user[0]
                        for store_row in stores_without_user:
                            store_id = store_row[0]
                            cursor.execute("UPDATE stores SET user_id = %s WHERE id = %s", (admin_id, store_id))
                        conn.commit()
                        logger.info(f"Assigned {len(stores_without_user)} stores to user {admin_id}")
                    else:
                        logger.warning("No admin user found to assign existing stores to")

                cursor.execute("""UPDATE stores s
                                  INNER JOIN users u ON u.id = s.user_id
                                  SET s.license_key = u.license_key
                                  WHERE s.license_key IS NULL AND u.license_key IS NOT NULL""")
                conn.commit()

                # Generate access tokens for existing stores that don't have them
                import secrets
                cursor.execute("SELECT id FROM stores WHERE access_token IS NULL OR access_token = ''")
                stores_without_token = cursor.fetchall()
                if stores_without_token:
                    logger.info(f"Generating access tokens for {len(stores_without_token)} existing stores...")
                    for store_row in stores_without_token:
                        store_id = store_row[0]
                        access_token = secrets.token_urlsafe(32)
                        cursor.execute("UPDATE stores SET access_token = %s WHERE id = %s", (access_token, store_id))
                    conn.commit()

                # Generate subdomains for existing stores that don't have them
                cursor.execute("SELECT id, store_name FROM stores WHERE subdomain IS NULL OR subdomain = ''")
                stores_without_subdomain = cursor.fetchall()
                if stores_without_subdomain:
                    logger.info(f"Generating subdomains for {len(stores_without_subdomain)} existing stores...")
                    for store_row in stores_without_subdomain:
                        store_id = store_row[0]
                        store_name = store_row[1]
                        # Generate subdomain from store name (lowercase, alphanumeric, hyphens)
                        import re
                        subdomain = re.sub(r'[^a-zA-Z0-9\s]', '', store_name).lower().replace(' ', '-')
                        subdomain = re.sub(r'-+', '-', subdomain).strip('-')
                        # Ensure uniqueness by adding random suffix if needed
                        cursor.execute("SELECT id FROM stores WHERE subdomain = %s", (subdomain,))
                        if cursor.fetchone():
                            subdomain = f"{subdomain}-{secrets.token_hex(3)}"
                        cursor.execute("UPDATE stores SET subdomain = %s WHERE id = %s", (subdomain, store_id))
                    conn.commit()

                # Check for responses table receipt_number column
                cursor.execute("SHOW COLUMNS FROM responses LIKE 'receipt_number'")
                if not cursor.fetchone():
                    logger.info("Adding 'receipt_number' column to responses table...")
                    cursor.execute("ALTER TABLE responses ADD COLUMN receipt_number VARCHAR(100) AFTER user_email")
                    conn.commit()

                # Reserve receipts already used before one-time enforcement was added.
                cursor.execute("""
                    INSERT IGNORE INTO receipt_usages
                        (store_id, receipt_number, response_id, used_at)
                    SELECT store_id, receipt_number, MIN(id), MIN(submitted_at)
                    FROM responses
                    WHERE receipt_number IS NOT NULL AND receipt_number <> ''
                    GROUP BY store_id, receipt_number
                """)
                conn.commit()

                # Enforce one-time receipt usage across every branch in the system.
                cursor.execute("""
                    INSERT IGNORE INTO global_receipt_usages
                        (receipt_number, store_id, response_id, used_at)
                    SELECT receipt_number, MIN(store_id), MIN(id), MIN(submitted_at)
                    FROM responses
                    WHERE receipt_number IS NOT NULL AND receipt_number <> ''
                    GROUP BY receipt_number
                """)
                conn.commit()
                
                # Add rating column to staff_commendations if missing
                cursor.execute("SHOW COLUMNS FROM staff_commendations LIKE 'rating'")
                if not cursor.fetchone():
                    logger.info("Adding 'rating' column to staff_commendations table...")
                    cursor.execute("ALTER TABLE staff_commendations ADD COLUMN rating INT DEFAULT 5 AFTER staff_id")
                    conn.commit()
                
                conn.commit()
                conn.close()
                logger.info("Master schema check/update completed.")
                break
            except Exception as e:
                logger.error(f"Database initialization error: {e}")
                retries -= 1
                if retries > 0:
                    time.sleep(5)
                else:
                    logger.critical("Could not initialize database schema after multiple attempts.")

    # Always initialize schema on startup
    try:
        init_master_schema()
        logger.info("Schema initialization completed successfully")
    except Exception as e:
        logger.critical(f"CRITICAL: Schema initialization failed: {e}")
        raise

    # --- ERROR HANDLERS ---
    @app.errorhandler(Exception)
    def handle_exception(e):
        """Global error handler to show tracebacks for ANY crash in Railway"""
        error_details = traceback.format_exc()
        logger.error(f"Global Crash: {e}\n{error_details}")
        
        return f"""
        <div style="font-family: sans-serif; padding: 20px; color: #721c24; background: #f8d7da; border: 1px solid #f5c6cb; border-radius: 8px;">
            <h2 style="margin-top: 0;">Oops! Something crashed.</h2>
            <p><b>Error:</b> {e}</p>
            <hr>
            <p><b>Traceback for Debugging:</b></p>
            <pre style="background: #fff; padding: 15px; border-radius: 4px; overflow: auto; font-size: 13px;">{error_details}</pre>
        </div>
        """, 500

    @app.errorhandler(404)
    def not_found_error(error):
        return "404 Not Found", 404

    @app.route("/debug/env")
    @login_required
    def debug_env():
        """Route to see available environment variable keys (NOT values)"""
        user = get_user_by_id(session['user_id'])
        if user['role'] != 'superadmin':
            return jsonify({"error": "Unauthorized"}), 403
        return jsonify({
            "available_keys": list(os.environ.keys()),
            "db_config_host": app.config["DB_CONFIG"].get("host"),
            "db_config_port": app.config["DB_CONFIG"].get("port"),
            "python_version": sys.version
        })

    def get_assigned_store_ids(user_id: int) -> List[int]:
        """Return store ids a view-only user has been granted access to."""
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT store_id FROM user_stores WHERE user_id = %s", (user_id,))
            return [row[0] for row in cursor.fetchall()]
        except Exception as e:
            logger.error(f"Error fetching assigned stores for user {user_id}: {e}")
            return []
        finally:
            conn.close()

    def can_manage_store(user_id: int, store_id: int) -> bool:
        """Superadmins manage all stores; clients manage only stores they own."""
        user = get_user_by_id(user_id)
        if not user or user.get('role') == 'user':
            return False
        if user.get('role') == 'superadmin':
            return True
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT 1 FROM stores WHERE id = %s AND user_id = %s", (store_id, user_id))
            return cursor.fetchone() is not None
        finally:
            conn.close()

    def can_manage_store_staff(user_id: int, store_id: int) -> bool:
        """Allow owners/superadmins, plus viewers assigned to this exact store."""
        user = get_user_by_id(user_id)
        if not user:
            return False
        if user.get('role') in ('admin', 'superadmin'):
            return can_manage_store(user_id, store_id)
        if user.get('role') != 'user':
            return False
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT 1 FROM user_stores WHERE user_id = %s AND store_id = %s LIMIT 1",
                (user_id, store_id),
            )
            return cursor.fetchone() is not None
        finally:
            conn.close()

    def fetch_stores(user_id: int | None = None, assigned_store_ids: List[int] | None = None) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)

            if assigned_store_ids is not None:
                # View-only user: filter by explicit list of assigned store ids
                if not assigned_store_ids:
                    logger.info("Fetching stores for view-only user with no assignments -> []")
                    return []
                placeholders = ",".join(["%s"] * len(assigned_store_ids))
                cursor.execute(
                    f"""
                    SELECT id, store_name, address, city, province, postal_code,
                           contact_number, email, store_manager_name, manager_contact,
                           store_type, status, created_at, logo_url, access_token, subdomain, user_id, license_key,
                           google_review_url, reward_type, google_review_mode
                    FROM stores
                    WHERE id IN ({placeholders})
                    ORDER BY id ASC
                    """,
                    tuple(assigned_store_ids),
                )
            elif user_id:
                # For client users, only show their own stores
                logger.info(f"Fetching stores for user_id: {user_id}")
                cursor.execute(
                    """
                    SELECT id, store_name, address, city, province, postal_code,
                           contact_number, email, store_manager_name, manager_contact,
                           store_type, status, created_at, logo_url, access_token, subdomain, user_id, license_key,
                           google_review_url, reward_type, google_review_mode
                    FROM stores
                    WHERE user_id = %s
                    ORDER BY id ASC
                    """,
                    (user_id,)
                )
            else:
                # For admin/dev/superadmin, show all stores
                logger.info("Fetching all stores (no user_id filter)")
                cursor.execute(
                    """
                    SELECT id, store_name, address, city, province, postal_code,
                           contact_number, email, store_manager_name, manager_contact,
                           store_type, status, created_at, logo_url, access_token, subdomain, user_id, license_key,
                           google_review_url, reward_type, google_review_mode
                    FROM stores
                    ORDER BY id ASC
                    """
                )
            rows = cursor.fetchall()
            logger.info(f"Fetched {len(rows)} stores")
        finally:
            conn.close()

        return rows

    def fetch_store_by_id(store_id: int) -> Dict[str, Any] | None:
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT id, store_name, address, city, province, postal_code,
                       contact_number, email, store_manager_name, manager_contact,
                       store_type, status, created_at, logo_url, access_token, subdomain, user_id, license_key,
                       google_review_url, reward_type, google_review_mode
                FROM stores
                WHERE id = %s
                LIMIT 1
                """,
                (store_id,),
            )
            store = cursor.fetchone()
        finally:
            conn.close()

        return store

    def create_store(
        store_name: str,
        address: str | None = None,
        city: str | None = None,
        province: str | None = None,
        postal_code: str | None = None,
        contact_number: str | None = None,
        email: str | None = None,
        store_manager_name: str | None = None,
        manager_contact: str | None = None,
        store_type: str | None = None,
        status: str = "active",
        logo_url: str | None = None,
        subdomain: str | None = None,
        google_review_url: str | None = None,
        google_review_mode: str = "reward",
        user_id: int | None = None,
        license_key: str | None = None
    ) -> int:
        """Create a new store with validation."""
        # Input validation
        if not store_name or not store_name.strip():
            raise ValueError("Store name is required")
        if status not in ["active", "inactive", "pending"]:
            raise ValueError("Invalid status value")
        if email and "@" not in email:
            raise ValueError("Invalid email format")
        if google_review_mode not in {"review_only", "reward"}:
            google_review_mode = "reward"
        
        import secrets
        import re
        access_token = secrets.token_urlsafe(32)
        
        # Generate subdomain from store name if not provided
        if not subdomain:
            subdomain = re.sub(r'[^a-zA-Z0-9\s]', '', store_name).lower().replace(' ', '-')
            subdomain = re.sub(r'-+', '-', subdomain).strip('-')
        
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                INSERT INTO stores (
                    store_name, address, city, province, postal_code,
                    contact_number, email, store_manager_name, manager_contact,
                    store_type, status, logo_url, access_token, subdomain, google_review_url, google_review_mode, user_id, license_key
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    store_name.strip(), address, city, province, postal_code,
                    contact_number, email, store_manager_name, manager_contact,
                    store_type, status, logo_url, access_token, subdomain, google_review_url, google_review_mode, user_id, license_key
                ),
            )
            new_store_id = int(cursor.lastrowid)

        return new_store_id

    def fetch_questionnaire_by_store(store_id: int) -> Dict[str, Any] | None:
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT id, store_id, owner_user_id, license_key, title, is_active, logo_url, created_at
                FROM questionnaires
                WHERE store_id = %s
                ORDER BY id ASC
                LIMIT 1
                """,
                (store_id,),
            )
            questionnaire = cursor.fetchone()
        finally:
            conn.close()

        return questionnaire

    def fetch_questions_for_questionnaire(questionnaire_id: int) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT id, question_text, question_type, target_scope, min_label, max_label, allow_comment, is_required, question_order
                FROM questions
                WHERE questionnaire_id = %s AND is_active = TRUE
                ORDER BY question_order ASC, id ASC
                """,
                (questionnaire_id,),
            )
            questions = cursor.fetchall()
        finally:
            conn.close()

        return questions

    def fetch_options_for_questions(question_ids: List[int]) -> Dict[int, List[Dict[str, Any]]]:
        if not question_ids:
            return {}

        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            placeholders = ", ".join(["%s"] * len(question_ids))
            cursor.execute(
                f"""
                SELECT question_id, id, option_text
                FROM question_options
                WHERE question_id IN ({placeholders})
                ORDER BY question_id ASC, id ASC
                """,
                tuple(question_ids),
            )
            rows = cursor.fetchall()
        finally:
            conn.close()

        by_question: Dict[int, List[Dict[str, Any]]] = {}
        for row in rows:
            qid = int(row["question_id"])
            by_question.setdefault(qid, []).append({"id": row["id"], "option_text": row["option_text"]})
        return by_question

    def get_store_public_url(store_id: int) -> str:
        """Return a customer-shareable URL, never an internal container IP."""
        public_domain = (os.getenv("RAILWAY_PUBLIC_DOMAIN") or "").strip().rstrip('/')
        static_url = (os.getenv("RAILWAY_STATIC_URL") or "").strip().rstrip('/')
        if public_domain:
            base_url = public_domain if public_domain.startswith(("http://", "https://")) else f"https://{public_domain}"
        elif static_url:
            base_url = static_url if static_url.startswith(("http://", "https://")) else f"https://{static_url}"
        else:
            base_url = request.url_root.rstrip('/')
        return f"{base_url}{url_for('public_survey', store_id=store_id)}"

    def generate_qr_data_uri(text: str) -> str:
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=6,
            border=2,
        )
        qr.add_data(text)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        encoded = base64.b64encode(buf.getvalue()).decode("utf-8")
        return f"data:image/png;base64,{encoded}"

    @app.route("/admin/stores/<int:store_id>/qr-download")
    def download_qr(store_id: int):
        """Download QR code as PNG file."""
        store = fetch_store_by_id(store_id=store_id)
        if not store:
            flash("Store not found", "danger")
            return redirect(url_for("stores_management"))
        public_url = get_store_public_url(store_id=store_id)
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=2,
        )
        qr.add_data(public_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        filename = f"QR_{store['store_name'].replace(' ', '_')}.png"
        return send_file(buf, mimetype="image/png", as_attachment=True, download_name=filename)

    # -------------------------
    # REPORT GENERATION (CSV & PDF)
    # -------------------------
    def _get_report_data(store_id: int, month: str = None):
        """Gather feedback data for reports, optionally filtered by month (YYYY-MM)."""
        store = fetch_store_by_id(store_id=store_id)
        if not store:
            return None, None, None, None, None
        all_feedback = fetch_responses_for_store(store_id=store_id, limit=10000)

        # Filter by month if provided
        if month:
            filtered = []
            for fb in all_feedback:
                submitted = fb.get("submitted_at")
                if submitted:
                    if isinstance(submitted, str):
                        try:
                            submitted = datetime.strptime(submitted, "%Y-%m-%d %H:%M:%S")
                        except (ValueError, TypeError):
                            continue
                    if submitted.strftime("%Y-%m") == month:
                        filtered.append(fb)
            all_feedback = filtered

        response_ids = [int(r["id"]) for r in all_feedback]
        answers_map = fetch_answers_for_responses(response_ids) if response_ids else {}

        # Get commendations
        commendations_map = {}
        if response_ids:
            conn = get_db_connection()
            try:
                cursor = conn.cursor(dictionary=True)
                ph = ','.join(['%s'] * len(response_ids))
                cursor.execute(f"""
                    SELECT sc.response_id, s.first_name, s.last_name, s.position
                    FROM staff_commendations sc
                    JOIN staff s ON s.id = sc.staff_id
                    WHERE sc.response_id IN ({ph})
                """, response_ids)
                for row in cursor.fetchall():
                    commendations_map.setdefault(int(row["response_id"]), []).append(row)
            finally:
                conn.close()

        return store, all_feedback, answers_map, commendations_map, response_ids

    @app.route("/admin/stores/<int:store_id>/report/csv")
    def download_report_csv(store_id: int):
        """Download feedback data as CSV."""
        month = request.args.get("month", "")
        store, feedback_list, answers_map, commendations_map, _ = _get_report_data(store_id, month or None)
        if not store:
            flash("Store not found", "danger")
            return redirect(url_for("stores_management"))

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Feedback ID", "Date", "Email", "Receipt #", "Status",
            "Question", "Type", "Rating", "Answer",
            "Commended Staff"
        ])

        for fb in feedback_list:
            fb_id = int(fb["id"])
            submitted = fb.get("submitted_at", "")
            if hasattr(submitted, "strftime"):
                submitted = submitted.strftime("%Y-%m-%d %H:%M:%S")
            email = fb.get("user_email", "")
            receipt = fb.get("receipt_number", "")
            status = fb.get("status", "")
            answers = answers_map.get(fb_id, [])
            comms = commendations_map.get(fb_id, [])
            comm_names = ", ".join(f"{c['first_name']} {c['last_name']}" for c in comms)

            if answers:
                for ans in answers:
                    writer.writerow([
                        fb_id, submitted, email, receipt, status,
                        ans.get("question_text", ""),
                        ans.get("question_type", ""),
                        ans.get("rating_value", ""),
                        ans.get("answer_text", ""),
                        comm_names
                    ])
            else:
                writer.writerow([fb_id, submitted, email, receipt, status, "", "", "", "", comm_names])

        buf = io.BytesIO()
        buf.write(output.getvalue().encode("utf-8"))
        buf.seek(0)
        month_label = month if month else "all"
        filename = f"Report_{store['store_name'].replace(' ', '_')}_{month_label}.csv"
        return send_file(buf, mimetype="text/csv", as_attachment=True, download_name=filename)

    @app.route("/admin/stores/<int:store_id>/report/pdf")
    def download_report_pdf(store_id: int):
        """Download feedback report as PDF."""
        month = request.args.get("month", "")
        store, feedback_list, answers_map, commendations_map, _ = _get_report_data(store_id, month or None)
        if not store:
            flash("Store not found", "danger")
            return redirect(url_for("stores_management"))

        total = len(feedback_list)
        resolved = sum(1 for f in feedback_list if f.get("status") == "resolved")
        unresolved = total - resolved
        resolution_rate = round(resolved / total * 100, 1) if total > 0 else 0

        # Rating stats
        rating_dist = [0, 0, 0, 0, 0]
        total_ratings = 0
        for fb in feedback_list:
            for ans in answers_map.get(int(fb["id"]), []):
                rv = ans.get("rating_value")
                if rv:
                    r = int(float(rv))
                    if 1 <= r <= 5:
                        rating_dist[r - 1] += 1
                        total_ratings += 1
        avg_rating = round(sum((i + 1) * c for i, c in enumerate(rating_dist)) / total_ratings, 2) if total_ratings > 0 else 0

        # Top commended staff
        staff_counts = defaultdict(int)
        for fb in feedback_list:
            for c in commendations_map.get(int(fb["id"]), []):
                staff_counts[f"{c['first_name']} {c['last_name']}"] += 1
        top_staff = sorted(staff_counts.items(), key=lambda x: x[1], reverse=True)[:10]

        # Build PDF
        month_label = month if month else "All Time"
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Title
        pdf.set_font("Helvetica", "B", 18)
        pdf.cell(0, 12, f"{store['store_name']} - Feedback Report", ln=True, align="C")
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 8, f"Period: {month_label}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True, align="C")
        pdf.ln(8)

        # KPI Summary
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 8, "Summary", ln=True)
        pdf.set_font("Helvetica", "", 10)
        col_w = 47.5
        pdf.set_fill_color(240, 240, 240)
        for label, val in [("Total Feedback", total), ("Resolved", resolved), ("Unresolved", unresolved), ("Resolution Rate", f"{resolution_rate}%")]:
            pdf.cell(col_w, 18, f"{label}\n{val}", border=1, align="C", fill=True)
        pdf.ln(18)
        pdf.ln(4)

        # Rating summary
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 8, "Ratings", ln=True)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(95, 8, f"Average Rating: {avg_rating} / 5", ln=False)
        pdf.cell(95, 8, f"Total Ratings: {total_ratings}", ln=True)
        # Rating distribution table
        pdf.set_font("Helvetica", "B", 9)
        for i in range(5):
            pdf.cell(38, 7, f"{i+1} Star", border=1, align="C", fill=True)
        pdf.ln(7)
        pdf.set_font("Helvetica", "", 9)
        for i in range(5):
            pct = round(rating_dist[i] / total_ratings * 100, 1) if total_ratings > 0 else 0
            pdf.cell(38, 7, f"{rating_dist[i]} ({pct}%)", border=1, align="C")
        pdf.ln(7)
        pdf.ln(6)

        # Top Commended Staff
        if top_staff:
            pdf.set_font("Helvetica", "B", 13)
            pdf.cell(0, 8, "Top Commended Staff", ln=True)
            pdf.set_font("Helvetica", "B", 9)
            pdf.set_fill_color(230, 230, 230)
            pdf.cell(10, 7, "#", border=1, align="C", fill=True)
            pdf.cell(100, 7, "Staff Member", border=1, fill=True)
            pdf.cell(40, 7, "Commendations", border=1, align="C", fill=True)
            pdf.ln(7)
            pdf.set_font("Helvetica", "", 9)
            for idx, (name, count) in enumerate(top_staff, 1):
                pdf.cell(10, 7, str(idx), border=1, align="C")
                pdf.cell(100, 7, name, border=1)
                pdf.cell(40, 7, str(count), border=1, align="C")
                pdf.ln(7)
            pdf.ln(6)

        # Feedback Detail Table
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 8, "Feedback Details", ln=True)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(230, 230, 230)
        col_widths = [15, 30, 45, 25, 20, 55]
        headers = ["ID", "Date", "Email", "Receipt #", "Status", "Avg Rating"]
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 7, h, border=1, align="C", fill=True)
        pdf.ln(7)

        pdf.set_font("Helvetica", "", 7)
        for fb in feedback_list:
            fb_id = int(fb["id"])
            submitted = fb.get("submitted_at", "")
            if hasattr(submitted, "strftime"):
                submitted = submitted.strftime("%m/%d/%y")
            elif isinstance(submitted, str) and len(submitted) > 10:
                submitted = submitted[:10]
            email = (fb.get("user_email", "") or "")[:25]
            receipt = (fb.get("receipt_number", "") or "")[:15]
            status = fb.get("status", "")
            answers = answers_map.get(fb_id, [])
            ratings = [float(a["rating_value"]) for a in answers if a.get("rating_value")]
            avg_r = round(sum(ratings) / len(ratings), 1) if ratings else "N/A"

            pdf.cell(col_widths[0], 6, str(fb_id), border=1, align="C")
            pdf.cell(col_widths[1], 6, str(submitted), border=1, align="C")
            pdf.cell(col_widths[2], 6, email, border=1)
            pdf.cell(col_widths[3], 6, receipt, border=1, align="C")
            pdf.cell(col_widths[4], 6, status, border=1, align="C")
            pdf.cell(col_widths[5], 6, str(avg_r), border=1, align="C")
            pdf.ln(6)

        buf = io.BytesIO()
        pdf.output(buf)
        buf.seek(0)
        filename = f"Report_{store['store_name'].replace(' ', '_')}_{month_label.replace(' ', '_')}.pdf"
        return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)

    # -------------------------
    # TEMPLATE QUESTIONNAIRE CRUD
    # -------------------------
    def _tenant_owner(user_id: int | None = None) -> Dict[str, Any]:
        uid = int(user_id or session['user_id'])
        user = get_user_by_id(uid)
        if not user:
            raise ValueError("Tenant owner not found")
        return user

    def fetch_template_questionnaire(owner_user_id: int | None = None, license_key: str | None = None) -> Dict[str, Any] | None:
        owner = _tenant_owner(owner_user_id)
        effective_license = license_key if license_key is not None else owner.get('license_key')
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT id, title, is_active, version, created_at, updated_at, logo_url,
                       owner_user_id, license_key
                FROM questionnaires
                WHERE is_template = TRUE AND owner_user_id = %s AND license_key <=> %s
                ORDER BY id ASC
                LIMIT 1
                """,
                (owner['id'], effective_license),
            )
            row = cursor.fetchone()
        finally:
            conn.close()
        return row

    def ensure_template_questionnaire(owner_user_id: int | None = None) -> Dict[str, Any]:
        owner = _tenant_owner(owner_user_id)
        existing = fetch_template_questionnaire(int(owner['id']))
        if existing:
            return existing
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            # An admin receives a detached snapshot of the Dev/Superadmin
            # starter questionnaire. The copied rows never point back to the
            # Dev questionnaire, so client edits stay inside their own license.
            starter = None
            if owner.get('role') == 'admin':
                cursor.execute(
                    """SELECT q.id, q.title, q.is_active, q.version, q.logo_url
                       FROM questionnaires q
                       JOIN users u ON u.id = q.owner_user_id
                       WHERE q.is_template = TRUE AND u.role = 'superadmin'
                       ORDER BY q.id ASC LIMIT 1"""
                )
                starter = cursor.fetchone()
            if not starter:
                cursor.execute("""SELECT id, title, is_active, version, logo_url
                                  FROM questionnaires
                                  WHERE is_template = TRUE AND owner_user_id IS NULL
                                  ORDER BY id ASC LIMIT 1""")
                starter = cursor.fetchone()
            default_title = starter[1] if starter else "Customer Feedback"
            default_active = bool(starter[2]) if starter else True
            default_version = int(starter[3] or 1) if starter else 1
            default_logo = starter[4] if starter else None
            cursor.execute(
                """
                INSERT INTO questionnaires
                    (store_id, owner_user_id, license_key, title, is_active, is_template, version, logo_url)
                VALUES (NULL, %s, %s, %s, %s, %s, %s, %s)
                """,
                (owner['id'], owner.get('license_key'), default_title, default_active, True, default_version, default_logo),
            )
            template_id = int(cursor.lastrowid)
            copied_count = 0
            if starter:
                cursor.execute(
                    """SELECT id, question_text, question_type, target_scope,
                              min_label, max_label, allow_comment, is_required,
                              question_order, is_active
                       FROM questions
                       WHERE questionnaire_id = %s AND is_template = TRUE
                       ORDER BY question_order ASC, id ASC LIMIT 5""",
                    (int(starter[0]),),
                )
                for source_question in cursor.fetchall():
                    cursor.execute(
                        """INSERT INTO questions
                           (questionnaire_id, question_text, question_type, target_scope,
                            min_label, max_label, allow_comment, is_required,
                            question_order, is_active, is_template, template_id)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE, NULL)""",
                        (template_id, source_question[1], source_question[2], source_question[3] or 'overall',
                         source_question[4], source_question[5], source_question[6], source_question[7],
                         source_question[8], source_question[9]),
                    )
                    copied_question_id = int(cursor.lastrowid)
                    cursor.execute(
                        "SELECT option_text FROM question_options WHERE question_id = %s ORDER BY id ASC",
                        (int(source_question[0]),),
                    )
                    source_options = cursor.fetchall()
                    if source_options:
                        cursor.executemany(
                            "INSERT INTO question_options (question_id, option_text, is_template) VALUES (%s, %s, TRUE)",
                            [(copied_question_id, option[0]) for option in source_options],
                        )
                    copied_count += 1

            fallback_questions = [
                ("How would you rate your overall experience and service?", "overall"),
                ("How would you rate the manager's service?", "manager"),
                ("How would you rate the staff member who assisted you?", "staff"),
                ("How satisfied are you with the quality you received?", "overall"),
                ("How likely are you to recommend this store?", "overall"),
            ]
            if copied_count < 5:
                cursor.executemany(
                    """INSERT INTO questions
                       (questionnaire_id, question_text, question_type, target_scope,
                        min_label, max_label, allow_comment, is_required,
                        question_order, is_active, is_template, template_id)
                       VALUES (%s, %s, 'rating', %s, 'Poor', 'Excellent', TRUE,
                               TRUE, %s, TRUE, TRUE, NULL)""",
                    [(template_id, text, scope, order)
                     for order, (text, scope) in enumerate(fallback_questions, 1)
                     if order > copied_count],
                )
            # Existing tenant questionnaires are preserved and never cleared.
        return {"id": template_id, "title": default_title, "is_active": default_active,
                "version": default_version, "created_at": None, "is_template": True,
                "owner_user_id": owner['id'], "license_key": owner.get('license_key'), "logo_url": default_logo}

    def fetch_tenant_branding(owner_user_id: int, license_key: str | None = None) -> Dict[str, Any]:
        defaults = {"primary_color": "#FF6B35", "secondary_color": "#F59E0B",
                    "accent_color": "#2563EB", "text_color": "#212529"}
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            owner = _tenant_owner(owner_user_id)
            effective_license = license_key if license_key is not None else owner.get('license_key')
            scope_key = effective_license or f"user:{owner_user_id}"
            cursor.execute("SELECT primary_color, secondary_color, accent_color, text_color FROM tenant_branding WHERE scope_key = %s", (scope_key,))
            return cursor.fetchone() or defaults
        finally:
            conn.close()

    @app.context_processor
    def inject_tenant_ui_branding():
        """Apply company colors to every authenticated client/admin page."""
        defaults = {"primary_color": "#FF6B35", "secondary_color": "#B03A14",
                    "accent_color": "#2563EB", "text_color": "#212529"}
        user_id = session.get("user_id")
        if not user_id:
            return {"ui_branding": defaults}
        try:
            user = get_user_by_id(int(user_id))
            if not user or user.get("role") == "superadmin":
                return {"ui_branding": defaults}
            if user.get("role") == "admin":
                return {"ui_branding": fetch_tenant_branding(int(user["id"]), user.get("license_key"))}
            conn = get_db_connection()
            try:
                cursor = conn.cursor(dictionary=True)
                cursor.execute(
                    """SELECT s.user_id, s.license_key FROM user_stores us
                       JOIN stores s ON s.id = us.store_id
                       WHERE us.user_id = %s ORDER BY s.id ASC LIMIT 1""",
                    (int(user["id"]),),
                )
                store_owner = cursor.fetchone()
            finally:
                conn.close()
            if store_owner:
                return {"ui_branding": fetch_tenant_branding(int(store_owner["user_id"]), store_owner.get("license_key"))}
        except Exception as exc:
            logger.warning("Unable to load UI branding: %s", exc)
        return {"ui_branding": defaults}

    def get_questionnaire_license_limit(owner: Dict[str, Any]) -> int:
        """Return licensed *additional* questions; zero means no extra questions."""
        if owner.get('role') != 'admin':
            return 0
        license_key = owner.get('license_key')
        if not license_key:
            raise ValueError("Configure a valid license key before publishing questionnaires.")
        config = get_license_config()
        portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
        try:
            import requests as http_requests
            response = http_requests.post(
                f"{portal_url}/api/validate/{license_key}",
                headers=licensing_api_headers(), timeout=10,
            )
            data = response.json() if response.content else {}
            if response.status_code != 200 or not data.get('valid'):
                raise ValueError(data.get('message') or "The license is invalid or inactive.")
            return max(0, int(data.get('max_questionnaires') or 0))
        except ValueError:
            raise
        except Exception as exc:
            logger.error("Unable to validate questionnaire limit: %s", exc)
            raise ValueError("Unable to validate the questionnaire license limit. Please try again.")

    def enforce_questionnaire_limit(owner: Dict[str, Any], _candidate_store_ids: List[int]) -> None:
        """Keep the five starter questions plus the licensed additional allowance."""
        max_questionnaires = get_questionnaire_license_limit(owner)
        allowed_total = 5 + max_questionnaires
        template = fetch_template_questionnaire(int(owner['id']), owner.get('license_key'))
        if not template:
            return
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                """SELECT COUNT(*) FROM questions
                   WHERE questionnaire_id = %s AND is_template = TRUE""",
                (int(template['id']),),
            )
            current_count = int(cursor.fetchone()[0])
        finally:
            conn.close()
        if current_count > allowed_total:
            raise ValueError(
                f"Question limit exceeded. Your plan includes 5 starter questions plus "
                f"{max_questionnaires} additional question(s) ({allowed_total} total)."
            )

    def questionnaire_quota_status(owner: Dict[str, Any]) -> Dict[str, Any]:
        additional_limit = get_questionnaire_license_limit(owner)
        allowed_total = 5 + additional_limit
        template = fetch_template_questionnaire(int(owner['id']), owner.get('license_key'))
        if not template:
            return {"used": 0, "max": allowed_total, "base": 5,
                    "additional": additional_limit, "remaining": allowed_total}
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                """SELECT COUNT(*) FROM questions
                   WHERE questionnaire_id = %s AND is_template = TRUE""",
                (int(template['id']),),
            )
            used = int(cursor.fetchone()[0])
        finally:
            conn.close()
        return {"used": used, "max": allowed_total, "base": 5,
                "additional": additional_limit, "remaining": max(0, allowed_total - used)}

    def update_template_questionnaire(title: str, is_active: bool, updated_at: str | None = None) -> None:
        """Update template questionnaire with validation."""
        # Input validation
        if not title or not title.strip():
            raise ValueError("Title is required")
        
        template = ensure_template_questionnaire()
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            if updated_at:
                cursor.execute(
                    """
                    UPDATE questionnaires
                    SET title = %s, is_active = %s, updated_at = %s
                    WHERE id = %s
                    """,
                    (title.strip(), is_active, updated_at, int(template["id"])),
                )
            else:
                cursor.execute(
                    """
                    UPDATE questionnaires
                    SET title = %s, is_active = %s, updated_at = NOW()
                    WHERE id = %s
                    """,
                    (title.strip(), is_active, int(template["id"])),
                )

    def fetch_template_questions(template_questionnaire_id: int) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT id, question_text, question_type, target_scope, min_label, max_label, allow_comment, is_required, question_order
                FROM questions
                WHERE questionnaire_id = %s
                ORDER BY question_order ASC, id ASC
                """,
                (template_questionnaire_id,),
            )
            rows = cursor.fetchall()
        finally:
            conn.close()
        return rows

    def fetch_template_options_by_question(template_question_ids: List[int]) -> Dict[int, List[Dict[str, Any]]]:
        if not template_question_ids:
            return {}
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            placeholders = ", ".join(["%s"] * len(template_question_ids))
            cursor.execute(
                f"""
                SELECT question_id, id, option_text
                FROM question_options
                WHERE question_id IN ({placeholders})
                ORDER BY question_id ASC, id ASC
                """,
                tuple(template_question_ids),
            )
            rows = cursor.fetchall()
        finally:
            conn.close()
        by_q: Dict[int, List[Dict[str, Any]]] = {}
        for r in rows:
            qid = int(r["question_id"])
            by_q.setdefault(qid, []).append({"id": r["id"], "option_text": r["option_text"]})
        return by_q

    def add_template_question(
        template_questionnaire_id: int,
        question_text: str,
        question_type: str,
        is_required: bool,
        question_order: int,
        min_label: str = "Poor",
        max_label: str = "Excellent",
        allow_comment: bool = False,
        target_scope: str = "overall",
    ) -> int:
        """Add a template question with validation."""
        # Input validation
        if not question_text or not question_text.strip():
            raise ValueError("Question text is required")
        if question_type not in ["rating", "text", "multiple_choice"]:
            raise ValueError("Invalid question type")
        if target_scope not in ["overall", "staff", "manager"]:
            raise ValueError("Invalid question target")
        if question_order < 0:
            raise ValueError("Question order must be non-negative")
        
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                INSERT INTO questions
                (questionnaire_id, question_text, question_type, target_scope, min_label, max_label, allow_comment, is_required, question_order, is_template)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (template_questionnaire_id, question_text.strip(), question_type, target_scope, min_label, max_label, allow_comment, is_required, question_order, True),
            )
            return int(cursor.lastrowid)

    def delete_template_question(template_question_id: int) -> None:
        """Delete a template question by ID."""
        template = ensure_template_questionnaire()
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM questions WHERE id = %s AND questionnaire_id = %s AND is_template = TRUE", (template_question_id, template['id']))

    def update_template_question(question_id: int, question_text: str, question_type: str, is_required: bool, min_label: str = "Poor", max_label: str = "Excellent", allow_comment: bool = False, target_scope: str = "overall") -> None:
        """Update a template question with validation."""
        # Input validation
        if not question_text or not question_text.strip():
            raise ValueError("Question text is required")
        if question_type not in ["rating", "text", "multiple_choice"]:
            raise ValueError("Invalid question type")
        if target_scope not in ["overall", "staff", "manager"]:
            raise ValueError("Invalid question target")
        
        template = ensure_template_questionnaire()
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                UPDATE questions
                SET question_text = %s, question_type = %s, target_scope = %s, is_required = %s, min_label = %s, max_label = %s, allow_comment = %s
                WHERE id = %s AND questionnaire_id = %s AND is_template = TRUE
                """,
                (question_text.strip(), question_type, target_scope, is_required, min_label, max_label, allow_comment, question_id, template['id']),
            )

    def add_template_option(template_question_id: int, option_text: str) -> int:
        """Add a template option with validation."""
        # Input validation
        if not option_text or not option_text.strip():
            raise ValueError("Option text is required")
        
        template = ensure_template_questionnaire()
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT 1 FROM questions WHERE id = %s AND questionnaire_id = %s AND is_template = TRUE", (template_question_id, template['id']))
            if not cursor.fetchone():
                raise ValueError("Question does not belong to this license")
            cursor.execute(
                """
                INSERT INTO question_options (question_id, option_text)
                VALUES (%s, %s)
                """,
                (template_question_id, option_text.strip()),
            )
            return int(cursor.lastrowid)

    def delete_template_option(template_option_id: int) -> None:
        """Delete a template option by ID."""
        template = ensure_template_questionnaire()
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            cursor.execute("""DELETE qo FROM question_options qo
                              INNER JOIN questions q ON q.id = qo.question_id
                              WHERE qo.id = %s AND q.questionnaire_id = %s AND q.is_template = TRUE""",
                           (template_option_id, template['id']))

    def publish_template_to_all_stores() -> int:
        """Publish only inside the signed-in admin's tenant/license scope."""
        owner = _tenant_owner()
        template = ensure_template_questionnaire()
        template_id = int(template["id"])
        template_questions = fetch_template_questions(template_questionnaire_id=template_id)
        template_options_by_question_id = fetch_template_options_by_question([int(q["id"]) for q in template_questions])

        stores = [store for store in fetch_stores(user_id=int(owner['id']))
                  if (store.get('license_key') or None) == (owner.get('license_key') or None)]
        enforce_questionnaire_limit(owner, [int(store['id']) for store in stores])
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor(dictionary=True)
            published_count = 0

            for store in stores:
                store_id = int(store["id"])

                # Check if store already has a questionnaire
                cursor.execute("SELECT id FROM questionnaires WHERE store_id = %s AND is_template = FALSE ORDER BY id ASC LIMIT 1", (store_id,))
                existing = cursor.fetchone()
                
                if existing:
                    # Update existing questionnaire metadata without deleting it
                    cursor.execute(
                        """
                        UPDATE questionnaires
                        SET title = %s, is_active = %s, template_id = %s,
                            owner_user_id = %s, license_key = %s, logo_url = %s
                        WHERE id = %s
                        """,
                        (template["title"], bool(template["is_active"]), template_id,
                         owner['id'], owner.get('license_key'), store.get('logo_url') or template.get('logo_url'), int(existing["id"])),
                    )
                    questionnaire_id = int(existing["id"])
                    
                    # Deactivate existing questions instead of deleting them
                    cursor.execute(
                        """
                        UPDATE questions
                        SET is_active = FALSE
                        WHERE questionnaire_id = %s AND is_template = FALSE
                        """,
                        (questionnaire_id,),
                    )
                else:
                    # Create new store questionnaire
                    cursor.execute(
                        """
                        INSERT INTO questionnaires
                            (store_id, owner_user_id, license_key, title, is_active, is_template, template_id, logo_url)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (store_id, owner['id'], owner.get('license_key'), template["title"],
                         bool(template["is_active"]), False, template_id, store.get('logo_url') or template.get('logo_url')),
                    )
                    questionnaire_id = int(cursor.lastrowid)

                # Add new active questions from template
                question_id_map: Dict[int, int] = {}
                for tq in template_questions:
                    cursor.execute(
                        """
                        INSERT INTO questions
                        (questionnaire_id, question_text, question_type, target_scope, min_label, max_label, allow_comment, is_required, question_order, is_template, template_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        (
                            questionnaire_id,
                            tq["question_text"],
                            tq["question_type"],
                            tq.get("target_scope", "overall"),
                            tq.get("min_label", "Poor"),
                            tq.get("max_label", "Excellent"),
                            bool(tq.get("allow_comment", False)),
                            bool(tq["is_required"]),
                            int(tq["question_order"]),
                            False,  # Store questions are not templates
                            int(tq["id"]),  # Link to template question
                        ),
                    )
                    new_qid = int(cursor.lastrowid)
                    question_id_map[int(tq["id"])] = new_qid

                for old_tq_id, opts in template_options_by_question_id.items():
                    new_qid = question_id_map.get(int(old_tq_id))
                    if not new_qid:
                        continue
                    for opt in opts:
                        cursor.execute(
                            """
                            INSERT INTO question_options (question_id, option_text)
                            VALUES (%s, %s)
                            """,
                            (new_qid, opt["option_text"]),
                        )

                published_count += 1

            return published_count

    @app.route("/")
    def index():
        if 'user_id' in session:
            return redirect(url_for("admin_dashboard"))
        return redirect(url_for("login"))

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if request.method == "POST":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            
            if not username or not password:
                flash("Username and password are required.", "danger")
                return redirect(url_for("login"))
            
            user = get_user_by_username(username)
            if not user:
                flash("Invalid username or password.", "danger")
                return redirect(url_for("login"))
            
            if not user['is_active']:
                flash("Your account has been deactivated.", "danger")
                return redirect(url_for("login"))
            
            if verify_password(password, user['password_hash']):
                if _expired_license_for_user(user, force=True):
                    flash("License Expired. Please Renew your license.", "danger")
                    return redirect(url_for("login"))
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['role'] = user['role']
                flash(f"Welcome back, {user['username']}!", "success")
                return redirect(url_for("admin_dashboard"))
            else:
                flash("Invalid username or password.", "danger")
                return redirect(url_for("login"))
        
        return render_template("auth/login.html")

    @app.route("/logout")
    def logout():
        session.clear()
        flash("You have been logged out.", "info")
        return redirect(url_for("login"))

    @app.route("/api/license/refresh", methods=["GET"])
    @login_required
    def api_refresh_license_status():
        """Bypass the hourly cache so a completed renewal clears immediately."""
        user = get_user_by_id(session["user_id"])
        if not user or user.get("role") != "admin" or not user.get("license_key"):
            return jsonify({"success": False, "error": "No client license connected"}), 403
        status = validate_tenant_license(user["license_key"], force=True)
        expiry = _parse_license_expiry(status)
        seconds_left = int((expiry - datetime.now(timezone.utc)).total_seconds()) if expiry else None
        return jsonify({
            "success": True,
            "expired": _license_is_expired(status),
            "show_warning": seconds_left is not None and 0 < seconds_left <= 30 * 24 * 60 * 60,
            "expires_at": expiry.isoformat() if expiry else None,
            "expiry_date": expiry.strftime("%B %d, %Y") if expiry else None,
        })

    @app.route("/admin/questionnaire", methods=["GET", "POST"])
    @login_required
    def master_questionnaire():
        user = get_user_by_id(session['user_id'])
        if not user or user.get('role') not in ('admin', 'superadmin'):
            flash("You don't have permission to manage questionnaires.", "danger")
            return redirect(url_for("admin_dashboard"))
        if request.method == "POST":
            title = request.form.get("title", "").strip()
            updated_at = request.form.get("updated_at", "").strip()
            if not title:
                flash("Template questionnaire title is required.", "danger")
                return redirect(url_for("master_questionnaire"))
            
            template = ensure_template_questionnaire()
            old_title = template.get("title", "")
            
            update_template_questionnaire(title=title, updated_at=updated_at if updated_at else None)
            
            # Log questionnaire changes
            changes = []
            if old_title != title:
                changes.append(f"Title: {old_title} → {title}")
            
            if changes:
                log_audit(
                    entity_type="questionnaire",
                    entity_id=int(template["id"]),
                    action="updated",
                    old_values=f"{', '.join(changes)}"
                )
            
            flash("Questionnaire Saved Successfully", "success")
            return redirect(url_for("master_questionnaire"))

        # Single database connection for better performance
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            
            # One master template per admin/license tenant.
            template = ensure_template_questionnaire()
            
            # Initialize questions and options
            questions = []
            options_by_question_id = {}
            
            if template:
                template_id = int(template["id"])
                
                # Get questions with single query
                cursor.execute(
                    """
                    SELECT q.id, q.question_text, q.question_type, q.target_scope, q.min_label, q.max_label,
                           q.allow_comment, q.is_required, q.question_order,
                           qo.id as option_id, qo.option_text
                    FROM questions q
                    LEFT JOIN question_options qo ON q.id = qo.question_id
                    WHERE q.questionnaire_id = %s
                    ORDER BY q.question_order ASC, q.id ASC, qo.id ASC
                    """,
                    (template_id,),
                )
                rows = cursor.fetchall()
                
                # Organize questions and options
                current_question = None
                
                for row in rows:
                    qid = int(row["id"])
                    
                    # Create question if not exists
                    if qid not in [q.get("id") for q in questions]:
                        questions.append({
                            "id": qid,
                            "question_text": row["question_text"],
                            "question_type": row["question_type"],
                            "target_scope": row["target_scope"] or "overall",
                            "min_label": row["min_label"],
                            "max_label": row["max_label"],
                            "allow_comment": bool(row["allow_comment"]),
                            "is_required": bool(row["is_required"]),
                            "question_order": int(row["question_order"])
                        })
                        options_by_question_id[qid] = []
                    
                    # Add option if exists
                    if row["option_id"]:
                        options_by_question_id[qid].append({
                            "id": row["option_id"],
                            "option_text": row["option_text"]
                        })
                        
        finally:
            conn.close()

        try:
            questionnaire_quota = questionnaire_quota_status(user)
            questionnaire_quota_error = None
        except ValueError as exc:
            questionnaire_quota = None
            questionnaire_quota_error = str(exc)

        return render_template(
            "master_questionnaire/master_questionnaire.html",
            master=template,
            questions=questions,
            options_by_question_id=options_by_question_id,
            branding=fetch_tenant_branding(int(user['id'])),
            questionnaire_quota=questionnaire_quota,
            questionnaire_quota_error=questionnaire_quota_error,
        )

    @app.route("/admin/questionnaire/questions/add", methods=["POST"])
    def master_add_question():
        template = ensure_template_questionnaire()
        template_id = int(template["id"])

        try:
            quota = questionnaire_quota_status(_tenant_owner())
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("master_questionnaire"))
        if quota["used"] >= quota["max"]:
            flash(
                f"Question limit reached. Your plan includes 5 starter questions plus "
                f"{quota['additional']} additional question(s) ({quota['max']} total). "
                "Edit or delete an existing question before adding another.",
                "danger",
            )
            return redirect(url_for("master_questionnaire"))

        question_text = request.form.get("question_text", "").strip()
        question_type = request.form.get("question_type", "").strip()
        is_required = request.form.get("is_required") == "on"
        min_label = request.form.get("min_label", "Poor").strip() or "Poor"
        max_label = request.form.get("max_label", "Excellent").strip() or "Excellent"
        allow_comment = request.form.get("allow_comment") == "on"
        target_scope = request.form.get("target_scope", "overall").strip()
        try:
            question_order = int(request.form.get("question_order", "0"))
        except ValueError:
            question_order = 0

        if not question_text:
            flash("Question text is required.", "danger")
            return redirect(url_for("master_questionnaire"))

        if question_type not in {"rating", "text", "multiple_choice"}:
            flash("Invalid question type.", "danger")
            return redirect(url_for("master_questionnaire"))

        new_question_id = add_template_question(
            template_questionnaire_id=template_id,
            question_text=question_text,
            question_type=question_type,
            is_required=is_required,
            question_order=question_order,
            min_label=min_label,
            max_label=max_label,
            allow_comment=allow_comment,
            target_scope=target_scope,
        )
        
        # Log the question addition
        log_audit(
            entity_type="question",
            entity_id=new_question_id,
            action="created",
            new_values=f"Text: {question_text}, Type: {question_type}, Required: {is_required}"
        )
        
        flash("Question Added Successfully", "success")
        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/questionnaire/questions/<int:master_question_id>/delete", methods=["POST"])
    def master_delete_question(master_question_id: int):
        # Get question text before deletion for logging
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT question_text FROM questions WHERE id = %s", (master_question_id,))
            question = cursor.fetchone()
            question_text = question[0] if question else "Unknown"
        finally:
            conn.close()
        
        delete_template_question(template_question_id=master_question_id)
        
        # Log the question deletion
        log_audit(
            entity_type="question",
            entity_id=master_question_id,
            action="deleted",
            old_values=f"Text: {question_text}"
        )
        
        flash("Question Deleted", "success")
        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/questionnaire/questions/<int:master_question_id>/edit", methods=["POST"])
    def master_edit_question(master_question_id: int):
        question_text = request.form.get("question_text", "").strip()
        question_type = request.form.get("question_type", "").strip()
        is_required = request.form.get("is_required") == "on"
        min_label = request.form.get("min_label", "Poor").strip() or "Poor"
        max_label = request.form.get("max_label", "Excellent").strip() or "Excellent"
        allow_comment = request.form.get("allow_comment") == "on"
        target_scope = request.form.get("target_scope", "overall").strip()

        if not question_text:
            flash("Question text is required.", "danger")
            return redirect(url_for("master_questionnaire"))

        update_template_question(master_question_id, question_text, question_type, is_required, min_label, max_label, allow_comment, target_scope)
        flash("Question Updated Successfully", "success")
        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/questionnaire/questions/<int:master_question_id>/target", methods=["POST"])
    def master_update_question_target(master_question_id: int):
        target_scope = request.form.get("target_scope", "overall").strip()
        if target_scope not in {"overall", "staff", "manager"}:
            flash("Invalid question target.", "danger")
            return redirect(url_for("master_questionnaire"))
        template = ensure_template_questionnaire()
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """UPDATE questions SET target_scope = %s
                   WHERE id = %s AND questionnaire_id = %s AND is_template = TRUE""",
                (target_scope, master_question_id, int(template['id'])),
            )
        flash("Question target updated. Publish or sync to apply it to stores.", "success")
        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/questionnaire/questions/<int:master_question_id>/options/add", methods=["POST"])
    def master_add_option(master_question_id: int):
        option_text = request.form.get("option_text", "").strip()
        if not option_text:
            flash("Option text is required.", "danger")
            return redirect(url_for("master_questionnaire"))
        add_template_option(template_question_id=master_question_id, option_text=option_text)
        flash("Option added.", "success")
        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/questionnaire/options/<int:master_option_id>/delete", methods=["POST"])
    def master_delete_option(master_option_id: int):
        delete_template_option(template_option_id=master_option_id)
        flash("Option deleted.", "success")
        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/questionnaire/upload-logo", methods=["POST"])
    def master_upload_logo():
        template = ensure_template_questionnaire()
        # Handle logo upload for master questionnaire - store as base64 in database
        logo_data = None
        if 'logo' in request.files:
            logo_file = request.files['logo']
            if logo_file and logo_file.filename:
                # Validate file type
                allowed_extensions = {'png', 'jpg', 'jpeg'}
                if '.' not in logo_file.filename or logo_file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
                    flash("Invalid file type. Only PNG, JPG, and JPEG files are allowed.", "danger")
                    return redirect(url_for("master_questionnaire"))
                
                # Validate file size (5MB max)
                logo_file.seek(0, os.SEEK_END)
                file_size = logo_file.tell()
                logo_file.seek(0)
                if file_size > 5 * 1024 * 1024:
                    flash("File size exceeds 5MB limit.", "danger")
                    return redirect(url_for("master_questionnaire"))
                
                # Convert image to base64 with data URI prefix
                import base64
                logo_bytes = logo_file.read()
                file_ext = logo_file.filename.rsplit('.', 1)[1].lower()
                mime_type = f"image/{file_ext}"
                logo_data = f"data:{mime_type};base64,{base64.b64encode(logo_bytes).decode('utf-8')}"

        # Update the master template questionnaire with the base64 logo data
        if logo_data:
            conn = get_db_connection()
            try:
                cursor = conn.cursor()
                cursor.execute("UPDATE questionnaires SET logo_url = %s WHERE id = %s", (logo_data, template['id']))
                conn.commit()
                flash("Brand logo uploaded successfully", "success")
            except Exception as e:
                logger.error(f"Error uploading logo: {e}")
                flash(f"Error uploading logo: {e}", "danger")
            finally:
                conn.close()
        else:
            flash("No file selected", "warning")

        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/questionnaire/delete-logo", methods=["POST"])
    def master_delete_logo():
        # Delete the logo from the master questionnaire
        template = ensure_template_questionnaire()
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("UPDATE questionnaires SET logo_url = NULL WHERE id = %s", (template['id'],))
            conn.commit()
            flash("Brand logo deleted successfully", "success")
        except Exception as e:
            logger.error(f"Error deleting logo: {e}")
            flash(f"Error deleting logo: {e}", "danger")
        finally:
            conn.close()

        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/questionnaire/branding", methods=["POST"])
    @login_required
    def save_tenant_branding():
        user = get_user_by_id(session['user_id'])
        if not user or user.get('role') not in ('admin', 'superadmin'):
            return jsonify({"success": False, "error": "Access denied"}), 403
        fields = ("primary_color", "secondary_color", "accent_color", "text_color")
        colors = {name: request.form.get(name, "").strip().upper() for name in fields}
        if any(not re.fullmatch(r"#[0-9A-F]{6}", value) for value in colors.values()):
            flash("Please select valid brand colors.", "danger")
            return redirect(url_for("master_questionnaire"))
        with get_db_connection_with_transaction() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """INSERT INTO tenant_branding
                       (owner_user_id, scope_key, license_key, primary_color, secondary_color, accent_color, text_color)
                   VALUES (%s, %s, %s, %s, %s, %s, %s)
                   ON DUPLICATE KEY UPDATE license_key = VALUES(license_key),
                       primary_color = VALUES(primary_color), secondary_color = VALUES(secondary_color),
                       accent_color = VALUES(accent_color), text_color = VALUES(text_color)""",
                (user['id'], user.get('license_key') or f"user:{user['id']}", user.get('license_key'), colors['primary_color'], colors['secondary_color'],
                 colors['accent_color'], colors['text_color']),
            )
        flash("Brand colors saved for your company license.", "success")
        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/questionnaire/publish", methods=["POST"])
    def master_publish():
        template = ensure_template_questionnaire()
        template_id = int(template["id"])
        questions = fetch_template_questions(template_questionnaire_id=template_id)
        if not questions:
            flash("Add at least 1 question before publishing.", "danger")
            return redirect(url_for("master_questionnaire"))

        try:
            count = publish_template_to_all_stores()
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("master_questionnaire"))
        
        # Log the publish action
        log_audit(
            entity_type="questionnaire",
            entity_id=template_id,
            action="published",
            new_values=f"Published to {count} store(s)"
        )
        
        flash(f"Published to {count} store(s) Successfully", "success")
        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/api/stores", methods=["GET"])
    def api_stores():
        """Return only stores owned by the current questionnaire tenant."""
        template = ensure_template_questionnaire()
        stores = [store for store in fetch_stores(user_id=session['user_id'])
                  if (store.get('license_key') or None) == (template.get('license_key') or None)]
        return jsonify(stores)

    @app.route("/admin/questionnaire/sync", methods=["POST"])
    def sync_to_selected_stores():
        """Sync master questionnaire to selected stores"""
        try:
            data = request.get_json() or {}
            try:
                store_ids = list(dict.fromkeys(int(store_id) for store_id in data.get('store_ids', [])))
            except (TypeError, ValueError):
                return {"success": False, "error": "Invalid store selection"}, 400
            
            if not store_ids:
                return {"success": False, "error": "No stores selected"}, 400
            
            template = ensure_template_questionnaire()
            template_id = int(template["id"])
            template_questions = fetch_template_questions(template_questionnaire_id=template_id)
            template_options_by_question_id = fetch_template_options_by_question([int(q["id"]) for q in template_questions])
            owner = _tenant_owner()
            enforce_questionnaire_limit(owner, store_ids)
            
            conn = get_db_connection()
            try:
                cursor = conn.cursor(dictionary=True)
                synced_count = 0
                
                for store_id in store_ids:
                    store_id = int(store_id)
                    
                    # Check if store exists
                    cursor.execute("""SELECT id, logo_url FROM stores
                                      WHERE id = %s AND user_id = %s AND license_key <=> %s""",
                                   (store_id, session['user_id'], template.get('license_key')))
                    scoped_store = cursor.fetchone()
                    if not scoped_store:
                        continue
                    
                    # Check if store already has a questionnaire
                    cursor.execute("SELECT id FROM questionnaires WHERE store_id = %s AND is_template = FALSE ORDER BY id ASC LIMIT 1", (store_id,))
                    existing = cursor.fetchone()
                    
                    if existing:
                        # Update existing questionnaire metadata
                        cursor.execute(
                            """
                            UPDATE questionnaires
                            SET title = %s, is_active = %s, template_id = %s,
                                owner_user_id = %s, license_key = %s, logo_url = %s
                            WHERE id = %s
                            """,
                            (template["title"], bool(template["is_active"]), template_id,
                             session['user_id'], template.get('license_key'), scoped_store.get('logo_url') or template.get('logo_url'), int(existing["id"])),
                        )
                        questionnaire_id = int(existing["id"])
                        
                        # Deactivate existing questions
                        cursor.execute(
                            """
                            UPDATE questions
                            SET is_active = FALSE
                            WHERE questionnaire_id = %s AND is_template = FALSE
                            """,
                            (questionnaire_id,),
                        )
                    else:
                        # Create new store questionnaire
                        cursor.execute(
                            """
                            INSERT INTO questionnaires
                                (store_id, owner_user_id, license_key, title, is_active, is_template, template_id, logo_url)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (store_id, session['user_id'], template.get('license_key'), template["title"],
                             bool(template["is_active"]), False, template_id, scoped_store.get('logo_url') or template.get('logo_url')),
                        )
                        questionnaire_id = int(cursor.lastrowid)
                    
                    # Copy questions from template
                    for template_question in template_questions:
                        cursor.execute(
                            """
                            INSERT INTO questions (questionnaire_id, question_text, question_type, target_scope, min_label, max_label, allow_comment, is_required, question_order, is_template, template_id)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (questionnaire_id, template_question["question_text"], template_question["question_type"],
                             template_question.get("target_scope", "overall"),
                             template_question["min_label"], template_question["max_label"], template_question["allow_comment"],
                             template_question["is_required"], template_question["question_order"], False, int(template_question["id"])),
                        )
                        new_question_id = int(cursor.lastrowid)
                        
                        # Copy options for this question
                        template_options = template_options_by_question_id.get(int(template_question["id"]), [])
                        for option in template_options:
                            cursor.execute(
                                """
                                INSERT INTO question_options (question_id, option_text, is_template)
                                VALUES (%s, %s, %s)
                                """,
                                (new_question_id, option["option_text"], False),
                            )
                    
                    synced_count += 1
                
                conn.commit()
                
                # Log the sync action
                log_audit(
                    entity_type="questionnaire",
                    entity_id=template_id,
                    action="synced",
                    new_values=f"Synced to {synced_count} store(s)"
                )
                
                return {"success": True, "count": synced_count}
                
            finally:
                conn.close()
                
        except ValueError as e:
            logger.warning(f"Questionnaire sync blocked: {e}")
            return {"success": False, "error": str(e)}, 409
        except Exception as e:
            logger.error(f"Error syncing to stores: {e}")
            return {"success": False, "error": str(e)}, 500

    @app.route("/admin/questionnaire/sync-status", methods=["GET"])
    def sync_status():
        """Check sync status of stores"""
        try:
            template = ensure_template_questionnaire()
            template_id = int(template["id"])
            
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            
            # Get total stores
            cursor.execute("SELECT COUNT(*) as total FROM stores WHERE user_id = %s AND license_key <=> %s",
                           (session['user_id'], template.get('license_key')))
            total_stores = cursor.fetchone()['total']
            
            # Get stores with synced questionnaire
            cursor.execute("""
                SELECT COUNT(DISTINCT s.id) as synced
                FROM stores s
                JOIN questionnaires q ON s.id = q.store_id
                WHERE q.is_template = FALSE AND q.template_id = %s AND s.user_id = %s
                  AND s.license_key <=> %s
            """, (template_id, session['user_id'], template.get('license_key')))
            synced_stores = cursor.fetchone()['synced']
            
            cursor.close()
            conn.close()
            
            unsynced_count = total_stores - synced_stores
            
            return jsonify({
                "synced": unsynced_count == 0,
                "synced_count": synced_stores,
                "unsynced_count": unsynced_count,
                "total_stores": total_stores
            })
            
        except Exception as e:
            logger.error(f"Error checking sync status: {e}")
            return jsonify({"synced": False, "error": str(e)}), 500

    @app.route("/admin/questionnaire/toggle-active", methods=["POST"])
    def master_toggle_active():
        template = ensure_template_questionnaire()
        current_active = template.get("is_active", False)
        new_active = not current_active
        
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                """
                UPDATE questionnaires
                SET is_active = %s
                WHERE id = %s AND is_template = TRUE
                """,
                (new_active, int(template["id"])),
            )
            conn.commit()
        finally:
            conn.close()
        
        # Log the toggle action
        log_audit(
            entity_type="questionnaire",
            entity_id=int(template["id"]),
            action="toggled",
            old_values=f"Active: {current_active} → {new_active}"
        )
        
        if new_active:
            flash("Survey enabled successfully. Stores can now accept feedback.", "success")
        else:
            flash("Survey disabled successfully. Stores can no longer accept feedback.", "warning")
        
        return redirect(url_for("master_questionnaire"))

    @app.route("/admin/questionnaire/preview")
    def master_preview():
        template = ensure_template_questionnaire()
        template_id = int(template["id"])
        questions = fetch_template_questions(template_questionnaire_id=template_id)
        question_ids = [q["id"] for q in questions]
        options_by_question_id = fetch_options_for_questions(question_ids)

        return render_template(
            "master_questionnaire/preview.html",
            master=template,
            questions=questions,
            options_by_question_id=options_by_question_id,
        )

    # -------------------------
    # DASHBOARD ANALYTICS
    # -------------------------
    # Bayesian-average smoothing constant: how many feedbacks a store/staff
    # needs before its own average dominates the global prior.
    BAYESIAN_C = 5

    def fetch_dashboard_analytics() -> Dict[str, Any]:
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)

            # One scope drives every dashboard query so totals, rankings,
            # activity, and staff never leak stores outside the user's access.
            dashboard_user = get_user_by_id(session['user_id'])
            scoped_store_ids: Optional[List[int]] = None
            if dashboard_user['role'] == 'admin':
                scoped_store_ids = [s['id'] for s in fetch_stores(user_id=session['user_id'])]
            elif dashboard_user['role'] == 'user':
                scoped_store_ids = get_assigned_store_ids(session['user_id'])

            scope_join = ""
            if scoped_store_ids is not None:
                cursor.execute("CREATE TEMPORARY TABLE dashboard_store_scope (store_id INT PRIMARY KEY)")
                if scoped_store_ids:
                    cursor.executemany(
                        "INSERT INTO dashboard_store_scope (store_id) VALUES (%s)",
                        [(store_id,) for store_id in scoped_store_ids],
                    )
                scope_join = "INNER JOIN dashboard_store_scope dss ON dss.store_id = s.id"

            # Global average rating across all stores (prior `m`).
            # Falls back to 4.0 (mid-high default) when there are no ratings yet.
            cursor.execute(
                f"""
                SELECT AVG(a.rating_value) as global_avg
                FROM stores s
                {scope_join}
                LEFT JOIN questionnaires q ON s.id = q.store_id
                LEFT JOIN responses r ON q.id = r.questionnaire_id
                LEFT JOIN answers a ON r.id = a.response_id
                JOIN questions q2 ON a.question_id = q2.id
                WHERE q2.question_type = 'rating' AND a.rating_value IS NOT NULL
                """
            )
            row = cursor.fetchone()
            global_avg_rating = float(row['global_avg']) if row and row['global_avg'] is not None else 4.0

            # Store overview data — `weighted_score` is now a Bayesian average:
            #   (C * m + sum_of_ratings) / (C + n)
            # which keeps the score on the 1–5 scale and pulls low-volume stores
            # toward the global mean until they accumulate enough feedback.
            cursor.execute(
                f"""
                SELECT s.id, s.store_name, s.address, s.city, s.created_at,
                       COUNT(DISTINCT r.id) as total_responses,
                       AVG(CASE WHEN q2.question_type = 'rating' THEN a.rating_value END) as avg_rating,
                       COUNT(DISTINCT CASE WHEN q2.question_type = 'rating' AND a.rating_value IS NOT NULL THEN r.id END) as rating_feedback_count,
                       (
                           (%s * %s) + COALESCE(SUM(CASE WHEN q2.question_type = 'rating' THEN a.rating_value END), 0)
                       ) / (
                           %s + COUNT(CASE WHEN q2.question_type = 'rating' AND a.rating_value IS NOT NULL THEN 1 END)
                       ) as weighted_score,
                       COUNT(DISTINCT r.user_email) as unique_users
                FROM stores s
                {scope_join}
                LEFT JOIN questionnaires q ON s.id = q.store_id
                LEFT JOIN responses r ON q.id = r.questionnaire_id
                LEFT JOIN answers a ON r.id = a.response_id
                LEFT JOIN questions q2 ON a.question_id = q2.id
                GROUP BY s.id, s.store_name, s.address, s.city, s.created_at
                ORDER BY weighted_score DESC, total_responses DESC
                """,
                (BAYESIAN_C, global_avg_rating, BAYESIAN_C),
            )
            stores_data = cursor.fetchall()
            
            # Convert Decimal values to float for template compatibility
            for store in stores_data:
                store['avg_rating'] = float(store['avg_rating']) if store['avg_rating'] is not None else 0.0
                store['rating_feedback_count'] = int(store['rating_feedback_count']) if store.get('rating_feedback_count') else 0
                store['weighted_score'] = float(store['weighted_score']) if store.get('weighted_score') is not None else 0.0
            
            # Overall statistics
            cursor.execute(
                f"""
                SELECT 
                    COUNT(DISTINCT r.id) as total_responses,
                    COUNT(DISTINCT s.id) as total_stores,
                    COUNT(DISTINCT r.user_email) as total_unique_users,
                    AVG(CASE WHEN q2.question_type = 'rating' THEN a.rating_value END) as overall_avg_rating,
                    COUNT(DISTINCT q.id) as total_questionnaires
                FROM stores s
                {scope_join}
                LEFT JOIN questionnaires q ON s.id = q.store_id
                LEFT JOIN responses r ON q.id = r.questionnaire_id
                LEFT JOIN answers a ON r.id = a.response_id
                LEFT JOIN questions q2 ON a.question_id = q2.id
                """
            )
            overall_stats = cursor.fetchone()
            
            if overall_stats:
                overall_stats['overall_avg_rating'] = float(overall_stats['overall_avg_rating']) if overall_stats['overall_avg_rating'] is not None else 0.0
            else:
                overall_stats = {
                    'total_responses': 0,
                    'total_stores': 0,
                    'total_unique_users': 0,
                    'overall_avg_rating': 0,
                    'total_questionnaires': 0
                }
            
            # Recent activity (last 7 days, linked to stores)
            cursor.execute(
                f"""
                SELECT DATE(r.submitted_at) as date, COUNT(DISTINCT r.id) as responses
                FROM responses r
                INNER JOIN questionnaires q ON r.questionnaire_id = q.id
                INNER JOIN stores s ON q.store_id = s.id
                {scope_join}
                WHERE r.submitted_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                GROUP BY DATE(r.submitted_at)
                ORDER BY date
                """
            )
            recent_activity = cursor.fetchall()
            
            # Top performing stores by feedback
            cursor.execute(
                f"""
                SELECT s.store_name, COUNT(r.id) as response_count
                FROM stores s
                {scope_join}
                LEFT JOIN questionnaires q ON s.id = q.store_id
                LEFT JOIN responses r ON q.id = r.questionnaire_id
                GROUP BY s.id, s.store_name
                ORDER BY response_count DESC
                LIMIT 5
                """
            )
            top_stores = cursor.fetchall()

            # Best overall store ranked by Bayesian-average score so a store
            # with one lucky 5★ doesn't beat a store with sustained 4.6★.
            cursor.execute(
                f"""
                SELECT s.id, s.store_name, s.address, s.city,
                       COUNT(DISTINCT r.id) as total_responses,
                       AVG(CASE WHEN q2.question_type = 'rating' THEN a.rating_value END) as avg_rating,
                       (
                           (%s * %s) + COALESCE(SUM(CASE WHEN q2.question_type = 'rating' THEN a.rating_value END), 0)
                       ) / (
                           %s + COUNT(CASE WHEN q2.question_type = 'rating' AND a.rating_value IS NOT NULL THEN 1 END)
                       ) as weighted_score
                FROM stores s
                {scope_join}
                LEFT JOIN questionnaires q ON s.id = q.store_id
                LEFT JOIN responses r ON q.id = r.questionnaire_id
                LEFT JOIN answers a ON r.id = a.response_id
                LEFT JOIN questions q2 ON a.question_id = q2.id
                WHERE q2.question_type = 'rating'
                GROUP BY s.id, s.store_name, s.address, s.city
                HAVING total_responses >= 1
                ORDER BY weighted_score DESC, total_responses DESC
                LIMIT 1
                """,
                (BAYESIAN_C, global_avg_rating, BAYESIAN_C),
            )
            best_overall_store = cursor.fetchone()
            if best_overall_store:
                best_overall_store['avg_rating'] = float(best_overall_store['avg_rating']) if best_overall_store['avg_rating'] is not None else 0.0

            # Global average commendation rating (Bayesian prior `m`).
            cursor.execute(
                f"""SELECT AVG(sc.rating) as global_avg
                    FROM stores s
                    {scope_join}
                    LEFT JOIN staff stf ON stf.store_id = s.id
                    LEFT JOIN staff_commendations sc ON sc.staff_id = stf.id
                    WHERE sc.rating IS NOT NULL"""
            )
            srow = cursor.fetchone()
            staff_global_avg = float(srow['global_avg']) if srow and srow['global_avg'] is not None else 4.0

            # Best overall staff (highest Bayesian-average score).
            cursor.execute(
                f"""
                SELECT s.id, s.first_name, s.last_name, s.position, s.role,
                       AVG(sc.rating) as avg_rating,
                       COUNT(sc.id) as commendation_count,
                       (
                           (%s * %s) + COALESCE(SUM(sc.rating), 0)
                       ) / (
                           %s + COUNT(sc.rating)
                       ) as weighted_score,
                       st.store_name
                FROM staff s
                LEFT JOIN staff_commendations sc ON s.id = sc.staff_id
                LEFT JOIN responses r ON sc.response_id = r.id
                LEFT JOIN questionnaires q ON r.questionnaire_id = q.id
                LEFT JOIN stores st ON q.store_id = st.id
                {scope_join.replace('s.id', 'st.id')}
                GROUP BY s.id, s.first_name, s.last_name, s.position, s.role, st.store_name
                HAVING avg_rating IS NOT NULL
                ORDER BY weighted_score DESC
                LIMIT 1
                """,
                (BAYESIAN_C, staff_global_avg, BAYESIAN_C),
            )
            best_overall_staff = cursor.fetchone()
            if best_overall_staff:
                best_overall_staff['avg_rating'] = float(best_overall_staff['avg_rating']) if best_overall_staff['avg_rating'] is not None else 0.0
                best_overall_staff['weighted_score'] = float(best_overall_staff['weighted_score']) if best_overall_staff['weighted_score'] is not None else 0.0

            return {
                'stores_data': stores_data,
                'overall_stats': overall_stats,
                'recent_activity': recent_activity,
                'top_stores': top_stores,
                'best_overall_store': best_overall_store,
                'best_overall_staff': best_overall_staff
            }
        finally:
            conn.close()

    @app.route("/admin/dashboard")
    @login_required
    def admin_dashboard():
        try:
            analytics = fetch_dashboard_analytics()
            return render_template("dashboard/dashboard.html", **analytics)
        except Exception as e:
            error_details = traceback.format_exc()
            logger.error(f"Dashboard Crash: {e}\n{error_details}")
            return f"Dashboard Error: {e}<br><pre>{error_details}</pre>", 500

    @app.route("/admin/users")
    @role_required('superadmin')
    def admin_users():
        """User management page"""
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users ORDER BY created_at DESC")
            users = cursor.fetchall()
            config = get_license_config()
            portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
            return render_template("admin/users.html", users=users, licensing_portal_url=portal_url)
        finally:
            conn.close()

    @app.route("/admin/users/add", methods=["POST"])
    @role_required('superadmin')
    def admin_add_user():
        """Add a new user"""
        username = request.form.get("username", "").strip()
        email = request.form.get("email", "").strip()
        password = request.form.get("password", "")
        role = request.form.get("role", "user")
        license_key = request.form.get("license_key", "").strip()
        max_stores = 0
        
        if not username or not email or not password:
            flash("Username, email, and password are required.", "danger")
            return redirect(url_for("admin_users"))
        
        if role not in ['superadmin', 'admin', 'user']:
            flash("Invalid role.", "danger")
            return redirect(url_for("admin_users"))

        # The licensing portal is the source of truth for client limits.
        if role == 'admin':
            if not license_key:
                flash("A License Key from the Licensing Portal is required for a client account.", "danger")
                return redirect(url_for("admin_users"))
            try:
                import requests as http_requests
                config = get_license_config()
                portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
                response = http_requests.post(
                    f"{portal_url}/api/validate/{license_key}",
                    headers=licensing_api_headers(), timeout=10,
                )
                license_data = response.json() if response.content else {}
                if response.status_code != 200 or not license_data.get("valid"):
                    flash(license_data.get("message") or "The License Key is invalid or inactive.", "danger")
                    return redirect(url_for("admin_users"))
                max_stores = int(license_data.get("max_stores") or 0)
            except Exception as exc:
                logger.error("Unable to validate client license: %s", exc)
                flash("Unable to connect to the Licensing Portal. Please try again.", "danger")
                return redirect(url_for("admin_users"))
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Check if username or email already exists
            cursor.execute("SELECT id FROM users WHERE username = %s OR email = %s", (username, email))
            if cursor.fetchone():
                flash("Username or email already exists.", "danger")
                return redirect(url_for("admin_users"))
            if license_key:
                cursor.execute("SELECT id FROM users WHERE license_key = %s", (license_key,))
                if cursor.fetchone():
                    flash("This License Key is already connected to another client.", "danger")
                    return redirect(url_for("admin_users"))
            
            password_hash = hash_password(password)
            cursor.execute(
                "INSERT INTO users (username, email, password_hash, role, max_stores, license_key) VALUES (%s, %s, %s, %s, %s, %s)",
                (username, email, password_hash, role, max_stores if role == 'admin' else 0, license_key if role == 'admin' else None)
            )
            conn.commit()
            conn.close()
            
            if role == 'admin':
                flash(f"Client account created and connected to its license ({max_stores or 'Unlimited'} stores).", "success")
                log_audit(
                    entity_type="user",
                    entity_id=0,
                    action="created",
                    new_values=f"Client {username} created with max_stores={max_stores}"
                )
            else:
                flash("User created successfully.", "success")
                log_audit(
                    entity_type="user",
                    entity_id=0,
                    action="created",
                    new_values=f"User {username} created with role {role}"
                )
        except Exception as e:
            logger.error(f"Error creating user: {e}")
            flash("Failed to create user.", "danger")
        
        return redirect(url_for("admin_users"))

    @app.route("/api/admin/licenses/validate", methods=["POST"])
    @role_required('superadmin')
    def admin_validate_license():
        """Validate a pasted portal key and return its licensed limits."""
        data = request.get_json(silent=True) or {}
        license_key = (data.get("license_key") or "").strip()
        if not license_key:
            return jsonify({"valid": False, "error": "License Key is required"}), 400
        try:
            import requests as http_requests
            config = get_license_config()
            portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
            response = http_requests.post(
                f"{portal_url}/api/validate/{license_key}",
                headers=licensing_api_headers(), timeout=10,
            )
            payload = response.json() if response.content else {}
            return jsonify(payload), response.status_code
        except Exception as exc:
            logger.error("License preview validation failed: %s", exc)
            return jsonify({"valid": False, "error": "Licensing Portal is unavailable"}), 502

    @app.route("/admin/users/<int:user_id>/toggle", methods=["POST"])
    @role_required('superadmin')
    def admin_toggle_user(user_id: int):
        """Toggle user active status"""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Don't allow deactivating yourself
            if user_id == session.get('user_id'):
                flash("You cannot deactivate your own account.", "danger")
                return redirect(url_for("admin_users"))
            
            cursor.execute("UPDATE users SET is_active = NOT is_active WHERE id = %s", (user_id,))
            conn.commit()
            conn.close()
            flash("User status updated successfully.", "success")
            log_audit(
                entity_type="user",
                entity_id=user_id,
                action="toggled",
                old_values="User status toggled"
            )
        except Exception as e:
            logger.error(f"Error toggling user: {e}")
            flash("Failed to update user status.", "danger")
        
        return redirect(url_for("admin_users"))

    @app.route("/admin/users/<int:user_id>/generate-license", methods=["POST"])
    @role_required('superadmin')
    def admin_generate_user_license(user_id: int):
        """Generate and attach a client license without a second portal login."""
        user = get_user_by_id(user_id)
        if not user or user.get("role") != "admin":
            flash("A license can only be generated for an Admin (Client) account.", "danger")
            return redirect(url_for("admin_users"))

        config = get_license_config()
        portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
        try:
            import requests as http_requests
            response = http_requests.post(
                f"{portal_url}/api/licenses/generate",
                headers=licensing_api_headers(),
                json={
                    "external_user_id": user_id,
                    "company_name": user.get("username") or user.get("email"),
                    "contact_email": user.get("email"),
                    "max_stores": int(user.get("max_stores") or 0),
                    "max_questionnaires": 0,
                },
                timeout=15,
            )
            payload = response.json() if response.content else {}
            if response.status_code not in (200, 201) or not payload.get("license_key"):
                logger.error("Portal license generation failed: %s %s", response.status_code, response.text)
                flash(payload.get("error") or "Unable to generate license from the licensing portal.", "danger")
                return redirect(url_for("admin_users"))

            conn = get_db_connection()
            try:
                cursor = conn.cursor()
                cursor.execute("UPDATE users SET license_key = %s WHERE id = %s", (payload["license_key"], user_id))
                conn.commit()
            finally:
                conn.close()

            flash(f"License generated and connected to {user['username']}.", "success")
            log_audit("user", user_id, "license_generated", new_values=f"Portal: {portal_url}")
        except Exception as exc:
            logger.error("Error generating client license: %s", exc)
            flash("Unable to connect to the licensing portal. Check the portal URL and shared API key.", "danger")
        return redirect(url_for("admin_users"))

    @app.route("/account/password", methods=["GET", "POST"])
    @login_required
    def account_change_password():
        """Self-service account settings: change username and/or password."""
        user_id = session.get('user_id')
        if not user_id:
            return redirect(url_for('login'))

        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id, username, email, role, password_hash, created_at, is_active FROM users WHERE id = %s", (user_id,))
            user = cursor.fetchone()

            if not user:
                session.clear()
                flash("Your account could not be found. Please log in again.", "danger")
                return redirect(url_for('login'))

            if request.method == "POST":
                form_type = request.form.get("form_type", "password")

                if form_type == "email":
                    new_email = request.form.get("new_email", "").strip()

                    if not new_email:
                        flash("Email is required.", "danger")
                        return redirect(url_for("account_change_password"))

                    if "@" not in new_email or "." not in new_email.split("@")[-1]:
                        flash("Please enter a valid email address.", "danger")
                        return redirect(url_for("account_change_password"))

                    if new_email == user['email']:
                        flash("New email must be different from your current email.", "warning")
                        return redirect(url_for("account_change_password"))

                    cursor.execute(
                        "SELECT id FROM users WHERE email = %s AND id != %s",
                        (new_email, user_id)
                    )
                    if cursor.fetchone():
                        flash("That email is already in use by another account.", "danger")
                        return redirect(url_for("account_change_password"))

                    cursor.execute(
                        "UPDATE users SET email = %s WHERE id = %s",
                        (new_email, user_id)
                    )
                    conn.commit()

                    try:
                        log_audit(
                            entity_type="user",
                            entity_id=user_id,
                            action="email_changed",
                            old_values=user['email'],
                            new_values=new_email,
                            user_id=user_id
                        )
                    except Exception:
                        pass

                    flash(f"Email updated to {new_email}.", "success")
                    return redirect(url_for("account_change_password"))

                if form_type == "username":
                    new_username = request.form.get("new_username", "").strip()

                    if not new_username:
                        flash("New username is required.", "danger")
                        return redirect(url_for("account_change_password"))

                    if len(new_username) < 2:
                        flash("Username must be at least 2 characters long.", "danger")
                        return redirect(url_for("account_change_password"))

                    if new_username == user['username']:
                        flash("New username must be different from your current username.", "warning")
                        return redirect(url_for("account_change_password"))

                    # Uniqueness check
                    cursor.execute(
                        "SELECT id FROM users WHERE username = %s AND id != %s",
                        (new_username, user_id)
                    )
                    if cursor.fetchone():
                        flash("That username is already taken.", "danger")
                        return redirect(url_for("account_change_password"))

                    cursor.execute(
                        "UPDATE users SET username = %s WHERE id = %s",
                        (new_username, user_id)
                    )
                    conn.commit()
                    session['username'] = new_username

                    try:
                        log_audit(
                            entity_type="user",
                            entity_id=user_id,
                            action="username_changed",
                            old_values=user['username'],
                            new_values=new_username,
                            user_id=user_id
                        )
                    except Exception:
                        pass

                    flash(f"Username changed to {new_username}.", "success")
                    return redirect(url_for("account_change_password"))

                # Password change (default)
                current_password = request.form.get("current_password", "")
                new_password = request.form.get("new_password", "")
                confirm_password = request.form.get("confirm_password", "")

                if not current_password or not new_password or not confirm_password:
                    flash("All password fields are required.", "danger")
                    return redirect(url_for("account_change_password"))

                if not verify_password(current_password, user['password_hash']):
                    flash("Current password is incorrect.", "danger")
                    return redirect(url_for("account_change_password"))

                if len(new_password) < 4:
                    flash("New password must be at least 4 characters long.", "danger")
                    return redirect(url_for("account_change_password"))

                if new_password != confirm_password:
                    flash("New password and confirmation do not match.", "danger")
                    return redirect(url_for("account_change_password"))

                if new_password == current_password:
                    flash("New password must be different from your current password.", "warning")
                    return redirect(url_for("account_change_password"))

                new_hash = hash_password(new_password)
                cursor.execute(
                    "UPDATE users SET password_hash = %s WHERE id = %s",
                    (new_hash, user_id)
                )
                conn.commit()

                try:
                    log_audit(
                        entity_type="user",
                        entity_id=user_id,
                        action="password_changed",
                        new_values="self-service password change",
                        user_id=user_id
                    )
                except Exception:
                    pass

                flash("Password updated successfully.", "success")
                return redirect(url_for("account_change_password"))

            return render_template("account/change_password.html", user=user)
        except Exception as e:
            logger.error(f"Error in account settings: {e}")
            flash(f"Error updating account: {e}", "danger")
            return redirect(url_for("admin_dashboard"))
        finally:
            conn.close()

    @app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
    @role_required('superadmin')
    def admin_edit_user(user_id: int):
        """Edit an existing user."""
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM users WHERE id = %s", (user_id,))
            user = cursor.fetchone()

            if not user:
                flash("User not found.", "danger")
                return redirect(url_for("admin_users"))

            if request.method == "POST":
                username = request.form.get("username", "").strip()
                email = request.form.get("email", "").strip()
                role = request.form.get("role", user['role'])
                try:
                    max_stores = int(request.form.get("max_stores", "0") or 0)
                except ValueError:
                    max_stores = 0
                is_active = request.form.get("is_active") == "on"
                new_password = request.form.get("new_password", "")

                if not username or not email:
                    flash("Username and email are required.", "danger")
                    return redirect(url_for("admin_edit_user", user_id=user_id))

                # Treat legacy 'dev' as superadmin (in case migration didn't run)
                if role == 'dev':
                    role = 'superadmin'

                if role not in ['superadmin', 'admin', 'user']:
                    flash(f"Invalid role: {role}", "danger")
                    return redirect(url_for("admin_edit_user", user_id=user_id))

                # Normalize stored legacy role for self-edit comparison
                stored_role = 'superadmin' if user['role'] in ('dev',) else user['role']

                # Don't allow demoting/deactivating yourself
                if user_id == session.get('user_id'):
                    if role != stored_role:
                        flash("You cannot change your own role.", "danger")
                        return redirect(url_for("admin_edit_user", user_id=user_id))
                    if not is_active:
                        flash("You cannot deactivate your own account.", "danger")
                        return redirect(url_for("admin_edit_user", user_id=user_id))

                # Check uniqueness of username/email (excluding self)
                cursor.execute(
                    "SELECT id FROM users WHERE (username = %s OR email = %s) AND id != %s",
                    (username, email, user_id)
                )
                if cursor.fetchone():
                    flash("Username or email already in use by another user.", "danger")
                    return redirect(url_for("admin_edit_user", user_id=user_id))

                old_values = {
                    'username': user['username'], 'email': user['email'],
                    'role': user['role'], 'max_stores': user.get('max_stores', 0),
                    'is_active': bool(user['is_active'])
                }

                # Update fields
                cursor.execute(
                    """
                    UPDATE users
                    SET username = %s,
                        email = %s,
                        role = %s,
                        max_stores = %s,
                        is_active = %s
                    WHERE id = %s
                    """,
                    (username, email, role, max_stores if role == 'admin' else 0, is_active, user_id)
                )

                # Optional password reset
                if new_password:
                    if len(new_password) < 4:
                        flash("Password must be at least 4 characters.", "danger")
                        conn.rollback()
                        return redirect(url_for("admin_edit_user", user_id=user_id))
                    password_hash = hash_password(new_password)
                    cursor.execute(
                        "UPDATE users SET password_hash = %s WHERE id = %s",
                        (password_hash, user_id)
                    )

                conn.commit()

                # If editing self, refresh session so new username/role show immediately
                if user_id == session.get('user_id'):
                    session['username'] = username
                    session['role'] = role

                new_values = {
                    'username': username, 'email': email, 'role': role,
                    'max_stores': max_stores if role == 'admin' else 0,
                    'is_active': is_active,
                    'password_changed': bool(new_password),
                }
                try:
                    log_audit(
                        entity_type="user",
                        entity_id=user_id,
                        action="updated",
                        old_values=str(old_values),
                        new_values=str(new_values),
                        user_id=session.get('user_id')
                    )
                except Exception:
                    pass

                flash("User updated successfully.", "success")
                return redirect(url_for("admin_users"))

            return render_template("admin/edit_user.html", user=user)
        except Exception as e:
            logger.error(f"Error editing user: {e}")
            flash(f"Error editing user: {e}", "danger")
            return redirect(url_for("admin_users"))
        finally:
            conn.close()

    @app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
    @role_required('superadmin')
    def admin_delete_user(user_id: int):
        """Delete a user"""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Don't allow deleting yourself
            if user_id == session.get('user_id'):
                flash("You cannot delete your own account.", "danger")
                return redirect(url_for("admin_users"))
            
            cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
            conn.commit()
            conn.close()
            flash("User deleted successfully.", "success")
            log_audit(
                entity_type="user",
                entity_id=user_id,
                action="deleted",
                user_id=session.get('user_id')
            )
        except Exception as e:
            logger.error(f"Error deleting user: {e}")
            flash(f"Error deleting user: {e}", "danger")
        return redirect(url_for("admin_users"))

    # ── Per-store view-only viewers (admin/superadmin manage who can view a store) ──
    def _user_can_manage_store(user: Dict[str, Any], store_id: int) -> bool:
        """Superadmin can manage any store; admin (client) can manage only stores they own."""
        if not user:
            return False
        if user['role'] == 'superadmin':
            return True
        if user['role'] != 'admin':
            return False
        # Verify ownership
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT user_id FROM stores WHERE id = %s", (store_id,))
            row = cursor.fetchone()
            return bool(row and row[0] == user['id'])
        finally:
            conn.close()

    @app.route("/admin/stores/<int:store_id>/viewers", methods=["GET"])
    @role_required('admin', 'superadmin')
    def store_viewers_list(store_id: int):
        """Return JSON: list of view-only users assigned to a store + available users to assign."""
        user = get_user_by_id(session['user_id'])
        if not _user_can_manage_store(user, store_id):
            return jsonify({"success": False, "error": "Forbidden"}), 403
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT u.id, u.username, u.email, us.created_at
                FROM user_stores us
                JOIN users u ON u.id = us.user_id
                WHERE us.store_id = %s
                ORDER BY u.username
                """,
                (store_id,)
            )
            assigned = cursor.fetchall()

            # A license may have at most one store viewer/manager per store,
            # and no more assigned viewers than the owning Admin's store limit.
            cursor.execute(
                """SELECT s.user_id AS owner_user_id, COALESCE(u.max_stores, 0) AS max_viewers,
                          u.license_key
                   FROM stores s
                   LEFT JOIN users u ON u.id = s.user_id
                   WHERE s.id = %s""",
                (store_id,),
            )
            ownership = cursor.fetchone() or {}
            owner_user_id = ownership.get('owner_user_id')
            max_viewers = int(ownership.get('max_viewers') or 0)
            if ownership.get('license_key'):
                license_status = validate_tenant_license(ownership['license_key'])
                if license_status.get('valid'):
                    max_viewers = int(license_status.get('max_stores') or 0)
            assigned_total = 0
            if owner_user_id:
                cursor.execute(
                    """SELECT COUNT(*) AS total
                       FROM user_stores us
                       JOIN stores s ON s.id = us.store_id
                       WHERE s.user_id = %s""",
                    (owner_user_id,),
                )
                assigned_total = int((cursor.fetchone() or {}).get('total') or 0)

            store_has_viewer = bool(assigned)
            limit_reached = max_viewers > 0 and assigned_total >= max_viewers

            # Available view-only users not yet assigned to this store
            if not store_has_viewer and not limit_reached:
                cursor.execute(
                    """SELECT u.id, u.username, u.email
                       FROM users u
                       WHERE u.role='user' AND u.is_active=TRUE
                         AND NOT EXISTS (
                           SELECT 1 FROM user_stores us WHERE us.user_id = u.id
                         )
                       ORDER BY u.username"""
                )
                available = cursor.fetchall()
            else:
                available = []
            return jsonify({
                "success": True,
                "assigned": assigned,
                "available": available,
                "assigned_total": assigned_total,
                "max_viewers": max_viewers,
                "limit_reached": limit_reached,
                "store_has_viewer": store_has_viewer,
            })
        finally:
            conn.close()

    @app.route("/admin/stores/<int:store_id>/viewers/add", methods=["POST"])
    @role_required('admin', 'superadmin')
    def store_viewers_add(store_id: int):
        """Assign an existing view-only user to a store."""
        user = get_user_by_id(session['user_id'])
        if not _user_can_manage_store(user, store_id):
            flash("You don't have permission to manage viewers for this store.", "danger")
            return redirect(url_for("stores_management"))

        try:
            target_user_id = int(request.form.get("user_id", "0") or 0)
        except ValueError:
            target_user_id = 0
        if not target_user_id:
            flash("Please choose a user to assign.", "danger")
            return redirect(url_for("stores_management", store_id=store_id))

        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT id, username, role FROM users WHERE id = %s", (target_user_id,))
            target = cursor.fetchone()
            if not target or target['role'] != 'user':
                flash("Selected account must be a view-only user.", "danger")
                return redirect(url_for("stores_management", store_id=store_id))

            cursor.execute("""SELECT s.user_id, u.license_key, COALESCE(u.max_stores, 0) AS max_viewers
                              FROM stores s LEFT JOIN users u ON u.id = s.user_id
                              WHERE s.id = %s""", (store_id,))
            store_row = cursor.fetchone()
            owner_user_id = store_row.get('user_id') if store_row else None

            cursor.execute("SELECT COUNT(*) AS total FROM user_stores WHERE store_id = %s", (store_id,))
            if int((cursor.fetchone() or {}).get('total') or 0) >= 1:
                flash("This store already has a viewer/manager. Remove the current viewer first.", "warning")
                return redirect(url_for("stores_management", store_id=store_id))

            cursor.execute("SELECT COUNT(*) AS total FROM user_stores WHERE user_id = %s", (target_user_id,))
            if int((cursor.fetchone() or {}).get('total') or 0) >= 1:
                flash("That viewer is already assigned to another store.", "warning")
                return redirect(url_for("stores_management", store_id=store_id))

            if owner_user_id:
                max_viewers = int(store_row.get('max_viewers') or 0)
                if store_row.get('license_key'):
                    license_status = validate_tenant_license(store_row['license_key'], force=True)
                    if not license_status.get('valid'):
                        flash("Unable to verify the store owner's license. Please try again.", "danger")
                        return redirect(url_for("stores_management", store_id=store_id))
                    max_viewers = int(license_status.get('max_stores') or 0)
                if max_viewers > 0:
                    cursor.execute(
                        """SELECT COUNT(*) AS total
                           FROM user_stores us
                           JOIN stores s ON s.id = us.store_id
                           WHERE s.user_id = %s""",
                        (owner_user_id,),
                    )
                    assigned_total = int((cursor.fetchone() or {}).get('total') or 0)
                    if assigned_total >= max_viewers:
                        flash(f"Viewer limit reached. This license allows up to {max_viewers} store viewers.", "danger")
                        return redirect(url_for("stores_management", store_id=store_id))

            try:
                cursor.execute(
                    "INSERT INTO user_stores (user_id, store_id) VALUES (%s, %s)",
                    (target_user_id, store_id)
                )
                conn.commit()
                flash(f"{target['username']} can now view this store.", "success")
                try:
                    log_audit(
                        entity_type="store",
                        entity_id=store_id,
                        action="viewer_added",
                        new_values=f"user_id={target_user_id}",
                        user_id=session.get('user_id')
                    )
                except Exception:
                    pass
            except Exception as e:
                # Likely unique constraint (already assigned)
                logger.info(f"Viewer add no-op: {e}")
                flash("That user is already assigned to this store.", "warning")
        finally:
            conn.close()

        return redirect(url_for("stores_management", store_id=store_id))

    @app.route("/admin/stores/<int:store_id>/viewers/<int:user_id>/remove", methods=["POST"])
    @role_required('admin', 'superadmin')
    def store_viewers_remove(store_id: int, user_id: int):
        """Remove a view-only user's access from a store."""
        user = get_user_by_id(session['user_id'])
        if not _user_can_manage_store(user, store_id):
            flash("You don't have permission to manage viewers for this store.", "danger")
            return redirect(url_for("stores_management"))

        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "DELETE FROM user_stores WHERE user_id = %s AND store_id = %s",
                (user_id, store_id)
            )
            conn.commit()
            flash("Viewer removed.", "success")
            try:
                log_audit(
                    entity_type="store",
                    entity_id=store_id,
                    action="viewer_removed",
                    new_values=f"user_id={user_id}",
                    user_id=session.get('user_id')
                )
            except Exception:
                pass
        finally:
            conn.close()

        return redirect(url_for("stores_management", store_id=store_id))

    @app.route("/api/licensing/users", methods=["GET"])
    def api_licensing_users():
        """API endpoint for licensing portal to fetch users"""
        # Simple API key check for security
        api_key = request.headers.get("X-Licensing-API-Key")
        expected_api_key = os.getenv("LICENSING_API_KEY")
        if not expected_api_key:
            return jsonify({"error": "Licensing API is not configured"}), 503
        if not api_key or not secrets.compare_digest(api_key, expected_api_key):
            return jsonify({"error": "Unauthorized"}), 401
        
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT id, username, email, role, max_stores, created_at, is_active
                FROM users
                WHERE role = 'admin'
                ORDER BY created_at DESC
            """)
            users = cursor.fetchall()
            return jsonify({"users": users})
        except Exception as e:
            logger.error(f"Error fetching users for licensing portal: {e}")
            return jsonify({"error": str(e)}), 500
        finally:
            conn.close()

    @app.route("/admin/license-config")
    @login_required
    @role_required('superadmin')
    def admin_license_config():
        """License configuration page"""
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            # Create table if it doesn't exist
            cursor.execute("CREATE TABLE IF NOT EXISTS license_config (id INT AUTO_INCREMENT PRIMARY KEY, license_key VARCHAR(255) NOT NULL, api_key VARCHAR(255) NOT NULL, licensing_portal_url VARCHAR(255) DEFAULT 'https://feedbacklicensing-production-c938.up.railway.app', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP)")
            cursor.execute("SELECT * FROM license_config ORDER BY id DESC LIMIT 1")
            config = cursor.fetchone()
            
            # Fetch license status from portal if config exists
            license_status = None
            license_error = None
            if config:
                try:
                    import requests
                    portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
                    
                    logger.info(f"Fetching license status from portal: {portal_url}/api/validate/{config['license_key']}")
                    response = requests.post(
                        f"{portal_url}/api/validate/{config['license_key']}",
                        headers=licensing_api_headers(),
                        timeout=10
                    )
                    
                    logger.info(f"License status response status: {response.status_code}")
                    if response.status_code == 200:
                        license_status = response.json()
                        logger.info(f"License status data: {license_status}")
                        if not license_status.get('valid'):
                            license_error = license_status.get('message', 'License validation failed')
                    else:
                        logger.error(f"License validation failed with status: {response.status_code}")
                        license_error = f"API error: {response.status_code}"
                except Exception as e:
                    logger.error(f"Error fetching license status: {e}")
                    license_error = "Unable to reach licensing portal. Please try again later."
            
            return render_template("admin/license_config.html", config=config, license_status=license_status, license_error=license_error)
        finally:
            conn.close()

    @app.route("/admin/license-config/save", methods=["POST"])
    @role_required('superadmin')
    def admin_save_license_config():
        """Save license configuration"""
        license_key = request.form.get("license_key", "").strip()
        api_key = request.form.get("api_key", "").strip()
        licensing_portal_url = normalize_portal_url(request.form.get("licensing_portal_url"))
        
        if not license_key or not api_key:
            flash("License key and API key are required.", "danger")
            return redirect(url_for("admin_license_config"))
        
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Check if config exists
            cursor.execute("SELECT id FROM license_config ORDER BY id DESC LIMIT 1")
            existing = cursor.fetchone()
            
            if existing:
                # Update existing config
                cursor.execute(
                    "UPDATE license_config SET license_key = %s, api_key = %s, licensing_portal_url = %s WHERE id = %s",
                    (license_key, api_key, licensing_portal_url, existing[0])
                )
            else:
                # Insert new config
                cursor.execute(
                    "INSERT INTO license_config (license_key, api_key, licensing_portal_url) VALUES (%s, %s, %s)",
                    (license_key, api_key, licensing_portal_url)
                )
            
            conn.commit()
            conn.close()
            
            flash("License configuration saved successfully.", "success")
            log_audit(
                entity_type="license_config",
                entity_id=0,
                action="updated",
                new_values=f"License configured for portal: {licensing_portal_url}"
            )
        except Exception as e:
            logger.error(f"Error saving license config: {e}")
            flash("Failed to save license configuration.", "danger")
        
        return redirect(url_for("admin_license_config"))

    @app.route("/client/license-config")
    @login_required
    def client_license_config():
        """Deprecated — license management is now part of the Support page."""
        return redirect(url_for("client_support"))

    @app.route("/client/license-config/save", methods=["POST"])
    @login_required
    def client_save_license_config():
        """Save client license configuration"""
        user = get_user_by_id(session['user_id'])
        if not user or user.get('role') != 'admin':
            flash("Only client administrator accounts can configure a license.", "danger")
            return redirect(url_for("client_support"))
        license_key = request.form.get("license_key", "").strip()
        
        if not license_key:
            flash("License key is required.", "danger")
            return redirect(url_for("client_license_config"))
        
        try:
            # Validate license against the licensing portal
            config = get_license_config()
            portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
            
            # Call the licensing portal API to validate the license
            import requests
            try:
                response = requests.post(
                    f"{portal_url}/api/validate/{license_key}",
                    headers=licensing_api_headers(),
                    timeout=10
                )
                
                if response.status_code != 200 or not response.json().get("valid"):
                    flash("Invalid license key. Please check with your administrator.", "danger")
                    return redirect(url_for("client_license_config"))
            except Exception as e:
                logger.error(f"Error validating license with portal: {e}")
                flash("Unable to validate license. Please try again later.", "danger")
                return redirect(url_for("client_license_config"))
            
            # If valid, save to user's account
            conn = get_db_connection()
            cursor = conn.cursor()
            
            cursor.execute(
                "UPDATE users SET license_key = %s WHERE id = %s",
                (license_key, session['user_id'])
            )
            
            conn.commit()
            conn.close()
            
            flash("License key configured successfully. You can now add stores.", "success")
            log_audit(
                entity_type="user",
                entity_id=session['user_id'],
                action="license_configured",
                new_values=f"License key configured for user {session['user_id']}"
            )
        except Exception as e:
            logger.error(f"Error saving client license config: {e}")
            flash("Failed to configure license key.", "danger")
        
        return redirect(url_for("client_license_config"))

    # ── Client Support Portal ──────────────────────────────────────
    def _user_license_key(user: Dict[str, Any]) -> Optional[str]:
        """Return only the license assigned to this client account.

        Never fall back to the system/global license: doing so would make
        superadmins and view-only users share the same support thread.
        """
        if user and user.get('role') == 'admin':
            return user.get('license_key') or None
        return None

    def _support_identity(user: Dict[str, Any]) -> str:
        """Stable, private conversation identity for the signed-in account."""
        return _user_license_key(user) or f"user:{int(user['id'])}"

    @app.route("/client/support")
    @login_required
    def client_support():
        """Client support page — renders instantly, data loads via AJAX"""
        user = get_user_by_id(session['user_id'])
        if user['role'] not in ('user', 'admin', 'superadmin'):
            flash("Access denied.", "danger")
            return redirect(url_for("admin_dashboard"))

        license_key = _user_license_key(user)

        return render_template("client/support.html",
                               user=user, license_key=license_key or '')

    @app.route("/api/support/status")
    @login_required
    def api_support_status():
        """AJAX endpoint — fetch license status + tickets from portal"""
        user = get_user_by_id(session['user_id'])
        config = get_license_config()
        portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
        license_key = _user_license_key(user)

        result = {"license_status": None, "license_error": None, "tickets": [], "renewals": []}

        if not license_key:
            return jsonify(result)

        import requests as http_requests
        # Fetch license status (short timeout)
        try:
            resp = http_requests.post(f"{portal_url}/api/validate/{license_key}", headers=licensing_api_headers(), timeout=5)
            if resp.status_code == 200:
                data = resp.json()
                result["license_status"] = data
                if not data.get('valid'):
                    result["license_error"] = data.get('message', 'License validation failed')
            else:
                result["license_error"] = f"API error: {resp.status_code}"
        except Exception as e:
            logger.error(f"Error fetching license status: {e}")
            result["license_error"] = "Unable to reach licensing portal"

        # Fetch tickets (short timeout, best-effort)
        try:
            resp = http_requests.get(f"{portal_url}/api/tickets/{license_key}", headers=licensing_api_headers(), timeout=5)
            if resp.status_code == 200:
                result["tickets"] = resp.json().get('tickets', [])
        except Exception:
            pass

        try:
            resp = http_requests.get(f"{portal_url}/api/renewals/{license_key}", headers=licensing_api_headers(), timeout=5)
            if resp.status_code == 200:
                result["renewals"] = resp.json().get("renewals", [])
        except Exception:
            pass

        return jsonify(result)

    @app.route("/client/support/ticket", methods=["POST"])
    @login_required
    def client_submit_ticket():
        """Submit a support ticket to the licensing portal"""
        user = get_user_by_id(session['user_id'])
        config = get_license_config()
        portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)

        license_key = _user_license_key(user) or ""
        subject = request.form.get("subject", "").strip()
        message = request.form.get("message", "").strip()
        ticket_type = request.form.get("ticket_type", "general")
        contact_email = user.get('email', '') or user.get('username', '')

        if not subject or not message:
            flash("Subject and message are required.", "danger")
            return redirect(url_for("client_support"))

        try:
            import requests as http_requests
            resp = http_requests.post(f"{portal_url}/api/tickets/create", json={
                "license_key": license_key,
                "contact_email": contact_email,
                "subject": subject,
                "message": message,
                "ticket_type": ticket_type
            }, headers=licensing_api_headers(), timeout=10)
            if resp.status_code in (200, 201):
                flash("Ticket submitted successfully. We'll get back to you soon.", "success")
            else:
                flash("Failed to submit ticket. Please try again.", "danger")
        except Exception as e:
            logger.error(f"Error submitting ticket to portal: {e}")
            flash("Unable to reach support. Please try again later.", "danger")

        return redirect(url_for("client_support"))

    @app.route("/client/support/renew", methods=["POST"])
    @login_required
    def client_request_renewal():
        """Create an Admin-confirmed renewal request for Superadmin review."""
        user = get_user_by_id(session['user_id'])
        config = get_license_config()
        portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)

        license_key = _user_license_key(user) or ""
        contact_email = user.get('email', '') or user.get('username', '')

        if not license_key:
            flash("No license key found.", "danger")
            return redirect(url_for("client_support"))

        if request.form.get("admin_confirmed") != "yes":
            flash("Please confirm that you want to submit a renewal request.", "warning")
            return redirect(url_for("client_support"))

        requested_plan = request.form.get("requested_plan", "Current plan").strip()
        payment_reference = request.form.get("payment_reference", "").strip()
        try:
            requested_days = int(request.form.get("requested_days", "365"))
        except ValueError:
            requested_days = 365

        try:
            import requests as http_requests
            resp = http_requests.post(f"{portal_url}/api/renewals", json={
                "license_key": license_key,
                "contact_email": contact_email,
                "requested_plan": requested_plan,
                "requested_days": requested_days,
                "payment_reference": payment_reference,
                "admin_confirmed": True,
            }, headers=licensing_api_headers(), timeout=10)
            if resp.status_code in (200, 201):
                flash("Renewal request confirmed and sent for Superadmin approval. Your license has not been extended yet.", "success")
            elif resp.status_code == 409:
                flash("You already have a renewal request waiting for Superadmin approval.", "warning")
            else:
                flash("Failed to submit renewal request.", "danger")
        except Exception as e:
            logger.error(f"Error submitting renewal request: {e}")
            flash("Unable to reach support. Please try again later.", "danger")

        return redirect(url_for("client_support"))

    @app.route("/client/support/renew/<int:request_id>/cancel", methods=["POST"])
    @login_required
    def client_cancel_renewal(request_id):
        user = get_user_by_id(session['user_id'])
        license_key = _user_license_key(user) or ""
        try:
            import requests as http_requests
            resp = http_requests.post(f"{_get_portal_url()}/api/renewals/{request_id}/cancel",
                json={"license_key": license_key}, headers=licensing_api_headers(), timeout=10)
            flash("Renewal request cancelled." if resp.status_code == 200 else "This request can no longer be cancelled.",
                  "success" if resp.status_code == 200 else "warning")
        except Exception as exc:
            logger.error("Unable to cancel renewal: %s", exc)
            flash("Unable to cancel the renewal request.", "danger")
        return redirect(url_for("client_support"))

    # ── Client Messaging System (proxies to licensing portal) ────────
    def _get_portal_url():
        config = get_license_config()
        return normalize_portal_url(config.get("licensing_portal_url") if config else None)

    def _ensure_portal_conversation(client_identifier, license_key, contact_email, company_name=""):
        """Ensure conversation exists on portal and return its ID"""
        # Fallback contact_email if empty (portal requires it)
        effective_email = contact_email or company_name or client_identifier or "client@unknown"
        portal_url = _get_portal_url()
        import requests as http_requests
        try:
            logger.info(f"Ensuring portal conversation at {portal_url} for {client_identifier}")
            resp = http_requests.post(f"{portal_url}/api/conversations/create", json={
                "client_identifier": client_identifier,
                "license_key": license_key or "",
                "contact_email": effective_email,
                "company_name": company_name or effective_email
            }, headers=licensing_api_headers(), timeout=10)
            if resp.status_code in (200, 201):
                conv_id = resp.json().get('conversation', {}).get('id')
                logger.info(f"Portal conversation ID: {conv_id}")
                return conv_id
            logger.error(f"Portal create failed: {resp.status_code} - {resp.text}")
        except Exception as e:
            logger.error(f"Failed to ensure portal conversation: {e}")
        return None

    @app.route("/api/client/messages", methods=["GET"])
    @login_required
    def api_get_client_messages():
        """Get messages for the current user from licensing portal"""
        user = get_user_by_id(session['user_id'])
        license_key = _user_license_key(user)
        client_identifier = _support_identity(user)
        contact_email = user.get('email', '') or user.get('username', '')

        if not license_key and not contact_email:
            return jsonify({"error": "No license key or email found"}), 400

        # Ensure conversation exists on portal
        conv_id = _ensure_portal_conversation(client_identifier, license_key, contact_email, user.get('username', ''))
        if not conv_id:
            return jsonify({"messages": [], "conversation_id": None, "error": "Portal unavailable"})

        # Fetch messages from portal
        portal_url = _get_portal_url()
        try:
            import requests as http_requests
            resp = http_requests.get(f"{portal_url}/api/conversations/{conv_id}/messages?viewer=client", headers=licensing_api_headers(), timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                return jsonify({"messages": data.get('messages', []), "conversation_id": conv_id})
        except Exception as e:
            logger.error(f"Failed to fetch messages from portal: {e}")
        return jsonify({"messages": [], "conversation_id": conv_id})

    @app.route("/api/client/messages/send", methods=["POST"])
    @login_required
    def api_send_client_message():
        """Send a client message directly to licensing portal"""
        user = get_user_by_id(session['user_id'])
        license_key = _user_license_key(user)
        client_identifier = _support_identity(user)
        contact_email = user.get('email', '') or user.get('username', '')

        data = request.get_json() or {}
        message = data.get("message", "").strip()
        if not message:
            return jsonify({"error": "Message is required"}), 400

        # Ensure conversation exists on portal
        conv_id = _ensure_portal_conversation(client_identifier, license_key, contact_email, user.get('username', ''))
        if not conv_id:
            return jsonify({"error": "Failed to reach licensing portal"}), 500

        # Send message to portal
        portal_url = _get_portal_url()
        try:
            import requests as http_requests
            resp = http_requests.post(f"{portal_url}/api/conversations/{conv_id}/send", json={
                "message": message,
                "sender_type": "client",
                "sender_name": contact_email or user.get('username', 'Client')
            }, headers=licensing_api_headers(), timeout=10)
            if resp.status_code in (200, 201):
                return jsonify({"success": True})
            logger.error(f"Failed to send message to portal: {resp.status_code} - {resp.text}")
            return jsonify({"error": "Failed to send message"}), 500
        except Exception as e:
            logger.error(f"Error sending message to portal: {e}")
            return jsonify({"error": "Failed to send message"}), 500

    # ── Admin Messaging Interface ───────────────────────────────────────
    @app.route("/admin/messages")
    @login_required
    @role_required('superadmin')
    def admin_messages():
        """Admin messages page - view all client conversations from licensing portal"""
        # Fetch conversations from licensing portal
        config = get_license_config()
        portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
        conversations = []
        try:
            import requests as http_requests
            resp = http_requests.get(f"{portal_url}/api/conversations", headers=licensing_api_headers(), timeout=5)
            if resp.status_code == 200:
                conversations = resp.json().get('conversations', [])
        except Exception as e:
            logger.error(f"Failed to fetch conversations from portal: {e}")
        return render_template("admin/messages.html", conversations=conversations)

    @app.route("/api/admin/conversations")
    @login_required
    @role_required('superadmin')
    def api_admin_get_conversations():
        """API endpoint to get all conversations from licensing portal"""
        config = get_license_config()
        portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
        try:
            import requests as http_requests
            logger.info(f"Fetching conversations from portal at {portal_url}")
            resp = http_requests.get(f"{portal_url}/api/conversations", headers=licensing_api_headers(), timeout=5)
            if resp.status_code == 200:
                logger.info(f"Successfully fetched {len(resp.json().get('conversations', []))} conversations from portal")
                return jsonify(resp.json())
            logger.error(f"Failed to fetch conversations: {resp.status_code}")
            return jsonify({"conversations": []})
        except Exception as e:
            logger.error(f"Failed to fetch conversations from portal: {e}")
            return jsonify({"conversations": []})

    @app.route("/api/admin/conversations/<int:conversation_id>/messages")
    @login_required
    @role_required('superadmin')
    def api_admin_get_conversation_messages(conversation_id):
        """API endpoint to get messages for a conversation from licensing portal"""
        config = get_license_config()
        portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
        try:
            import requests as http_requests
            resp = http_requests.get(f"{portal_url}/api/conversations/{conversation_id}/messages", headers=licensing_api_headers(), timeout=5)
            if resp.status_code == 200:
                return jsonify(resp.json())
            return jsonify({"messages": []})
        except Exception as e:
            logger.error(f"Failed to fetch messages from portal: {e}")
            return jsonify({"messages": []})

    @app.route("/api/admin/conversations/<int:conversation_id>/send", methods=["POST"])
    @login_required
    @role_required('superadmin')
    def api_admin_send_message(conversation_id):
        """API endpoint to send a message as admin to licensing portal"""
        data = request.get_json() or {}
        message = data.get("message", "").strip()
        if not message:
            return jsonify({"error": "Message is required"}), 400

        config = get_license_config()
        portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
        try:
            import requests as http_requests
            resp = http_requests.post(f"{portal_url}/api/conversations/{conversation_id}/send", json={
                "message": message,
                "sender_type": "admin",
                "sender_name": "Support Team"
            }, headers=licensing_api_headers(), timeout=5)
            if resp.status_code in (200, 201):
                return jsonify({"success": True})
            return jsonify({"error": "Failed to send message"}), 500
        except Exception as e:
            logger.error(f"Error sending message to portal: {e}")
            return jsonify({"error": "Failed to send message"}), 500

    @app.route("/api/messages/unread-count")
    @login_required
    def api_message_unread_count():
        """Return the private message unread total for the signed-in account."""
        user = get_user_by_id(session["user_id"])
        portal_url = _get_portal_url()
        try:
            import requests as http_requests
            if user.get("role") == "superadmin":
                resp = http_requests.get(f"{portal_url}/api/conversations", headers=licensing_api_headers(), timeout=5)
                conversations = resp.json().get("conversations", []) if resp.status_code == 200 else []
                return jsonify({"success": True, "count": sum(int(c.get("unread_count") or 0) for c in conversations)})

            license_key = _user_license_key(user)
            contact_email = user.get("email", "") or user.get("username", "")
            conv_id = _ensure_portal_conversation(_support_identity(user), license_key, contact_email, user.get("username", ""))
            if not conv_id:
                return jsonify({"success": True, "count": 0})
            resp = http_requests.get(f"{portal_url}/api/conversations/{conv_id}/messages?viewer=count", headers=licensing_api_headers(), timeout=5)
            messages = resp.json().get("messages", []) if resp.status_code == 200 else []
            unread = sum(1 for m in messages if m.get("sender_type") == "admin" and not bool(m.get("is_read", m.get("seen", False))))
            return jsonify({"success": True, "count": unread})
        except Exception as exc:
            logger.warning("Unable to fetch message unread count: %s", exc)
            return jsonify({"success": True, "count": 0})

    @app.route("/admin/reset-database", methods=["POST"])
    @role_required('superadmin')
    def admin_reset_database():
        """Reset all data in the database (keeps users)"""
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Disable foreign key checks temporarily
            cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
            
            # Delete all data from tables (keeping users and schema)
            tables_to_clear = [
                "feedback",
                "stores",
                "questionnaires",
                "questions",
                "audit_logs",
                "notifications"
            ]
            
            for table in tables_to_clear:
                try:
                    cursor.execute(f"DELETE FROM {table}")
                except Exception as e:
                    logger.warning(f"Could not clear table {table}: {e}")
            
            # Re-enable foreign key checks
            cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
            
            conn.commit()
            conn.close()
            
            flash("Database reset successfully. All data cleared except users.", "success")
            log_audit(
                entity_type="database",
                entity_id=0,
                action="reset",
                old_values="Database reset"
            )
        except Exception as e:
            logger.error(f"Error resetting database: {e}")
            flash(f"Failed to reset database: {e}", "danger")
        
        return redirect(url_for("admin_users"))

    @app.route("/dashboard/staff-overall")
    def staff_overall():
        """Staff overall page showing all staff with ratings and performance metrics."""
        try:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)

            # Global average commendation rating (prior `m`); fall back to 4.0.
            cursor.execute(
                "SELECT AVG(rating) as global_avg FROM staff_commendations WHERE rating IS NOT NULL"
            )
            row = cursor.fetchone()
            global_avg_rating = float(row['global_avg']) if row and row['global_avg'] is not None else 4.0

            # Fetch all staff with their commendation ratings and metrics.
            # `weighted_score` is the Bayesian average:
            #   (C * m + sum_of_ratings) / (C + n)
            cursor.execute(
                """
                SELECT s.id, s.first_name, s.last_name, s.email, s.phone, s.position, s.role, s.status, st.store_name, s.store_id,
                       AVG(sc.rating) as avg_rating,
                       COUNT(sc.id) as commendation_count,
                       (
                           (%s * %s) + COALESCE(SUM(sc.rating), 0)
                       ) / (
                           %s + COUNT(sc.rating)
                       ) as weighted_score
                FROM staff s
                LEFT JOIN staff_commendations sc ON s.id = sc.staff_id
                LEFT JOIN stores st ON s.store_id = st.id
                GROUP BY s.id, s.first_name, s.last_name, s.email, s.phone, s.position, s.role, s.status, st.store_name, s.store_id
                ORDER BY weighted_score DESC, s.last_name, s.first_name
                """,
                (BAYESIAN_C, global_avg_rating, BAYESIAN_C),
            )
            staff_data = cursor.fetchall()
            
            # Format the data
            for staff in staff_data:
                staff['avg_rating'] = float(staff['avg_rating']) if staff['avg_rating'] else 0.0
                staff['commendation_count'] = int(staff['commendation_count']) if staff['commendation_count'] else 0
                staff['weighted_score'] = float(staff['weighted_score']) if staff['weighted_score'] else 0.0
            
            conn.close()
            return render_template("dashboard/staff_overall.html", staff_data=staff_data)
        except Exception as e:
            error_details = traceback.format_exc()
            logger.error(f"Staff Overall Error: {e}\n{error_details}")
            return f"Staff Overall Error: {e}<br><pre>{error_details}</pre>", 500

    @app.route("/api/dashboard/analytics")
    def api_dashboard_analytics():
        """JSON endpoint for overall dashboard analytics (used by store filter)."""
        try:
            analytics = fetch_dashboard_analytics()
            # Serialize for JSON
            stores_data = analytics.get('stores_data', [])
            overall = analytics.get('overall_stats', {})
            recent = analytics.get('recent_activity', [])
            top = analytics.get('top_stores', [])
            best_store = analytics.get('best_overall_store')
            best_staff = analytics.get('best_overall_staff')

            # Format recent_activity dates
            formatted_activity = []
            for a in recent:
                d = a.get('date')
                formatted_activity.append({
                    'date_label': d.strftime('%b %d') if d else '?',
                    'responses': a.get('responses', 0)
                })

            return jsonify({
                'stores_data': [
                    {
                        'id': s['id'],
                        'store_name': s['store_name'],
                        'address': s.get('address', ''),
                        'city': s.get('city', ''),
                        'total_responses': s.get('total_responses', 0),
                        'avg_rating': float(s['avg_rating']) if s.get('avg_rating') else 0.0,
                        'unique_users': s.get('unique_users', 0)
                    } for s in stores_data
                ],
                'overall_stats': {
                    'total_responses': overall.get('total_responses', 0),
                    'total_stores': overall.get('total_stores', 0),
                    'total_unique_users': overall.get('total_unique_users', 0),
                    'overall_avg_rating': float(overall.get('overall_avg_rating', 0)),
                    'total_questionnaires': overall.get('total_questionnaires', 0)
                },
                'recent_activity': formatted_activity,
                'top_stores': [
                    {'store_name': t['store_name'], 'response_count': t['response_count']}
                    for t in top
                ],
                'best_overall_store': {
                    'store_name': best_store['store_name'],
                    'avg_rating': float(best_store['avg_rating'])
                } if best_store else None,
                'best_overall_staff': {
                    'first_name': best_staff['first_name'],
                    'last_name': best_staff['last_name'],
                    'avg_rating': float(best_staff['avg_rating']) if best_staff.get('avg_rating') else 0.0
                } if best_staff else None
            })
        except Exception as e:
            logger.error(f"Dashboard API error: {e}")
            return jsonify({'error': str(e)}), 500

    @app.route("/admin/stores/performance")
    def stores_performance():
        analytics = fetch_dashboard_analytics()
        return render_template(
            "dashboard/store_performance.html",
            stores_data=analytics.get("stores_data", []),
            overall_stats=analytics.get("overall_stats", {}),
        )

    @app.route("/admin/stores", methods=["GET"])
    @login_required
    def stores_management():
        user = get_user_by_id(session['user_id'])
        # Role-based store visibility:
        #   superadmin -> all stores
        #   admin (client) -> own stores (filtered by user_id ownership)
        #   user (view-only) -> only stores explicitly assigned via user_stores
        if user['role'] == 'user':
            assigned_ids = get_assigned_store_ids(session['user_id'])
            logger.info(f"View-only user {session['user_id']} assigned to stores: {assigned_ids}")
            stores = fetch_stores(assigned_store_ids=assigned_ids)
            user_id = session['user_id']
        else:
            user_id = session['user_id'] if user['role'] == 'admin' else None
            logger.info(f"User {session['user_id']} (role: {user['role']}) viewing stores. Filtering by user_id: {user_id}")
            stores = fetch_stores(user_id=user_id)
        logger.info(f"User {session['user_id']} sees {len(stores)} stores")

        # Batch-load feedback + staff counts and (for non-clients) user info in 3 queries
        # instead of running 2 queries per store + 1 per user.
        store_ids = [s["id"] for s in stores]
        feedback_counts = {}
        staff_counts = {}
        user_info = {}
        if store_ids:
            conn = get_db_connection()
            try:
                cursor = conn.cursor()
                placeholders = ",".join(["%s"] * len(store_ids))
                cursor.execute(
                    f"SELECT store_id, COUNT(*) FROM responses WHERE store_id IN ({placeholders}) GROUP BY store_id",
                    tuple(store_ids),
                )
                feedback_counts = {row[0]: int(row[1]) for row in cursor.fetchall()}
                cursor.execute(
                    f"SELECT store_id, COUNT(*) FROM staff WHERE store_id IN ({placeholders}) GROUP BY store_id",
                    tuple(store_ids),
                )
                staff_counts = {row[0]: int(row[1]) for row in cursor.fetchall()}

                if user['role'] == 'superadmin':
                    user_ids = sorted({s.get("user_id") for s in stores if s.get("user_id")})
                    if user_ids:
                        uph = ",".join(["%s"] * len(user_ids))
                        cursor2 = conn.cursor(dictionary=True)
                        cursor2.execute(
                            f"SELECT id, username, role FROM users WHERE id IN ({uph})",
                            tuple(user_ids),
                        )
                        user_info = {u["id"]: u for u in cursor2.fetchall()}
            finally:
                conn.close()

        # Enhance stores with counts (single pass)
        enhanced_stores = []
        for store in stores:
            sid = store["id"]
            store_with_counts = dict(store)
            store_with_counts["feedback_count"] = feedback_counts.get(sid, 0)
            store_with_counts["staff_count"] = staff_counts.get(sid, 0)
            enhanced_stores.append(store_with_counts)

        # Group enhanced stores by client (admin) for superadmin only
        stores_by_user_enhanced = None
        if user['role'] == 'superadmin' and enhanced_stores:
            stores_by_user_enhanced = {}
            for store in enhanced_stores:
                uid = store.get("user_id") or "unassigned"
                stores_by_user_enhanced.setdefault(uid, []).append(store)

        selected_store_id_param = request.args.get("store_id")
        selected_store_id = None
        if selected_store_id_param:
            try:
                selected_store_id = int(selected_store_id_param)
            except ValueError:
                selected_store_id = None

        selected_store = None
        if selected_store_id is not None:
            for store in stores:
                if store["id"] == selected_store_id:
                    selected_store = store
                    break

        public_url = None
        qr_data_uri = None
        if selected_store:
            public_url = get_store_public_url(store_id=int(selected_store["id"]))
            qr_data_uri = generate_qr_data_uri(public_url)

        return render_template(
            "manage_stores/stores.html",
            stores=enhanced_stores if user['role'] != 'superadmin' else None,
            all_stores=enhanced_stores,
            stores_by_user=stores_by_user_enhanced if user['role'] == 'superadmin' else None,
            user_info=user_info if user['role'] == 'superadmin' else None,
            selected_store=selected_store,
            public_url=public_url,
            qr_data_uri=qr_data_uri,
        )

    def get_feedback_count_for_store(store_id: int) -> int:
        """Get the total number of feedback responses for a store."""
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT COUNT(*) FROM responses WHERE store_id = %s",
                (store_id,)
            )
            count = cursor.fetchone()[0]
            return int(count) if count else 0
        finally:
            conn.close()

    def get_staff_count_for_store(store_id: int) -> int:
        """Get the total number of staff members for a store."""
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT COUNT(*) FROM staff WHERE store_id = %s",
                (store_id,)
            )
            count = cursor.fetchone()[0]
            return int(count) if count else 0
        finally:
            conn.close()

    def get_staff_performance_for_store(store_id: int) -> List[Dict[str, Any]]:
        """Get staff for a store ranked by Bayesian-average score.

        Score = (C * m + sum_of_ratings) / (C + n), where m is the global
        commendation rating and C is the smoothing constant (`BAYESIAN_C`).
        """
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)

            cursor.execute(
                "SELECT AVG(rating) as global_avg FROM staff_commendations WHERE rating IS NOT NULL"
            )
            row = cursor.fetchone()
            global_avg_rating = float(row['global_avg']) if row and row['global_avg'] is not None else 4.0

            cursor.execute(
                """
                SELECT s.id, s.first_name, s.last_name, s.position, s.role,
                       AVG(sc.rating) as avg_rating,
                       COUNT(sc.id) as commendation_count,
                       (
                           (%s * %s) + COALESCE(SUM(sc.rating), 0)
                       ) / (
                           %s + COUNT(sc.rating)
                       ) as weighted_score
                FROM staff s
                LEFT JOIN staff_commendations sc ON s.id = sc.staff_id
                WHERE s.store_id = %s
                GROUP BY s.id, s.first_name, s.last_name, s.position, s.role
                ORDER BY weighted_score DESC
                """,
                (BAYESIAN_C, global_avg_rating, BAYESIAN_C, store_id),
            )
            return cursor.fetchall()
        finally:
            conn.close()

    # API endpoint for store feedback data
    @app.route("/api/stores/<int:store_id>/feedback", methods=["GET"])
    @login_required
    def api_store_feedback(store_id: int):
        """API endpoint to get feedback data for a store."""
        if not can_manage_store_staff(session['user_id'], store_id):
            return jsonify({"error": "You can only view your assigned store."}), 403
        store = fetch_store_by_id(store_id=store_id)
        if not store:
            return jsonify({"error": "Store not found"}), 404
        
        feedback = fetch_responses_for_store(store_id=store_id, limit=5)
        return jsonify(feedback)

    # API endpoint for store analytics data
    @app.route("/api/stores/<int:store_id>/analytics", methods=["GET"])
    @login_required
    def api_store_analytics(store_id: int):
        """API endpoint to get analytics data for a store."""
        if not can_manage_store_staff(session['user_id'], store_id):
            return jsonify({"error": "You can only view your assigned store."}), 403
        store = fetch_store_by_id(store_id=store_id)
        if not store:
            return jsonify({"error": "Store not found"}), 404
        
        # Fetch all feedback for analytics
        all_feedback = fetch_responses_for_store(store_id=store_id, limit=1000)
        total_feedback = len(all_feedback)
        
        # Resolved / unresolved counts
        resolved_count = sum(1 for f in all_feedback if f.get("status") == "resolved")
        unresolved_count = total_feedback - resolved_count
        resolution_rate = round((resolved_count / total_feedback * 100), 1) if total_feedback > 0 else 0
        
        # Calculate ratings
        all_response_ids = [int(r["id"]) for r in all_feedback]
        answers_by_response_id = fetch_answers_for_responses(all_response_ids) if all_feedback else {}
        
        # Rating distribution
        rating_distribution = [0, 0, 0, 0, 0]  # 1-5 stars
        total_ratings = 0
        for response_id, answers in answers_by_response_id.items():
            for answer in answers:
                if answer.get("rating_value"):
                    rating = int(float(answer["rating_value"]))
                    if 1 <= rating <= 5:
                        rating_distribution[rating - 1] += 1
                        total_ratings += 1
        
        # Calculate percentages
        five_star_count = rating_distribution[4]
        four_star_count = rating_distribution[3]
        
        five_star_rate = round((five_star_count / total_ratings * 100), 1) if total_ratings > 0 else 0
        four_plus_star_rate = round(((four_star_count + five_star_count) / total_ratings * 100), 1) if total_ratings > 0 else 0
        
        # Rating distribution percentages
        rating_pcts = [round(c / total_ratings * 100, 1) if total_ratings > 0 else 0 for c in rating_distribution]
        
        # Quality score
        quality_score = round(
            (rating_distribution[0] * 1 + rating_distribution[1] * 2 + 
             rating_distribution[2] * 3 + rating_distribution[3] * 4 + 
             rating_distribution[4] * 5) / total_ratings, 1
        ) if total_ratings > 0 else 0
        
        # Monthly feedback trend (last 6 months)
        monthly_trend = defaultdict(int)
        now = datetime.now()
        for fb in all_feedback:
            submitted = fb.get("submitted_at")
            if submitted:
                if isinstance(submitted, str):
                    try:
                        submitted = datetime.strptime(submitted, "%Y-%m-%d %H:%M:%S")
                    except (ValueError, TypeError):
                        continue
                key = submitted.strftime("%Y-%m")
                monthly_trend[key] += 1
        
        # Build last 6 months labels and values
        trend_labels = []
        trend_values = []
        for i in range(5, -1, -1):
            d = now - timedelta(days=i * 30)
            key = d.strftime("%Y-%m")
            label = d.strftime("%b")
            trend_labels.append(label)
            trend_values.append(monthly_trend.get(key, 0))
        
        # Staff commendations count + top staff
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            if all_response_ids:
                placeholders = ','.join(['%s'] * len(all_response_ids))
                cursor.execute(f"""
                    SELECT COUNT(*) as cnt FROM staff_commendations 
                    WHERE response_id IN ({placeholders})
                """, all_response_ids)
                total_commendations = cursor.fetchone()["cnt"]
                
                # Global average commendation rating (Bayesian prior `m`).
                cursor.execute(
                    "SELECT AVG(rating) as global_avg FROM staff_commendations WHERE rating IS NOT NULL"
                )
                grow = cursor.fetchone()
                staff_global_avg = float(grow['global_avg']) if grow and grow['global_avg'] is not None else 4.0

                # Top 5 commended staff ranked by Bayesian-average score.
                cursor.execute(f"""
                    SELECT s.first_name, s.last_name, s.position, s.role,
                           AVG(sc.rating) as avg_rating,
                           COUNT(sc.id) as commendation_count,
                           (
                               (%s * %s) + COALESCE(SUM(sc.rating), 0)
                           ) / (
                               %s + COUNT(sc.rating)
                           ) as weighted_score
                    FROM staff_commendations sc
                    JOIN staff s ON s.id = sc.staff_id
                    WHERE sc.response_id IN ({placeholders})
                    GROUP BY s.id, s.first_name, s.last_name, s.position, s.role
                    ORDER BY weighted_score DESC
                    LIMIT 5
                """, [BAYESIAN_C, staff_global_avg, BAYESIAN_C, *all_response_ids])
                top_staff = cursor.fetchall()
            else:
                total_commendations = 0
                top_staff = []
        finally:
            conn.close()
        
        formatted_top_staff = []
        for s in top_staff:
            formatted_top_staff.append({
                "name": f"{s['first_name']} {s['last_name']}",
                "position": s["position"] or (s["role"].title() if s["role"] else "Staff"),
                "avg_rating": float(s["avg_rating"]) if s["avg_rating"] else 0.0
            })
        
        return jsonify({
            "overview": {
                "total_feedback": total_feedback,
                "resolved": resolved_count,
                "unresolved": unresolved_count,
                "resolution_rate": resolution_rate,
                "total_ratings": total_ratings
            },
            "rating_metrics": {
                "five_star_rate": five_star_rate,
                "four_plus_star_rate": four_plus_star_rate,
                "quality_score": quality_score,
                "distribution": rating_distribution,
                "distribution_pcts": rating_pcts
            },
            "trend": {
                "labels": trend_labels,
                "values": trend_values
            },
            "staff_metrics": {
                "total_commendations": total_commendations,
                "top_staff": formatted_top_staff
            }
        })

    # API endpoint for store staff data
    @app.route("/api/stores/<int:store_id>/staff", methods=["GET"])
    @login_required
    def api_store_staff(store_id: int):
        """API endpoint to get staff data for a store."""
        if not can_manage_store_staff(session['user_id'], store_id):
            return jsonify({"error": "You can only view staff in your assigned store."}), 403
        store = fetch_store_by_id(store_id=store_id)
        if not store:
            return jsonify({"error": "Store not found"}), 404
        
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            
            # Global average commendation rating (Bayesian prior `m`).
            cursor.execute(
                "SELECT AVG(rating) as global_avg FROM staff_commendations WHERE rating IS NOT NULL"
            )
            row = cursor.fetchone()
            global_avg_rating = float(row['global_avg']) if row and row['global_avg'] is not None else 4.0

            # Fetch staff with commendation ratings; `weighted_score` uses the
            # Bayesian average so low-volume staff don't outrank high-volume
            # staff just because of a single 5★ commendation.
            cursor.execute(
                """
                SELECT s.id, s.first_name, s.last_name, s.email, s.phone, s.position, s.photo_url, s.role, s.status,
                       AVG(sc.rating) as avg_rating,
                       COUNT(sc.id) as commendation_count,
                       (
                           (%s * %s) + COALESCE(SUM(sc.rating), 0)
                       ) / (
                           %s + COUNT(sc.rating)
                       ) as weighted_score
                FROM staff s
                LEFT JOIN staff_commendations sc ON s.id = sc.staff_id
                WHERE s.store_id = %s
                GROUP BY s.id
                ORDER BY weighted_score DESC, s.last_name, s.first_name
                """,
                (BAYESIAN_C, global_avg_rating, BAYESIAN_C, store_id),
            )
            
            staff_members = cursor.fetchall()
            
            # Format staff data
            total_commendations = sum(s["commendation_count"] or 0 for s in staff_members) if staff_members else 0
            max_commendations = max((s["commendation_count"] or 0 for s in staff_members), default=0)
            formatted_staff = []
            for staff in staff_members:
                formatted_staff.append({
                    "id": staff["id"],
                    "name": f"{staff['first_name']} {staff['last_name']}",
                    "first_name": staff["first_name"],
                    "last_name": staff["last_name"],
                    "position": staff["position"] or staff["role"].title(),
                    "email": staff.get("email", "") or "",
                    "phone": staff.get("phone", "") or "",
                    "photo_url": staff.get("photo_url", "") or "",
                    "role": staff["role"],
                    "status": staff["status"],
                    "avg_rating": float(staff["avg_rating"]) if staff["avg_rating"] else 0.0,
                    "commendation_count": staff["commendation_count"] or 0,
                    "store_id": store_id
                })
            
            return jsonify(formatted_staff)
        finally:
            conn.close()

    # -------------------------
    # PUBLIC SURVEY
    # -------------------------
    @app.route("/dashboard", methods=["GET"])
    def public_store_dashboard_subdomain():
        # Extract subdomain from request host
        host = request.host.split(':')[0]  # Remove port if present
        parts = host.split('.')
        
        # Check if accessing via subdomain
        if len(parts) >= 3:
            subdomain = parts[0]
            
            # Validate subdomain and fetch store
            conn = get_db_connection()
            try:
                cursor = conn.cursor(dictionary=True)
                cursor.execute(
                    """
                    SELECT id, store_name, address, city, province, postal_code,
                           contact_number, email, store_manager_name, manager_contact,
                           store_type, status, logo_url
                    FROM stores
                    WHERE subdomain = %s
                    LIMIT 1
                    """,
                    (subdomain,)
                )
                store = cursor.fetchone()
            finally:
                conn.close()

            if store:
                # Fetch store performance data
                conn = get_db_connection()
                try:
                    cursor = conn.cursor(dictionary=True)

                    # Total feedback count
                    cursor.execute("SELECT COUNT(*) as total FROM responses WHERE store_id = %s", (store['id'],))
                    total_feedback = cursor.fetchone()['total']

                    # Average rating
                    cursor.execute("SELECT AVG(a.rating_value) as avg_rating FROM answers a JOIN responses r ON a.response_id = r.id WHERE r.store_id = %s AND a.rating_value IS NOT NULL", (store['id'],))
                    avg_rating = cursor.fetchone()['avg_rating']

                    # Total commendations
                    cursor.execute("SELECT COUNT(*) as total FROM staff_commendations sc JOIN responses r ON sc.response_id = r.id WHERE r.store_id = %s", (store['id'],))
                    total_commendations = cursor.fetchone()['total']

                    # Total staff
                    cursor.execute("SELECT COUNT(*) as total FROM staff WHERE store_id = %s", (store['id'],))
                    total_staff = cursor.fetchone()['total']

                    # Staff performance
                    cursor.execute(
                        """
                        SELECT s.id, s.first_name, s.last_name, s.position, s.role, s.status,
                               AVG(sc.rating) as avg_rating,
                               COUNT(sc.id) as commendation_count
                        FROM staff s
                        LEFT JOIN staff_commendations sc ON s.id = sc.staff_id
                        LEFT JOIN responses r ON sc.response_id = r.id
                        WHERE s.store_id = %s AND r.store_id = %s
                        GROUP BY s.id, s.first_name, s.last_name, s.position, s.role, s.status
                        ORDER BY avg_rating DESC
                        """,
                        (store['id'], store['id'])
                    )
                    staff_performance = cursor.fetchall()

                    # Recent feedback
                    cursor.execute(
                        """
                        SELECT a.rating_value as rating, a.answer_text as comment, r.created_at
                        FROM responses r
                        LEFT JOIN answers a ON r.id = a.response_id
                        WHERE r.store_id = %s
                        ORDER BY r.created_at DESC
                        LIMIT 10
                        """,
                        (store['id'],)
                    )
                    recent_feedback = cursor.fetchall()

                    # Fetch master questionnaire logo
                    cursor.execute("SELECT logo_url FROM questionnaires WHERE is_template = 1 AND owner_user_id = %s AND license_key <=> %s LIMIT 1", (store['user_id'], store.get('license_key')))
                    master_logo = cursor.fetchone()

                finally:
                    conn.close()

                return render_template(
                    "public/store_dashboard.html",
                    store=store,
                    master_logo=master_logo.get('logo_url') if master_logo else None,
                    total_feedback=total_feedback,
                    avg_rating=avg_rating,
                    total_commendations=total_commendations,
                    total_staff=total_staff,
                    staff_performance=staff_performance,
                    recent_feedback=recent_feedback
                )
        
        # If no subdomain match, redirect to main domain or show error
        return render_template("layout.html", error="Store not found or invalid subdomain"), 404

    @app.route("/d/<access_token>", methods=["GET"])
    def public_store_dashboard(access_token: str):
        # Validate access token and fetch store
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT id, store_name, address, city, province, postal_code,
                       contact_number, email, store_manager_name, manager_contact,
                       store_type, status, logo_url
                FROM stores
                WHERE access_token = %s
                LIMIT 1
                """,
                (access_token,)
            )
            store = cursor.fetchone()
        finally:
            conn.close()

        if not store:
            return render_template("layout.html", error="Invalid access token or store not found"), 404

        # Fetch store performance data
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)

            # Total feedback count
            cursor.execute("SELECT COUNT(*) as total FROM responses WHERE store_id = %s", (store['id'],))
            total_feedback = cursor.fetchone()['total']

            # Average rating
            cursor.execute("SELECT AVG(a.rating_value) as avg_rating FROM answers a JOIN responses r ON a.response_id = r.id WHERE r.store_id = %s AND a.rating_value IS NOT NULL", (store['id'],))
            avg_rating = cursor.fetchone()['avg_rating']

            # Total commendations
            cursor.execute("SELECT COUNT(*) as total FROM staff_commendations sc JOIN responses r ON sc.response_id = r.id WHERE r.store_id = %s", (store['id'],))
            total_commendations = cursor.fetchone()['total']

            # Total staff
            cursor.execute("SELECT COUNT(*) as total FROM staff WHERE store_id = %s", (store['id'],))
            total_staff = cursor.fetchone()['total']

            # Staff performance
            cursor.execute(
                """
                SELECT s.id, s.first_name, s.last_name, s.position, s.role, s.status,
                       AVG(sc.rating) as avg_rating,
                       COUNT(sc.id) as commendation_count
                FROM staff s
                LEFT JOIN staff_commendations sc ON s.id = sc.staff_id
                LEFT JOIN responses r ON sc.response_id = r.id
                WHERE s.store_id = %s AND r.store_id = %s
                GROUP BY s.id, s.first_name, s.last_name, s.position, s.role, s.status
                ORDER BY avg_rating DESC
                """,
                (store['id'], store['id'])
            )
            staff_performance = cursor.fetchall()

            # Recent feedback
            cursor.execute(
                """
                SELECT a.rating_value as rating, a.answer_text as comment, r.created_at
                FROM responses r
                LEFT JOIN answers a ON r.id = a.response_id
                WHERE r.store_id = %s
                ORDER BY r.created_at DESC
                LIMIT 10
                """,
                (store['id'],)
            )
            recent_feedback = cursor.fetchall()

            # Fetch master questionnaire logo
            cursor.execute("SELECT logo_url FROM questionnaires WHERE is_template = 1 AND owner_user_id = %s AND license_key <=> %s LIMIT 1", (store['user_id'], store.get('license_key')))
            master_logo = cursor.fetchone()

        finally:
            conn.close()

        return render_template(
            "public/store_dashboard.html",
            store=store,
            master_logo=master_logo.get('logo_url') if master_logo else None,
            total_feedback=total_feedback,
            avg_rating=avg_rating,
            total_commendations=total_commendations,
            total_staff=total_staff,
            staff_performance=staff_performance,
            recent_feedback=recent_feedback
        )

    @app.route("/s/<int:store_id>", methods=["GET"])
    def public_survey(store_id: int):
        store = fetch_store_by_id(store_id=store_id)
        if not store:
            return render_template("survey_error.html", store=None, error="Page not found"), 404

        # Check if master questionnaire is active
        master_template = fetch_template_questionnaire(int(store['user_id']), store.get('license_key'))
        if not master_template or not master_template.get("is_active"):
            return render_template("survey_error.html", store=store, error="Sorry, the system is not accepting any feedbacks right now"), 404

        questionnaire = fetch_questionnaire_by_store(store_id=store_id)
        if not questionnaire:
            return render_template(
                "survey_error.html", store=store,
                error="This store does not have a published questionnaire yet."
            ), 404
        if not questionnaire.get("is_active"):
            return render_template("survey_error.html", store=store, error="Sorry, the system is not accepting any feedbacks right now"), 404

        questions = fetch_questions_for_questionnaire(questionnaire_id=int(questionnaire["id"]))
        question_ids = [int(q["id"]) for q in questions]
        options_by_question_id = fetch_options_for_questions(question_ids=question_ids)

        # The restaurant/store logo always wins over template branding.
        master_logo = store.get('logo_url') or questionnaire.get('logo_url') or master_template.get('logo_url')
        branding = fetch_tenant_branding(int(store['user_id']), store.get('license_key'))

        # Fetch active staff for this store
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, first_name, last_name, position, role, photo_url
            FROM staff
            WHERE store_id = %s AND status = 'active'
            ORDER BY role DESC, last_name, first_name
        """, (store_id,))
        staff_members = cursor.fetchall()
        cursor.close()
        conn.close()

        return render_template(
            "master_questionnaire/survey.html",
            store=store,
            master_logo=master_logo,
            branding=branding,
            questionnaire=questionnaire,
            questions=questions,
            options_by_question_id=options_by_question_id,
            staff_members=staff_members,
            staff_photo_map={str(member['id']): member.get('photo_url') for member in staff_members if member.get('photo_url')},
        )

    @app.route("/s/<int:store_id>/submit", methods=["POST"])
    def submit_survey(store_id: int):
        store = fetch_store_by_id(store_id=store_id)
        if not store:
            return render_template("survey_error.html", store=None, error="Page not found"), 404

        # Check if master questionnaire is active
        master_template = fetch_template_questionnaire(int(store['user_id']), store.get('license_key'))
        if not master_template or not master_template.get("is_active"):
            return render_template("survey_error.html", store=store, error="Sorry, the system is not accepting any feedbacks right now"), 404

        questionnaire = fetch_questionnaire_by_store(store_id=store_id)
        if not questionnaire or not questionnaire.get("is_active"):
            return render_template("survey_error.html", store=store, error="Sorry, the system is not accepting any feedbacks right now"), 404

        questions = fetch_questions_for_questionnaire(questionnaire_id=int(questionnaire["id"]))
        options_by_question_id = fetch_options_for_questions([int(q["id"]) for q in questions])

        # Get and validate receipt number
        receipt_number = request.form.get("receipt_number", "").strip()
        if not receipt_number:
            flash("Receipt/Transaction number is required.", "danger")
            return redirect(url_for("public_survey", store_id=store_id))
        
        # The SI/transaction number printed on the receipt is exactly 8 digits.
        if not re.fullmatch(r'\d{8}', receipt_number):
            flash("Receipt/Transaction number must contain exactly 8 digits.", "danger")
            return redirect(url_for("public_survey", store_id=store_id))

        # Get and validate email
        user_email = request.form.get("user_email", "").strip()
        if not user_email:
            flash("Email address is required.", "danger")
            return redirect(url_for("public_survey", store_id=store_id))
        
        # Basic email validation
        if "@" not in user_email or "." not in user_email.split("@")[1]:
            flash("Please enter a valid email address.", "danger")
            return redirect(url_for("public_survey", store_id=store_id))

        errors: List[str] = []
        answers_to_save: List[Dict[str, Any]] = []

        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""SELECT id, role FROM staff
                              WHERE store_id = %s AND status = 'active'""", (store_id,))
            eligible_staff = {int(row['id']): row['role'] for row in cursor.fetchall()}
        finally:
            conn.close()

        for q in questions:
            qid = int(q["id"])
            key = f"q_{qid}"
            q_type = q["question_type"]
            is_required = bool(q["is_required"])
            target_scope = q.get("target_scope") or "overall"
            target_staff_id = None
            if target_scope in ("staff", "manager"):
                raw_target = request.form.get(f"target_{qid}", "").strip()
                if raw_target.isdigit():
                    candidate_id = int(raw_target)
                    candidate_role = eligible_staff.get(candidate_id)
                    role_matches = (
                        candidate_role == "manager" if target_scope == "manager"
                        else candidate_role in ("staff", "supervisor")
                    )
                    if role_matches:
                        target_staff_id = candidate_id
                if target_staff_id is None and is_required:
                    errors.append(f"Select a {target_scope}: {q['question_text']}")
                    continue

            if q_type == "rating":
                raw = request.form.get(key, "").strip()
                if not raw:
                    if is_required:
                        errors.append(f"Rating required: {q['question_text']}")
                    continue
                try:
                    rating_value = int(raw)
                except ValueError:
                    errors.append(f"Invalid rating: {q['question_text']}")
                    continue
                if rating_value < 1 or rating_value > 5:
                    errors.append(f"Rating must be 1-5: {q['question_text']}")
                    continue
                comment = request.form.get(f"{key}_comment", "").strip()
                answers_to_save.append(
                    {"question_id": qid, "staff_id": target_staff_id, "answer_text": comment if comment else None, "rating_value": rating_value}
                )

            elif q_type == "text":
                text = request.form.get(key, "")
                text = text.strip()
                if not text:
                    if is_required:
                        errors.append(f"Answer required: {q['question_text']}")
                    continue
                answers_to_save.append({"question_id": qid, "staff_id": target_staff_id, "answer_text": text, "rating_value": None})

            elif q_type == "multiple_choice":
                raw = request.form.get(key, "").strip()
                if not raw:
                    if is_required:
                        errors.append(f"Choice required: {q['question_text']}")
                    continue

                try:
                    selected_option_id = int(raw)
                except ValueError:
                    errors.append(f"Invalid choice: {q['question_text']}")
                    continue

                options = options_by_question_id.get(qid, [])
                selected_text = None
                for opt in options:
                    if int(opt["id"]) == selected_option_id:
                        selected_text = opt["option_text"]
                        break

                if not selected_text:
                    errors.append(f"Invalid choice: {q['question_text']}")
                    continue

                answers_to_save.append(
                    {"question_id": qid, "staff_id": target_staff_id, "answer_text": selected_text, "rating_value": None}
                )
            else:
                errors.append(f"Unsupported question type: {q_type}")

        if errors:
            for e in errors[:5]:
                flash(e, "danger")
            return redirect(url_for("public_survey", store_id=store_id))

        overall_question_ids = {int(q["id"]) for q in questions
                                if (q.get("target_scope") or "overall") == "overall"}
        rating_values = [float(a["rating_value"]) for a in answers_to_save
                         if a.get("rating_value") is not None
                         and int(a["question_id"]) in overall_question_ids]
        average_rating = (sum(rating_values) / len(rating_values)) if rating_values else 0
        reward_claim_token = None

        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            # Use Philippine time (UTC+08:00)
            ph_tz = timezone(timedelta(hours=8))
            now_ph = datetime.now(ph_tz).strftime("%Y-%m-%d %H:%M:%S")
            try:
                cursor.execute(
                    """INSERT INTO global_receipt_usages (receipt_number, store_id, used_at)
                       VALUES (%s, %s, %s)""",
                    (receipt_number, store_id, now_ph),
                )
            except mysql.connector.IntegrityError as exc:
                conn.rollback()
                if exc.errno == 1062:
                    flash("This Receipt/Transaction Number can only be used once across all branches and has already been submitted.", "danger")
                    return redirect(url_for("public_survey", store_id=store_id))
                raise
            cursor.execute(
                """
                INSERT INTO responses (questionnaire_id, store_id, user_email, receipt_number, submitted_at)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (int(questionnaire["id"]), store_id, user_email, receipt_number, now_ph),
            )
            response_id = int(cursor.lastrowid)
            cursor.execute(
                """UPDATE global_receipt_usages SET response_id = %s
                   WHERE receipt_number = %s""",
                (response_id, receipt_number),
            )

            show_google_review = average_rating >= 4 and bool(store.get("google_review_url"))
            if (show_google_review and store.get("google_review_mode") == "reward" and user_email):
                reward_claim_token = secrets.token_urlsafe(32)
                cursor.execute(
                    """INSERT INTO review_rewards
                       (response_id, store_id, owner_user_id, license_key,
                        customer_email, claim_token, reward_type, status)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, 'pending')""",
                    (response_id, store_id, int(store["user_id"]), store.get("license_key"),
                     user_email, reward_claim_token,
                     store.get("reward_type") or "Store Reward or Discount"),
                )

            for a in answers_to_save:
                cursor.execute(
                    """
                    INSERT INTO answers (response_id, question_id, staff_id, answer_text, rating_value)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (response_id, a["question_id"], a.get("staff_id"), a["answer_text"], a["rating_value"]),
                )
                if a.get("staff_id") and a.get("rating_value") is not None:
                    cursor.execute(
                        """INSERT INTO staff_commendations
                           (response_id, staff_id, rating, commendation_type, comment)
                           VALUES (%s, %s, %s, 'excellent_service', %s)""",
                        (response_id, a["staff_id"], int(a["rating_value"]), a.get("answer_text")),
                    )

            # Handle staff commendation if provided
            staff_commendation = request.form.get("staff_commendation", "").strip()
            if staff_commendation and staff_commendation.isdigit():
                staff_id = int(staff_commendation)
                commendation_type = request.form.get("commendation_type", "excellent_service")
                commendation_comment = request.form.get("commendation_comment", "").strip()
                commendation_rating = request.form.get("commendation_rating", "5").strip()
                if not commendation_rating or not commendation_rating.isdigit():
                    commendation_rating = 5
                commendation_rating = int(commendation_rating)
                
                # Verify staff exists and belongs to this store
                cursor.execute("SELECT id FROM staff WHERE id = %s AND store_id = %s", (staff_id, store_id))
                if cursor.fetchone():
                    cursor.execute("""
                        INSERT INTO staff_commendations (response_id, staff_id, rating, commendation_type, comment)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (response_id, staff_id, commendation_rating, commendation_type, commendation_comment if commendation_comment else None))

            conn.commit()
        finally:
            conn.close()

        if reward_claim_token:
            return redirect(url_for("survey_thank_you", store_id=store_id,
                                    claim=reward_claim_token, review=1))
        return redirect(url_for("survey_thank_you", store_id=store_id,
                                review=1 if show_google_review else None))

    @app.route("/s/<int:store_id>/thanks", methods=["GET"])
    def survey_thank_you(store_id: int):
        store = fetch_store_by_id(store_id=store_id)
        if not store:
            return render_template("layout.html", store=None, error="Page not found"), 404
        claim_token = request.args.get("claim", "").strip()
        reward = None
        if claim_token:
            conn = get_db_connection()
            try:
                cursor = conn.cursor(dictionary=True)
                cursor.execute(
                    """SELECT rr.id, rr.status, rr.reward_code, rr.reward_type,
                              r.receipt_number
                       FROM review_rewards rr
                       JOIN responses r ON r.id = rr.response_id
                       WHERE rr.store_id = %s AND rr.claim_token = %s""",
                    (store_id, claim_token),
                )
                reward = cursor.fetchone()
            finally:
                conn.close()
        show_google_review = request.args.get("review") == "1" and bool(store.get("google_review_url"))
        return render_template("master_questionnaire/thank_you.html", store=store,
                               reward=reward, claim_token=claim_token,
                               show_google_review=show_google_review)

    @app.route("/s/<int:store_id>/review-reward/<claim_token>", methods=["POST"])
    def claim_review_reward(store_id: int, claim_token: str):
        """Issue a reward only after uploaded proof passes OCR text validation."""
        review_file = request.files.get("google_review_proof")
        review_ocr_text = request.form.get("review_ocr_text", "").strip()[:12000]

        def validated_image_data(upload):
            if not upload or not upload.filename:
                return None
            if upload.mimetype not in {"image/jpeg", "image/png", "image/webp"}:
                return None
            content = upload.read(2 * 1024 * 1024 + 1)
            if not content or len(content) > 2 * 1024 * 1024:
                return None
            return f"data:{upload.mimetype};base64,{base64.b64encode(content).decode('ascii')}"

        review_proof = validated_image_data(review_file)
        if not review_proof:
            flash("Upload a clear Google Review screenshot (PNG, JPG, or WEBP; maximum 2 MB).", "danger")
            return redirect(url_for("survey_thank_you", store_id=store_id, claim=claim_token))

        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                """SELECT rr.*, s.store_name, r.receipt_number
                   FROM review_rewards rr
                   JOIN stores s ON s.id = rr.store_id
                   JOIN responses r ON r.id = rr.response_id
                   WHERE rr.store_id = %s AND rr.claim_token = %s FOR UPDATE""",
                (store_id, claim_token),
            )
            reward = cursor.fetchone()
            if not reward:
                return "Invalid or expired reward claim", 404
            if reward["status"] == "pending":
                review_text_lower = review_ocr_text.lower()
                review_matches = ("review" in review_text_lower and
                                  any(term in review_text_lower for term in ("done", "point", "posted", "contribute", "published")))
                if not review_matches:
                    conn.rollback()
                    flash("OCR verification failed. Upload a clear full screenshot showing the completed Google Review.", "danger")
                    return redirect(url_for("survey_thank_you", store_id=store_id, claim=claim_token))

                code = "RWD-" + secrets.token_hex(5).upper()
                cursor.execute(
                    """UPDATE review_rewards SET reward_code = %s, status = 'issued',
                       google_review_proof = %s, review_ocr_text = %s,
                       proof_verified_at = NOW(), issued_at = NOW()
                       WHERE id = %s AND status = 'pending'""",
                    (code, review_proof, review_ocr_text, int(reward["id"])),
                )
                conn.commit()
                reward["reward_code"] = code
                reward["status"] = "issued"
            else:
                conn.commit()
        finally:
            conn.close()

        if reward["status"] == "issued" and not reward.get("email_sent"):
            message = (f"Thank you for reviewing {reward['store_name']} on Google. "
                       f"Your unique reward code is {reward['reward_code']}. "
                       f"Reward: {reward['reward_type']}. Bring your original receipt and a screenshot "
                       f"of this code to the store for verification. This code can only be redeemed once.")
            success, email_message = email_config.send_feedback_reply(
                to_email=reward["customer_email"],
                customer_name=reward["customer_email"].split("@")[0].replace(".", " ").title(),
                reply_message=message,
                store_name=reward["store_name"],
                feedback_summary="Google Review Reward",
                template_type="appreciation",
            )
            conn = get_db_connection()
            try:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE review_rewards SET email_sent = %s, email_error = %s WHERE id = %s",
                    (bool(success), None if success else str(email_message)[:1000], int(reward["id"])),
                )
                conn.commit()
            finally:
                conn.close()
        return redirect(url_for("survey_thank_you", store_id=store_id, claim=claim_token, issued=1))

    @app.route("/admin/rewards", methods=["GET"])
    @role_required('admin', 'superadmin')
    def admin_rewards():
        user = get_user_by_id(session["user_id"])
        search = request.args.get("search", "").strip()[:100]
        status_filter = request.args.get("status", "").strip().lower()
        store_filter = request.args.get("store", "").strip()
        try:
            per_page = int(request.args.get("per_page", 20))
        except (TypeError, ValueError):
            per_page = 20
        if per_page not in (20, 50, 100):
            per_page = 20
        try:
            page = max(1, int(request.args.get("page", 1)))
        except (TypeError, ValueError):
            page = 1

        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            if user["role"] == "superadmin":
                cursor.execute("SELECT * FROM stores ORDER BY store_name")
                stores = cursor.fetchall()
                scope_sql = "1 = 1"
                scope_params = []
            else:
                cursor.execute("SELECT * FROM stores WHERE user_id = %s ORDER BY store_name", (user["id"],))
                stores = cursor.fetchall()
                scope_sql = "rr.owner_user_id = %s AND rr.license_key <=> %s"
                scope_params = [user["id"], user.get("license_key")]

            filters = [scope_sql]
            params = list(scope_params)
            if search:
                term = f"%{search}%"
                filters.append("(rr.reward_code LIKE %s OR rr.customer_email LIKE %s OR s.store_name LIKE %s OR rr.reward_type LIKE %s)")
                params.extend([term, term, term, term])
            if status_filter in {"pending", "issued", "used"}:
                filters.append("rr.status = %s")
                params.append(status_filter)
            if store_filter.isdigit():
                filters.append("rr.store_id = %s")
                params.append(int(store_filter))

            where_sql = " AND ".join(filters)
            cursor.execute(f"""SELECT COUNT(*) AS total
                               FROM review_rewards rr
                               JOIN stores s ON s.id = rr.store_id
                               WHERE {where_sql}""", tuple(params))
            total_rewards = int(cursor.fetchone()["total"])
            total_pages = max(1, (total_rewards + per_page - 1) // per_page)
            page = min(page, total_pages)
            start_page = max(1, page - 2)
            end_page = min(total_pages, page + 2)
            offset = (page - 1) * per_page
            cursor.execute(f"""SELECT rr.*, s.store_name, s.google_review_url
                               FROM review_rewards rr
                               JOIN stores s ON s.id = rr.store_id
                               WHERE {where_sql}
                               ORDER BY rr.created_at DESC
                               LIMIT %s OFFSET %s""", tuple(params + [per_page, offset]))
            rewards = cursor.fetchall()
        finally:
            conn.close()
        return render_template(
            "admin/rewards.html", stores=stores, rewards=rewards,
            search=search, status_filter=status_filter, store_filter=store_filter,
            per_page=per_page, page=page, total_pages=total_pages,
            total_rewards=total_rewards, start_page=start_page, end_page=end_page,
        )

    @app.route("/admin/stores/<int:store_id>/reward-settings", methods=["POST"])
    @role_required('admin', 'superadmin')
    def save_reward_settings(store_id: int):
        if not can_manage_store(session["user_id"], store_id):
            flash("You don't have permission to update this store.", "danger")
            return redirect(url_for("admin_rewards"))
        reward_type = request.form.get("reward_type", "").strip()
        google_review_mode = request.form.get("google_review_mode", "reward").strip()
        if google_review_mode not in {"review_only", "reward"}:
            google_review_mode = "reward"
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("UPDATE stores SET reward_type = %s, google_review_mode = %s WHERE id = %s",
                           (reward_type or "Store Reward or Discount", google_review_mode, store_id))
            conn.commit()
        finally:
            conn.close()
        flash("Google Review option saved.", "success")
        return redirect(url_for("admin_rewards"))

    @app.route("/admin/rewards/<int:reward_id>/use", methods=["POST"])
    @role_required('admin', 'superadmin')
    def use_review_reward(reward_id: int):
        user = get_user_by_id(session["user_id"])
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            if user["role"] == "superadmin":
                cursor.execute("UPDATE review_rewards SET status='used', used_at=NOW() WHERE id=%s AND status='issued'", (reward_id,))
            else:
                cursor.execute("""UPDATE review_rewards SET status='used', used_at=NOW()
                                  WHERE id=%s AND owner_user_id=%s AND license_key <=> %s AND status='issued'""",
                               (reward_id, user["id"], user.get("license_key")))
            conn.commit()
        finally:
            conn.close()
        flash("Reward code marked as used.", "success")
        return redirect(url_for("admin_rewards"))

    @app.route("/admin/stores/add", methods=["POST"])
    @login_required
    def add_store():
        store_name = request.form.get("store_name", "").strip()
        address = request.form.get("address", "").strip()
        city = request.form.get("city", "").strip()
        province = request.form.get("province", "").strip()
        postal_code = request.form.get("postal_code", "").strip()
        contact_number = request.form.get("contact_number", "").strip()
        email = request.form.get("email", "").strip()
        store_manager_name = request.form.get("store_manager_name", "").strip()
        manager_contact = request.form.get("manager_contact", "").strip()
        store_type = request.form.get("store_type", "").strip()
        status = request.form.get("status", "active")
        google_review_url = request.form.get("google_review_url", "").strip()
        google_review_mode = request.form.get("google_review_mode", "reward").strip()
        if google_review_mode not in {"review_only", "reward"}:
            google_review_mode = "reward"

        if not store_name:
            flash("Store name is required.", "danger")
            return redirect(url_for("stores_management"))

        # Basic email validation if provided
        if email and ("@" not in email or "." not in email.split("@")[1]):
            flash("Please enter a valid email address.", "danger")
            return redirect(url_for("stores_management"))
        if google_review_url and not google_review_url.startswith("https://"):
            flash("Google Review Link must start with https://", "danger")
            return redirect(url_for("stores_management"))

        # Check store limit based on user role and membership
        user = get_user_by_id(session['user_id'])
        
        if user['role'] == 'admin':
            # Client account - check if license is configured
            if not user.get('license_key'):
                flash("Please configure your license key first. Contact your administrator for your license key.", "danger")
                return redirect(url_for("client_license_config"))
            
            # Validate license against portal and get max_stores
            config = get_license_config()
            portal_url = normalize_portal_url(config.get("licensing_portal_url") if config else None)
            
            try:
                import requests
                response = requests.post(
                    f"{portal_url}/api/validate/{user['license_key']}",
                    headers=licensing_api_headers(),
                    timeout=10
                )
                
                logger.info(f"License validation response status: {response.status_code}")
                
                if response.status_code != 200 or not response.json().get("valid"):
                    flash("Invalid license. Please contact your administrator.", "danger")
                    return redirect(url_for("client_license_config"))
                
                license_data = response.json()
                max_stores = license_data.get("max_stores", 0)
                logger.info(f"License data: {license_data}")
                
                if max_stores > 0:
                    conn = get_db_connection()
                    try:
                        cursor = conn.cursor()
                        cursor.execute("SELECT COUNT(*) FROM stores WHERE user_id = %s", (session['user_id'],))
                        current_count = cursor.fetchone()[0]
                        
                        cursor.execute("SELECT COUNT(*) FROM stores")
                        total_stores = cursor.fetchone()[0]
                        
                        if current_count >= max_stores:
                            flash(f"Your license limit reached. You can only create up to {max_stores} stores. Contact support to upgrade.", "danger")
                            return redirect(url_for("stores_management"))
                    finally:
                        conn.close()
            except Exception as e:
                logger.error(f"Error validating license: {e}")
                flash("Unable to validate license. Please try again later.", "danger")
                return redirect(url_for("stores_management"))
        elif user['role'] != 'superadmin':
            flash("You don't have permission to add stores.", "danger")
            return redirect(url_for("stores_management"))

        # Handle logo upload
        logo_url = None
        if 'logo' in request.files:
            logo_file = request.files['logo']
            if logo_file and logo_file.filename:
                # Validate file type
                allowed_extensions = {'png', 'jpg', 'jpeg'}
                if '.' not in logo_file.filename or logo_file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
                    flash("Invalid file type. Only PNG, JPG, and JPEG files are allowed.", "danger")
                    return redirect(url_for("stores_management"))
                
                # Validate file size (5MB max)
                logo_file.seek(0, os.SEEK_END)
                file_size = logo_file.tell()
                logo_file.seek(0)
                if file_size > 5 * 1024 * 1024:
                    flash("File size exceeds 5MB limit.", "danger")
                    return redirect(url_for("stores_management"))
                
                import base64
                ext = logo_file.filename.rsplit('.', 1)[1].lower()
                mime = "image/jpeg" if ext in ('jpg', 'jpeg') else "image/png"
                logo_url = f"data:{mime};base64,{base64.b64encode(logo_file.read()).decode('utf-8')}"

        new_store_id = create_store(
            store_name=store_name,
            address=address if address else None,
            city=city if city else None,
            province=province if province else None,
            postal_code=postal_code if postal_code else None,
            contact_number=contact_number if contact_number else None,
            email=email if email else None,
            store_manager_name=store_manager_name if store_manager_name else None,
            manager_contact=manager_contact if manager_contact else None,
            store_type=store_type if store_type else None,
            status=status,
            logo_url=logo_url,
            google_review_url=google_review_url or None,
            google_review_mode=google_review_mode,
            user_id=session.get('user_id'),
            license_key=user.get('license_key')
        )
        
        logger.info(f"Created store {new_store_id} for user {session.get('user_id')}")
        
        # Log the store addition
        log_audit(
            entity_type="store",
            entity_id=new_store_id,
            action="created",
            new_values=f"Store Name: {store_name}, Address: {address}, City: {city}, Status: {status}"
        )
        
        flash(f"Store \"{store_name}\" added Successfully", "success")
        return redirect(url_for("stores_management"))

    def update_store(
        store_id: int,
        store_name: str,
        store_type: str | None,
        address: str | None,
        city: str | None,
        province: str | None,
        postal_code: str | None,
        contact_number: str | None,
        email: str | None,
        store_manager_name: str | None,
        manager_contact: str | None,
        status: str,
        logo_url: str | None = None,
        google_review_url: str | None = None,
        google_review_mode: str = "reward"
    ) -> bool:
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute(
                """
                UPDATE stores
                SET store_name = %s, store_type = %s, address = %s, city = %s,
                    province = %s, postal_code = %s, contact_number = %s,
                    email = %s, store_manager_name = %s, manager_contact = %s,
                    status = %s, logo_url = COALESCE(%s, logo_url), google_review_url = %s,
                    google_review_mode = %s
                WHERE id = %s
                """,
                (
                    store_name,
                    store_type,
                    address,
                    city,
                    province,
                    postal_code,
                    contact_number,
                    email,
                    store_manager_name,
                    manager_contact,
                    status,
                    logo_url,
                    google_review_url,
                    google_review_mode,
                    store_id,
                ),
            )
            conn.commit()
            return cursor.rowcount > 0
        finally:
            conn.close()

    @app.route("/admin/stores/<int:store_id>/upload-logo", methods=["POST"])
    @login_required
    def upload_store_logo(store_id: int):
        if not can_manage_store(session['user_id'], store_id):
            flash("You don't have permission to update this store.", "danger")
            return redirect(url_for("stores_management"))
        # Handle logo upload only
        logo_url = None
        if 'logo' in request.files:
            logo_file = request.files['logo']
            if logo_file and logo_file.filename:
                # Validate file type
                allowed_extensions = {'png', 'jpg', 'jpeg'}
                if '.' not in logo_file.filename or logo_file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
                    flash("Invalid file type. Only PNG, JPG, and JPEG files are allowed.", "danger")
                    return redirect(url_for("store_details", store_id=store_id))
                
                # Validate file size (5MB max)
                logo_file.seek(0, os.SEEK_END)
                file_size = logo_file.tell()
                logo_file.seek(0)
                if file_size > 5 * 1024 * 1024:
                    flash("File size exceeds 5MB limit.", "danger")
                    return redirect(url_for("store_details", store_id=store_id))
                
                import base64
                ext = logo_file.filename.rsplit('.', 1)[1].lower()
                mime = "image/jpeg" if ext in ('jpg', 'jpeg') else "image/png"
                logo_url = f"data:{mime};base64,{base64.b64encode(logo_file.read()).decode('utf-8')}"

        # Update only the logo_url in the database
        if logo_url:
            conn = get_db_connection()
            try:
                cursor = conn.cursor()
                cursor.execute("UPDATE stores SET logo_url = %s WHERE id = %s", (logo_url, store_id))
                conn.commit()
                flash("Logo uploaded successfully", "success")
            except Exception as e:
                logger.error(f"Error uploading logo: {e}")
                flash(f"Error uploading logo: {e}", "danger")
            finally:
                conn.close()
        else:
            flash("No file selected", "warning")

        return redirect(url_for("store_details", store_id=store_id))

    @app.route("/admin/stores/<int:store_id>/edit", methods=["POST"])
    @login_required
    def edit_store(store_id: int):
        if not can_manage_store(session['user_id'], store_id):
            flash("You don't have permission to edit this store.", "danger")
            return redirect(url_for("stores_management"))
        store_name = request.form.get("store_name", "").strip()
        store_type = request.form.get("store_type", "").strip() or None
        address = request.form.get("address", "").strip() or None
        city = request.form.get("city", "").strip() or None
        province = request.form.get("province", "").strip() or None
        postal_code = request.form.get("postal_code", "").strip() or None
        contact_number = request.form.get("contact_number", "").strip() or None
        email = request.form.get("email", "").strip() or None
        store_manager_name = request.form.get("store_manager_name", "").strip() or None
        manager_contact = request.form.get("manager_contact", "").strip() or None
        status = request.form.get("status", "active")
        google_review_url = request.form.get("google_review_url", "").strip()
        google_review_mode = request.form.get("google_review_mode", "reward").strip()
        if google_review_mode not in {"review_only", "reward"}:
            google_review_mode = "reward"

        if not store_name:
            flash("Store name is required.", "danger")
            return redirect(url_for("stores_management"))
        if google_review_url and not google_review_url.startswith("https://"):
            flash("Google Review Link must start with https://", "danger")
            return redirect(url_for("stores_management"))

        # Handle logo upload
        logo_url = None
        if 'logo' in request.files:
            logo_file = request.files['logo']
            if logo_file and logo_file.filename:
                # Validate file type
                allowed_extensions = {'png', 'jpg', 'jpeg'}
                if '.' not in logo_file.filename or logo_file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
                    flash("Invalid file type. Only PNG, JPG, and JPEG files are allowed.", "danger")
                    return redirect(url_for("stores_management"))
                
                # Validate file size (5MB max)
                logo_file.seek(0, os.SEEK_END)
                file_size = logo_file.tell()
                logo_file.seek(0)
                if file_size > 5 * 1024 * 1024:
                    flash("File size exceeds 5MB limit.", "danger")
                    return redirect(url_for("stores_management"))
                
                import base64
                ext = logo_file.filename.rsplit('.', 1)[1].lower()
                mime = "image/jpeg" if ext in ('jpg', 'jpeg') else "image/png"
                logo_url = f"data:{mime};base64,{base64.b64encode(logo_file.read()).decode('utf-8')}"

        success = update_store(
            store_id=store_id,
            store_name=store_name,
            store_type=store_type,
            address=address,
            city=city,
            province=province,
            postal_code=postal_code,
            contact_number=contact_number,
            email=email,
            store_manager_name=store_manager_name,
            manager_contact=manager_contact,
            status=status,
            logo_url=logo_url,
            google_review_url=google_review_url or None,
            google_review_mode=google_review_mode
        )

        if success:
            # Log the store edit
            log_audit(
                entity_type="store",
                entity_id=store_id,
                action="updated",
                new_values=f"Store Name: {store_name}, Address: {address}, City: {city}, Status: {status}"
            )
            flash(f"Store \"{store_name}\" Edited", "success")
        else:
            flash("Store not found or update failed.", "danger")

        return redirect(url_for("stores_management", store_id=store_id))

    @app.route("/admin/stores/<int:store_id>/delete", methods=["POST"])
    @role_required('admin', 'superadmin')
    def delete_store_route(store_id: int):
        if not can_manage_store(session['user_id'], store_id):
            flash("You don't have permission to delete this store.", "danger")
            return redirect(url_for("stores_management"))
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            # Fetch store name before deletion for the notification
            cursor.execute("SELECT store_name FROM stores WHERE id = %s", (store_id,))
            store_row = cursor.fetchone()
            store_name = store_row[0] if store_row else "Unknown"

            if not store_row:
                flash("Store not found.", "warning")
                return redirect(url_for("stores_management"))

            # These receipt tables intentionally have no foreign keys, so
            # remove their store-scoped records before deleting the store.
            cursor.execute("DELETE FROM receipt_usages WHERE store_id = %s", (store_id,))
            cursor.execute("DELETE FROM global_receipt_usages WHERE store_id = %s", (store_id,))

            # Cascading delete: delete staff_commendations first
            cursor.execute("""
                DELETE sc FROM staff_commendations sc
                JOIN responses r ON sc.response_id = r.id
                WHERE r.store_id = %s
            """, (store_id,))

            # Delete answers
            cursor.execute("""
                DELETE a FROM answers a
                JOIN responses r ON a.response_id = r.id
                WHERE r.store_id = %s
            """, (store_id,))
            
            # Delete responses
            cursor.execute("DELETE FROM responses WHERE store_id = %s", (store_id,))

            # Delete staff
            cursor.execute("DELETE FROM staff WHERE store_id = %s", (store_id,))
            
            # Delete question options for store's questionnaires
            cursor.execute("""
                DELETE qo FROM question_options qo
                JOIN questions q ON qo.question_id = q.id
                JOIN questionnaires qn ON q.questionnaire_id = qn.id
                WHERE qn.store_id = %s
            """, (store_id,))
            
            # Delete questions
            cursor.execute("""
                DELETE q FROM questions q
                JOIN questionnaires qn ON q.questionnaire_id = qn.id
                WHERE qn.store_id = %s
            """, (store_id,))
            
            # Delete questionnaires
            cursor.execute("DELETE FROM questionnaires WHERE store_id = %s", (store_id,))
            
            # Delete store itself
            cursor.execute("DELETE FROM stores WHERE id = %s", (store_id,))
            
            conn.commit()
            
            # Log the store deletion
            log_audit(
                entity_type="store",
                entity_id=store_id,
                action="deleted",
                old_values=f"Store Name: {store_name}"
            )
            
            flash(f"Store \"{store_name}\" Deleted", "success")
        except Exception as e:
            logger.error(f"Error deleting store: {e}")
            flash(f"Error deleting store: {e}", "danger")
        finally:
            conn.close()
            
        return redirect(url_for("stores_management"))

    @app.route("/admin/history")
    @role_required('superadmin')
    def history():
        # Run automatic pruning of old logs (90 days retention)
        try:
            prune_audit_logs(days=90)
        except Exception as e:
            logger.error(f"Error pruning audit logs: {e}")
        
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT id, entity_type, entity_id, action, old_values, new_values, user_id, created_at
                FROM audit_logs
                ORDER BY created_at DESC
                LIMIT 100
            """)
            logs = cursor.fetchall()
        finally:
            conn.close()
        
        return render_template("history.html", logs=logs)

    @app.route("/admin/history/clear", methods=["POST"])
    @role_required('superadmin')
    def clear_history():
        deleted_count = prune_audit_logs(days=0)  # Delete all logs
        flash(f"Cleared {deleted_count} history entries", "success")
        return redirect(url_for("history"))

    @app.route("/admin/clear-feedback", methods=["POST"])
    def clear_feedback_route():
        """Clear all feedback data while preserving stores and their configuration."""
        conn = get_db_connection()
        try:
            cursor = conn.cursor()

            # Count before deletion for feedback
            cursor.execute("SELECT COUNT(*) FROM responses")
            responses_count = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM answers")
            answers_count = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM staff_commendations")
            commendations_count = cursor.fetchone()[0]

            # Delete in correct order (respecting foreign keys)
            # 1. Delete staff_commendations
            cursor.execute("DELETE FROM staff_commendations")

            # 2. Delete answers
            cursor.execute("DELETE FROM answers")

            # 3. Delete responses
            cursor.execute("DELETE FROM responses")

            conn.commit()

            flash(f"Cleared {responses_count} responses, {answers_count} answers, and {commendations_count} commendations. Stores and configurations preserved.", "success")
        except Exception as e:
            logger.error(f"Error clearing feedback data: {e}")
            flash(f"Error clearing feedback data: {e}", "danger")
        finally:
            conn.close()

        return redirect(url_for("stores_management"))

    @app.route("/admin/backup/csv", methods=["GET"])
    def backup_csv_route():
        """Export all data to a CSV organized by store.

        Layout:
          # FEEDBACK SYSTEM BACKUP
          # Generated: <timestamp>
          # Total stores: N

          === STORES SUMMARY ===
          <stores table>

          === STORE: <name> (id=<id>) ===
            -- Staff --
            <staff for this store>
            -- Feedback --
            <responses joined with answers, one row per question>
            -- Commendations --
            <commendations for this store>
          (repeat per store)
        """
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)

            cursor.execute("SELECT * FROM stores ORDER BY id")
            stores = cursor.fetchall()

            output = io.StringIO()
            # UTF-8 BOM so Excel opens it correctly with special characters
            output.write('\ufeff')
            writer = csv.writer(output)

            timestamp_human = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            writer.writerow(["# FEEDBACK SYSTEM BACKUP"])
            writer.writerow([f"# Generated: {timestamp_human}"])
            writer.writerow([f"# Total stores: {len(stores)}"])
            writer.writerow([])

            # ---- Stores summary ----
            writer.writerow(["=== STORES SUMMARY ==="])
            if stores:
                # Counts per store for quick overview
                cursor.execute("SELECT store_id, COUNT(*) AS cnt FROM responses GROUP BY store_id")
                resp_counts = {r['store_id']: r['cnt'] for r in cursor.fetchall()}
                cursor.execute("SELECT store_id, COUNT(*) AS cnt FROM staff GROUP BY store_id")
                staff_counts = {r['store_id']: r['cnt'] for r in cursor.fetchall()}

                summary_cols = list(stores[0].keys()) + ["total_responses", "total_staff"]
                writer.writerow(summary_cols)
                for s in stores:
                    row = list(s.values()) + [
                        resp_counts.get(s['id'], 0),
                        staff_counts.get(s['id'], 0),
                    ]
                    writer.writerow(row)
            else:
                writer.writerow(["(no stores)"])
            writer.writerow([])
            writer.writerow([])

            # ---- Per-store sections ----
            for store in stores:
                store_id = store['id']
                store_name = store.get('store_name') or f"Store #{store_id}"

                writer.writerow([f"=== STORE: {store_name} (id={store_id}) ==="])

                # Staff
                cursor.execute("SELECT * FROM staff WHERE store_id = %s ORDER BY id", (store_id,))
                staff_rows = cursor.fetchall()
                writer.writerow(["-- Staff --"])
                if staff_rows:
                    writer.writerow(staff_rows[0].keys())
                    for r in staff_rows:
                        writer.writerow(r.values())
                else:
                    writer.writerow(["(no staff)"])
                writer.writerow([])

                # Feedback (responses joined with answers)
                cursor.execute(
                    """
                    SELECT r.id AS response_id,
                           r.user_email,
                           r.submitted_at,
                           r.is_read,
                           r.status,
                           a.id AS answer_id,
                           a.question_id,
                           a.answer_text,
                           a.rating_value
                    FROM responses r
                    LEFT JOIN answers a ON a.response_id = r.id
                    WHERE r.store_id = %s
                    ORDER BY r.submitted_at DESC, r.id, a.id
                    """,
                    (store_id,)
                )
                feedback_rows = cursor.fetchall()
                writer.writerow(["-- Feedback --"])
                if feedback_rows:
                    writer.writerow(feedback_rows[0].keys())
                    for r in feedback_rows:
                        writer.writerow(r.values())
                else:
                    writer.writerow(["(no feedback)"])
                writer.writerow([])

                # Commendations for this store's staff
                cursor.execute(
                    """
                    SELECT c.*
                    FROM staff_commendations c
                    JOIN staff s ON c.staff_id = s.id
                    WHERE s.store_id = %s
                    ORDER BY c.id
                    """,
                    (store_id,)
                )
                commend_rows = cursor.fetchall()
                writer.writerow(["-- Commendations --"])
                if commend_rows:
                    writer.writerow(commend_rows[0].keys())
                    for r in commend_rows:
                        writer.writerow(r.values())
                else:
                    writer.writerow(["(no commendations)"])
                writer.writerow([])
                writer.writerow([])

            output.seek(0)
            csv_data = output.getvalue()

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"feedback_system_backup_{timestamp}.csv"

            return send_file(
                io.BytesIO(csv_data.encode('utf-8')),
                mimetype='text/csv',
                as_attachment=True,
                download_name=filename
            )
        except Exception as e:
            logger.error(f"Error creating CSV backup: {e}")
            flash(f"Error creating backup: {e}", "danger")
            return redirect(url_for("stores_management"))
        finally:
            conn.close()

    @app.route("/admin/seed-feedback", methods=["POST"])
    def seed_feedback_route():
        """Seed sample feedback data for each store with answers and staff commendations."""
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)

            # Fetch all stores
            cursor.execute("SELECT * FROM stores WHERE user_id = %s", (session['user_id'],))
            stores = cursor.fetchall()

            # Fetch template questionnaire
            user = get_user_by_id(session['user_id'])
            cursor.execute("SELECT * FROM questionnaires WHERE is_template = TRUE AND owner_user_id = %s AND license_key <=> %s LIMIT 1", (session['user_id'], user.get('license_key') if user else None))
            template_questionnaire = cursor.fetchone()

            if not template_questionnaire:
                flash("No template questionnaire found. Please create one first.", "danger")
                return redirect(url_for("stores_management"))

            # Fetch questions from template questionnaire
            cursor.execute("SELECT * FROM questions WHERE questionnaire_id = %s", (template_questionnaire["id"],))
            template_questions = cursor.fetchall()

            total_responses = 0
            total_answers = 0
            total_commendations = 0

            # Sample data for feedback
            sample_emails = [
                "customer1@example.com", "customer2@example.com", "customer3@example.com",
                "customer4@example.com", "customer5@example.com", "customer6@example.com",
                "customer7@example.com", "customer8@example.com", "customer9@example.com",
                "customer10@example.com"
            ]

            sample_receipts = [
                "REC-001", "REC-002", "REC-003", "REC-004", "REC-005",
                "REC-006", "REC-007", "REC-008", "REC-009", "REC-010"
            ]

            sample_answers_text = [
                "Great service!", "Very satisfied", "Excellent experience",
                "Good quality", "Friendly staff", "Quick service",
                "Clean environment", "Helpful team", "Professional",
                "Will return again"
            ]

            for store in stores:
                store_id = int(store["id"])

                # Fetch or create store-specific questionnaire
                cursor.execute("SELECT * FROM questionnaires WHERE store_id = %s", (store_id,))
                store_questionnaire = cursor.fetchone()

                if not store_questionnaire:
                    # Create a new questionnaire for this store from template
                    cursor.execute(
                        """
                        INSERT INTO questionnaires (title, store_id, is_active, is_template)
                        VALUES (%s, %s, %s, %s)
                        """,
                        (template_questionnaire["title"], store_id, True, False),
                    )
                    store_questionnaire_id = int(cursor.lastrowid)

                    # Copy questions from template to store questionnaire
                    for template_q in template_questions:
                        cursor.execute(
                            """
                            INSERT INTO questions (questionnaire_id, question_text, question_type, is_required, min_label, max_label, allow_comment, question_order)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (store_questionnaire_id, template_q["question_text"], template_q["question_type"],
                             template_q["is_required"], template_q["min_label"], template_q["max_label"],
                             template_q["allow_comment"], template_q["question_order"]),
                        )
                        new_question_id = int(cursor.lastrowid)

                        # Copy options if it's a multiple choice question
                        if template_q["question_type"] == "multiple_choice":
                            cursor.execute("SELECT * FROM question_options WHERE question_id = %s", (template_q["id"],))
                            options = cursor.fetchall()
                            for opt in options:
                                cursor.execute(
                                    """
                                    INSERT INTO question_options (question_id, option_text)
                                    VALUES (%s, %s)
                                    """,
                                    (new_question_id, opt["option_text"]),
                                )
                else:
                    store_questionnaire_id = int(store_questionnaire["id"])

                # Fetch questions for this store's questionnaire
                cursor.execute("SELECT * FROM questions WHERE questionnaire_id = %s", (store_questionnaire_id,))
                questions = cursor.fetchall()

                # Fetch staff for this store
                cursor.execute("SELECT * FROM staff WHERE store_id = %s", (store_id,))
                staff_list = cursor.fetchall()

                # Determine number of feedbacks for this store (5-15)
                num_feedbacks = random.randint(5, 15)

                for i in range(num_feedbacks):
                    # Use Philippine time
                    ph_tz = timezone(timedelta(hours=8))
                    now_ph = datetime.now(ph_tz).strftime("%Y-%m-%d %H:%M:%S")

                    # Create response
                    cursor.execute(
                        """
                        INSERT INTO responses (questionnaire_id, store_id, user_email, receipt_number, submitted_at, status)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        """,
                        (store_questionnaire_id, store_id, sample_emails[i % len(sample_emails)],
                         sample_receipts[i % len(sample_receipts)], now_ph, random.choice(['resolved', 'unresolved'])),
                    )
                    response_id = int(cursor.lastrowid)
                    total_responses += 1

                    # Add answers for each question
                    for question in questions:
                        question_type = question["question_type"]

                        if question_type == "rating":
                            # Random rating 1-5
                            rating = str(random.randint(1, 5))
                            cursor.execute(
                                """
                                INSERT INTO answers (response_id, question_id, rating_value)
                                VALUES (%s, %s, %s)
                                """,
                                (response_id, question["id"], rating),
                            )
                            total_answers += 1
                        elif question_type == "text":
                            # Random text answer
                            answer_text = sample_answers_text[random.randint(0, len(sample_answers_text) - 1)]
                            cursor.execute(
                                """
                                INSERT INTO answers (response_id, question_id, answer_text)
                                VALUES (%s, %s, %s)
                                """,
                                (response_id, question["id"], answer_text),
                            )
                            total_answers += 1
                        elif question_type == "multiple_choice":
                            # Fetch options for this question
                            cursor.execute("SELECT * FROM question_options WHERE question_id = %s", (question["id"],))
                            options = cursor.fetchall()
                            if options:
                                selected_option = random.choice(options)
                                cursor.execute(
                                    """
                                    INSERT INTO answers (response_id, question_id, answer_text)
                                    VALUES (%s, %s, %s)
                                    """,
                                    (response_id, question["id"], selected_option["option_text"]),
                                )
                                total_answers += 1

                    # Add staff commendations (if staff exists and rating was good)
                    if staff_list and random.random() > 0.5:  # 50% chance
                        num_commendations = random.randint(1, min(3, len(staff_list)))
                        commended_staff = random.sample(staff_list, num_commendations)
                        for staff in commended_staff:
                            cursor.execute(
                                """
                                INSERT INTO staff_commendations (response_id, staff_id)
                                VALUES (%s, %s)
                                """,
                                (response_id, staff["id"]),
                            )
                            total_commendations += 1

            conn.commit()
            flash(f"Seeded {total_responses} feedback responses, {total_answers} answers, and {total_commendations} staff commendations across {len(stores)} stores.", "success")
            logger.info(f"Seeded feedback data: {total_responses} responses, {total_answers} answers, {total_commendations} commendations")
        except Exception as e:
            logger.error(f"Error seeding feedback data: {e}")
            flash(f"Error seeding feedback data: {e}", "danger")
        finally:
            conn.close()

        return redirect(url_for("stores_management"))

    # -------------------------
    # STAFF MANAGEMENT
    # -------------------------

    def _uploaded_staff_photo(field_name: str = "photo") -> Optional[str]:
        photo = request.files.get(field_name)
        if not photo or not photo.filename:
            return None
        if '.' not in photo.filename or photo.filename.rsplit('.', 1)[1].lower() not in {'png', 'jpg', 'jpeg'}:
            raise ValueError("Staff photo must be a PNG or JPG image.")
        photo.seek(0, os.SEEK_END)
        file_size = photo.tell()
        photo.seek(0)
        if file_size > 5 * 1024 * 1024:
            raise ValueError("Staff photo must be 5MB or smaller.")
        ext = photo.filename.rsplit('.', 1)[1].lower()
        mime = "image/jpeg" if ext in ('jpg', 'jpeg') else "image/png"
        return f"data:{mime};base64,{base64.b64encode(photo.read()).decode('utf-8')}"

    @app.route("/admin/stores/<int:store_id>/staff")
    @role_required('user', 'admin', 'superadmin')
    def staff_management(store_id: int):
        if not can_manage_store_staff(session['user_id'], store_id):
            flash("You can only manage staff in your assigned store.", "danger")
            return redirect(url_for("stores_management"))
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            
            # Get store information
            cursor.execute("SELECT * FROM stores WHERE id = %s", (store_id,))
            store = cursor.fetchone()
            
            if not store:
                flash("Store not found", "danger")
                return redirect(url_for("stores_management"))
            
            # Get staff for this store
            cursor.execute("""
                SELECT * FROM staff 
                WHERE store_id = %s 
                ORDER BY role DESC, last_name, first_name
            """, (store_id,))
            staff = cursor.fetchall()
            
            # Generate QR code for the store
            public_url = get_store_public_url(store_id=store_id)
            qr_data_uri = generate_qr_data_uri(public_url)
            
            return render_template("manage_staff/staff.html", store=store, staff=staff, public_url=public_url, qr_data_uri=qr_data_uri)
        except Exception as e:
            logger.error(f"Error loading staff management: {e}")
            flash(f"Error loading staff: {e}", "danger")
            return redirect(url_for("stores_management"))
        finally:
            conn.close()

    @app.route("/admin/stores/<int:store_id>/staff/add", methods=["POST"])
    @role_required('user', 'admin', 'superadmin')
    def add_staff(store_id: int):
        if not can_manage_store_staff(session['user_id'], store_id):
            flash("You can only add staff to your assigned store.", "danger")
            return redirect(url_for("stores_management"))
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip() or None
        phone = request.form.get("phone", "").strip() or None
        position = request.form.get("position", "").strip() or None
        role = request.form.get("role", "staff")
        hire_date = request.form.get("hire_date", "").strip() or None
        try:
            photo_url = _uploaded_staff_photo()
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("staff_management", store_id=store_id))
        
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            
            # Verify store exists
            cursor.execute("SELECT id FROM stores WHERE id = %s", (store_id,))
            if not cursor.fetchone():
                flash("Store not found", "danger")
                return redirect(url_for("stores_management"))
            
            # Insert new staff member
            cursor.execute("""
                INSERT INTO staff (store_id, first_name, last_name, email, phone, position, photo_url, role, hire_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (store_id, first_name, last_name, email, phone, position, photo_url, role, hire_date))
            
            new_staff_id = cursor.lastrowid
            conn.commit()
            
            # Log the staff addition
            log_audit(
                entity_type="staff",
                entity_id=new_staff_id,
                action="created",
                new_values=f"Name: {first_name} {last_name}, Position: {position}, Role: {role}, Store ID: {store_id}"
            )
            
            flash(f"Staff member \"{first_name} {last_name}\" added successfully", "success")
        except Exception as e:
            logger.error(f"Error adding staff: {e}")
            flash(f"Error adding staff: {e}", "danger")
        finally:
            conn.close()
            
        return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))

    @app.route("/admin/stores/<int:store_id>/staff/import-template", methods=["GET"])
    @role_required('user', 'admin', 'superadmin')
    def staff_import_template(store_id: int):
        if not can_manage_store_staff(session['user_id'], store_id):
            flash("You can only manage staff in your assigned store.", "danger")
            return redirect(url_for("stores_management"))

        output = io.StringIO(newline='')
        writer = csv.writer(output)
        writer.writerow(["first_name", "last_name", "email", "phone", "position", "role", "hire_date", "status"])
        writer.writerow(["Juan", "Dela Cruz", "juan@example.com", "09171234567", "Sales Associate", "staff", "2026-09-01", "active"])
        payload = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        payload.seek(0)
        return send_file(
            payload,
            mimetype="text/csv; charset=utf-8",
            as_attachment=True,
            download_name=f"staff_import_template_store_{store_id}.csv",
        )

    @app.route("/admin/stores/<int:store_id>/staff/import", methods=["POST"])
    @role_required('user', 'admin', 'superadmin')
    def import_staff(store_id: int):
        """Bulk-import staff from CSV or XLSX. Photos remain manual per staff profile."""
        if not can_manage_store_staff(session['user_id'], store_id):
            flash("You can only import staff into your assigned store.", "danger")
            return redirect(url_for("stores_management"))

        upload = request.files.get("staff_file")
        if not upload or not upload.filename:
            flash("Please choose a CSV or Excel (.xlsx) file.", "danger")
            return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))

        filename = upload.filename.lower()
        if not filename.endswith((".csv", ".xlsx")):
            flash("Unsupported file type. Please upload a .csv or .xlsx file.", "danger")
            return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))

        raw = upload.read(5 * 1024 * 1024 + 1)
        if len(raw) > 5 * 1024 * 1024:
            flash("Staff import file must be 5MB or smaller.", "danger")
            return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))

        def normalize_header(value: Any) -> str:
            return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")

        aliases = {
            "firstname": "first_name", "first": "first_name",
            "lastname": "last_name", "last": "last_name", "surname": "last_name",
            "mobile": "phone", "phone_number": "phone", "contact_number": "phone",
            "job_title": "position", "designation": "position",
            "date_hired": "hire_date", "hired_date": "hire_date",
        }

        try:
            if filename.endswith(".csv"):
                text_stream = io.StringIO(raw.decode("utf-8-sig"))
                csv_rows = list(csv.reader(text_stream))
                if not csv_rows:
                    raise ValueError("The uploaded file is empty.")
                headers = [aliases.get(normalize_header(h), normalize_header(h)) for h in csv_rows[0]]
                rows = [dict(zip(headers, row)) for row in csv_rows[1:]]
            else:
                from openpyxl import load_workbook
                workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                worksheet = workbook.active
                values = worksheet.iter_rows(values_only=True)
                first_row = next(values, None)
                if not first_row:
                    raise ValueError("The uploaded workbook is empty.")
                headers = [aliases.get(normalize_header(h), normalize_header(h)) for h in first_row]
                rows = [dict(zip(headers, row)) for row in values]
                workbook.close()
        except (UnicodeDecodeError, ValueError) as exc:
            flash(f"Unable to read staff file: {exc}", "danger")
            return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))
        except Exception as exc:
            logger.error("Staff import parsing failed: %s", exc)
            flash("Unable to read the file. Check that it is a valid CSV or .xlsx workbook.", "danger")
            return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))

        if "first_name" not in headers or "last_name" not in headers:
            flash("Missing required columns: first_name and last_name.", "danger")
            return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))
        if len(rows) > 1000:
            flash("A single import can contain up to 1,000 staff rows.", "danger")
            return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))

        valid_roles = {"staff", "manager", "supervisor"}
        valid_statuses = {"active", "inactive"}
        prepared = []
        errors = []
        for row_number, row in enumerate(rows, start=2):
            cleaned = {key: str(value).strip() if value is not None else "" for key, value in row.items()}
            if not any(cleaned.values()):
                continue
            first_name = cleaned.get("first_name", "")
            last_name = cleaned.get("last_name", "")
            email = cleaned.get("email", "")
            role = cleaned.get("role", "staff").lower() or "staff"
            status = cleaned.get("status", "active").lower() or "active"
            hire_date = row.get("hire_date")

            if not first_name or not last_name:
                errors.append(f"Row {row_number}: first_name and last_name are required")
                continue
            if email and ("@" not in email or "." not in email.rsplit("@", 1)[-1]):
                errors.append(f"Row {row_number}: invalid email")
                continue
            if role not in valid_roles:
                errors.append(f"Row {row_number}: role must be staff, manager, or supervisor")
                continue
            if status not in valid_statuses:
                errors.append(f"Row {row_number}: status must be active or inactive")
                continue
            if isinstance(hire_date, datetime):
                hire_date = hire_date.date().isoformat()
            elif isinstance(hire_date, date):
                hire_date = hire_date.isoformat()
            elif hire_date:
                hire_date = str(hire_date).strip()
                try:
                    datetime.strptime(hire_date, "%Y-%m-%d")
                except ValueError:
                    errors.append(f"Row {row_number}: hire_date must use YYYY-MM-DD")
                    continue
            else:
                hire_date = None

            prepared.append((
                store_id, first_name, last_name, email or None,
                cleaned.get("phone") or None, cleaned.get("position") or None,
                role, status, hire_date,
            ))

        if not prepared:
            detail = errors[0] if errors else "No staff rows were found."
            flash(f"No staff imported. {detail}", "danger")
            return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))

        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM stores WHERE id = %s", (store_id,))
            if not cursor.fetchone():
                flash("Store not found.", "danger")
                return redirect(url_for("stores_management"))
            cursor.executemany(
                """INSERT INTO staff
                   (store_id, first_name, last_name, email, phone, position, role, status, hire_date)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                prepared,
            )
            conn.commit()
        except Exception as exc:
            conn.rollback()
            logger.error("Staff bulk import failed: %s", exc)
            flash("Staff import failed. No new rows were saved.", "danger")
            return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))
        finally:
            conn.close()

        try:
            log_audit(
                entity_type="staff",
                entity_id=store_id,
                action="bulk_imported",
                new_values=f"Imported {len(prepared)} staff; skipped {len(errors)} invalid rows; Store ID: {store_id}",
            )
        except Exception:
            pass

        message = f"Imported {len(prepared)} staff member(s) successfully. Add profile pictures manually using Edit Staff."
        if errors:
            message += f" Skipped {len(errors)} invalid row(s): " + "; ".join(errors[:3])
        flash(message, "warning" if errors else "success")
        return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))

    @app.route("/admin/stores/<int:store_id>/staff/<int:staff_id>/edit", methods=["POST"])
    @role_required('user', 'admin', 'superadmin')
    def edit_staff(store_id: int, staff_id: int):
        if not can_manage_store_staff(session['user_id'], store_id):
            flash("You can only edit staff in your assigned store.", "danger")
            return redirect(url_for("stores_management"))
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip() or None
        phone = request.form.get("phone", "").strip() or None
        position = request.form.get("position", "").strip() or None
        role = request.form.get("role", "staff")
        status = request.form.get("status", "active")
        hire_date = request.form.get("hire_date", "").strip() or None
        try:
            photo_url = _uploaded_staff_photo()
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("staff_management", store_id=store_id))
        
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            
            # Update staff member
            cursor.execute("""
                UPDATE staff 
                SET first_name = %s, last_name = %s, email = %s, phone = %s,
                    position = %s, photo_url = COALESCE(%s, photo_url), role = %s, status = %s, hire_date = %s
                WHERE id = %s AND store_id = %s
            """, (first_name, last_name, email, phone, position, photo_url, role, status, hire_date, staff_id, store_id))
            
            if cursor.rowcount == 0:
                flash("Staff member not found", "danger")
            else:
                conn.commit()
                
                # Log the staff edit
                log_audit(
                    entity_type="staff",
                    entity_id=staff_id,
                    action="updated",
                    new_values=f"Name: {first_name} {last_name}, Position: {position}, Role: {role}, Status: {status}, Store ID: {store_id}"
                )
                
                flash(f"Staff member \"{first_name} {last_name}\" updated successfully", "success")
        except Exception as e:
            logger.error(f"Error updating staff: {e}")
            flash(f"Error updating staff: {e}", "danger")
        finally:
            conn.close()
            
        return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))

    @app.route("/admin/stores/<int:store_id>/staff/<int:staff_id>/delete", methods=["POST"])
    @role_required('admin', 'superadmin')
    def delete_staff(store_id: int, staff_id: int):
        if not can_manage_store(session['user_id'], store_id):
            flash("You don't have permission to manage staff for this store.", "danger")
            return redirect(url_for("stores_management"))
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            
            # Get staff member name for flash message
            cursor.execute("SELECT first_name, last_name FROM staff WHERE id = %s AND store_id = %s", (staff_id, store_id))
            staff = cursor.fetchone()
            
            if not staff:
                flash("Staff member not found", "danger")
                return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))
            
            # Delete staff member
            cursor.execute("DELETE FROM staff WHERE id = %s AND store_id = %s", (staff_id, store_id))
            conn.commit()
            
            # Log the staff deletion
            log_audit(
                entity_type="staff",
                entity_id=staff_id,
                action="deleted",
                old_values=f"Name: {staff[0]} {staff[1]}, Store ID: {store_id}"
            )
            
            flash(f"Staff member \"{staff[0]} {staff[1]}\" deleted successfully", "success")
        except Exception as e:
            logger.error(f"Error deleting staff: {e}")
            flash(f"Error deleting staff: {e}", "danger")
        finally:
            conn.close()
            
        return redirect(url_for("store_feedback", store_id=store_id, tab='staff'))

    @app.route("/admin/responses/<int:response_id>/delete", methods=["POST"])
    def delete_response_route(response_id: int):
        store_id = request.args.get("store_id")
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            # Delete answers first
            cursor.execute("DELETE FROM answers WHERE response_id = %s", (response_id,))
            # Delete response
            cursor.execute("DELETE FROM responses WHERE id = %s", (response_id,))
            conn.commit()
            flash("Feedback Deleted", "success")
        except Exception as e:
            logger.error(f"Error deleting response: {e}")
            flash(f"Error deleting response: {e}", "danger")
        finally:
            conn.close()
            
        if store_id:
            return redirect(url_for("store_feedback", store_id=store_id))
        return redirect(url_for("admin_dashboard"))

    # -------------------------
    # QUESTION ORDER MANAGEMENT
    # -------------------------
    @app.route("/admin/questions/<int:question_id>/order", methods=["POST"])
    def update_question_order(question_id: int):
        if request.method == "POST":
            try:
                data = request.get_json()
                new_order = int(data.get("question_order", 0))
                template = ensure_template_questionnaire()
                
                conn = get_db_connection()
                try:
                    cursor = conn.cursor()
                    cursor.execute(
                        """
                        UPDATE questions 
                        SET question_order = %s 
                        WHERE id = %s AND questionnaire_id = %s AND is_template = TRUE
                        """,
                        (new_order, question_id, int(template['id'])),
                    )
                    conn.commit()
                    if cursor.rowcount == 0:
                        return {"success": False, "error": "Question does not belong to this license"}, 403
                    return {"success": True, "message": "Question order updated"}
                finally:
                    conn.close()
                    
            except Exception as e:
                return {"success": False, "error": str(e)}, 400
                
        return {"success": False, "error": "Method not allowed"}, 405

    # -------------------------
    # FEEDBACK VIEWER (ADMIN)
    # -------------------------
    def fetch_responses_for_store(store_id: int, limit: int = 50, status: str = None) -> List[Dict[str, Any]]:
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            
            if status == "unresolved":
                cursor.execute(
                    """
                    SELECT id, questionnaire_id, store_id, user_email, receipt_number, submitted_at, status
                    FROM responses
                    WHERE store_id = %s AND status = 'unresolved'
                    ORDER BY submitted_at DESC, id DESC
                    LIMIT %s
                    """,
                    (store_id, limit),
                )
            elif status == "resolved":
                cursor.execute(
                    """
                    SELECT id, questionnaire_id, store_id, user_email, receipt_number, submitted_at, status
                    FROM responses
                    WHERE store_id = %s AND status = 'resolved'
                    ORDER BY submitted_at DESC, id DESC
                    LIMIT %s
                    """,
                    (store_id, limit),
                )
            else:
                cursor.execute(
                    """
                    SELECT id, questionnaire_id, store_id, user_email, receipt_number, submitted_at, status
                    FROM responses
                    WHERE store_id = %s
                    ORDER BY submitted_at DESC, id DESC
                    LIMIT %s
                    """,
                    (store_id, limit),
                )
            rows = cursor.fetchall()
        finally:
            conn.close()
        return rows

    def fetch_answers_for_responses(response_ids: List[int]) -> Dict[int, List[Dict[str, Any]]]:
        if not response_ids:
            return {}
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            placeholders = ", ".join(["%s"] * len(response_ids))
            cursor.execute(
                f"""
                SELECT a.response_id, a.question_id, a.answer_text, a.rating_value, q.question_text, q.question_type
                FROM answers a
                JOIN questions q ON q.id = a.question_id
                WHERE a.response_id IN ({placeholders})
                ORDER BY a.response_id ASC, q.question_order ASC, a.id ASC
                """,
                tuple(response_ids),
            )
            rows = cursor.fetchall()
        finally:
            conn.close()

        by_response: Dict[int, List[Dict[str, Any]]] = {}
        for r in rows:
            rid = int(r["response_id"])
            by_response.setdefault(rid, []).append(r)
        return by_response

    @app.route("/admin/stores/<int:store_id>/details", methods=["GET"])
    def store_details(store_id: int):
        store = fetch_store_by_id(store_id=store_id)
        if not store:
            flash("Store not found.", "danger")
            return redirect(url_for("admin_dashboard"))

        # Fetch recent feedback
        recent_feedback = fetch_responses_for_store(store_id=store_id, limit=5)
        
        # Calculate analytics data
        all_feedback = fetch_responses_for_store(store_id=store_id, limit=1000)
        total_feedback = len(all_feedback)
        
        # Calculate average rating
        avg_rating = 0
        answers_by_response_id = {}
        if all_feedback:
            all_response_ids = [int(r["id"]) for r in all_feedback]
            answers_by_response_id = fetch_answers_for_responses(all_response_ids)
            all_ratings = []
            for response_id, answers in answers_by_response_id.items():
                for answer in answers:
                    if answer.get("rating_value"):
                        all_ratings.append(float(answer["rating_value"]))
            if all_ratings:
                avg_rating = sum(all_ratings) / len(all_ratings)
        
        # Enhanced 5-star rating analytics
        rating_distribution = [0, 0, 0, 0, 0]  # 1-5 stars
        total_ratings = 0
        for response_id, answers in answers_by_response_id.items():
            for answer in answers:
                if answer.get("rating_value"):
                    rating = int(float(answer["rating_value"]))
                    if 1 <= rating <= 5:
                        rating_distribution[rating - 1] += 1
                        total_ratings += 1
        
        # Calculate 5-star specific metrics
        five_star_count = rating_distribution[4]  # 5 stars
        four_star_count = rating_distribution[3]  # 4 stars
        three_star_count = rating_distribution[2]  # 3 stars
        two_star_count = rating_distribution[1]   # 2 stars
        one_star_count = rating_distribution[0]    # 1 star
        
        # Calculate percentages
        five_star_percentage = (five_star_count / total_ratings * 100) if total_ratings > 0 else 0
        four_plus_star_percentage = ((four_star_count + five_star_count) / total_ratings * 100) if total_ratings > 0 else 0
        three_plus_star_percentage = ((three_star_count + four_star_count + five_star_count) / total_ratings * 100) if total_ratings > 0 else 0
        
        # Rating quality score (weighted average)
        rating_quality_score = (
            (one_star_count * 1) +
            (two_star_count * 2) +
            (three_star_count * 3) +
            (four_star_count * 4) +
            (five_star_count * 5)
        ) / total_ratings if total_ratings > 0 else 0
        
        # Fetch staff members
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, first_name, last_name, email, phone, position, role, status
            FROM staff 
            WHERE store_id = %s
            ORDER BY role DESC, last_name, first_name
        """, (store_id,))
        staff_members = cursor.fetchall()
        cursor.close()
        conn.close()
        
        total_staff = len(staff_members)
        
        # Fetch commendations
        commendations_by_response_id = {}
        commendations = []
        if all_feedback:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            response_ids = [int(r["id"]) for r in all_feedback]
            placeholders = ','.join(['%s'] * len(response_ids))
            cursor.execute(f"""
                SELECT sc.*, s.first_name, s.last_name, s.position, s.role
                FROM staff_commendations sc
                JOIN staff s ON sc.staff_id = s.id
                WHERE sc.response_id IN ({placeholders})
                ORDER BY sc.created_at DESC
            """, response_ids)
            commendations = cursor.fetchall()
            cursor.close()
            conn.close()
            
            for commendation in commendations:
                response_id = commendation['response_id']
                if response_id not in commendations_by_response_id:
                    commendations_by_response_id[response_id] = []
                commendations_by_response_id[response_id].append(commendation)
        
        total_commendations = sum(len(comms) for comms in commendations_by_response_id.values())
        
        # Staff analytics - commendations per staff
        staff_commendations = {}
        if commendations:
            for commendation in commendations:
                staff_id = commendation['staff_id']
                if staff_id not in staff_commendations:
                    staff_commendations[staff_id] = {
                        'staff_name': f"{commendation['first_name']} {commendation['last_name']}",
                        'staff_position': commendation['position'] or commendation['role'].title(),
                        'total_commendations': 0,
                        'rating_sum': 0,
                        'commendation_types': {},
                        'comments': []
                    }
                
                staff_commendations[staff_id]['total_commendations'] += 1
                staff_commendations[staff_id]['rating_sum'] += (commendation['rating'] or 5)
                
                # Count by type
                c_type = commendation['commendation_type']
                if c_type not in staff_commendations[staff_id]['commendation_types']:
                    staff_commendations[staff_id]['commendation_types'][c_type] = 0
                staff_commendations[staff_id]['commendation_types'][c_type] += 1
                
                # Collect comments
                if commendation['comment']:
                    staff_commendations[staff_id]['comments'].append(commendation['comment'])
        
        # Calculate avg_rating and weighted_score for each staff, sort by weighted_score
        for staff_id in staff_commendations:
            if staff_commendations[staff_id]['total_commendations'] > 0:
                staff_commendations[staff_id]['avg_rating'] = staff_commendations[staff_id]['rating_sum'] / staff_commendations[staff_id]['total_commendations']
                staff_commendations[staff_id]['weighted_score'] = staff_commendations[staff_id]['avg_rating'] * (staff_commendations[staff_id]['total_commendations'] ** 0.5)
            else:
                staff_commendations[staff_id]['avg_rating'] = 0
                staff_commendations[staff_id]['weighted_score'] = 0
        top_staff = sorted(staff_commendations.values(), key=lambda x: x['weighted_score'], reverse=True)
        
        # Identify staff with potential issues (low or no commendations)
        staff_performance = []
        for staff_member in staff_members:
            staff_id = staff_member['id']
            staff_name = f"{staff_member['first_name']} {staff_member['last_name']}"
            staff_position = staff_member['position'] or staff_member['role'].title()
            
            commendation_count = staff_commendations.get(staff_id, {}).get('total_commendations', 0)
            
            # Calculate performance score based on commendations
            performance_score = 'excellent'
            if commendation_count == 0:
                performance_score = 'needs_attention'
            elif commendation_count < 3:
                performance_score = 'average'
            
            staff_performance.append({
                'staff_id': staff_id,
                'staff_name': staff_name,
                'staff_position': staff_position,
                'commendation_count': commendation_count,
                'performance_score': performance_score,
                'role': staff_member['role']
            })
        
        # Sort by performance (those needing attention first)
        staff_performance.sort(key=lambda x: (x['performance_score'] != 'needs_attention', x['commendation_count']))
        
        # Calculate metrics (mock data for now)
        resolution_rate = 85 if total_feedback > 0 else 0
        response_time = 2.5
        commendation_rate = round((total_commendations / total_feedback * 100) if total_feedback > 0 else 0)
        repeat_rate = 42
        
        # Feedback trend data (mock data for now)
        feedback_trend_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']
        feedback_trend_data = [12, 19, 15, 25, 22, 30]
        
        # Generate QR code for the store
        public_url = get_store_public_url(store_id=store['id'])
        qr_data_uri = generate_qr_data_uri(public_url)
        
        return render_template(
            "manage_stores/store_details.html",
            store=store,
            recent_feedback=recent_feedback,
            total_feedback=total_feedback,
            avg_rating=avg_rating,
            rating_distribution=rating_distribution,
            staff_members=staff_members,
            total_staff=total_staff,
            total_commendations=total_commendations,
            resolution_rate=resolution_rate,
            response_time=response_time,
            commendation_rate=commendation_rate,
            repeat_rate=repeat_rate,
            feedback_trend_labels=feedback_trend_labels,
            feedback_trend_data=feedback_trend_data,
            public_url=public_url,
            qr_data_uri=qr_data_uri,
            top_staff=top_staff,
            staff_performance=staff_performance,
            staff_commendations=staff_commendations,
            # Enhanced 5-star rating analytics
            total_ratings=total_ratings,
            five_star_count=five_star_count,
            five_star_percentage=five_star_percentage,
            four_plus_star_percentage=four_plus_star_percentage,
            three_plus_star_percentage=three_plus_star_percentage,
            rating_quality_score=rating_quality_score,
        )

    @app.route("/admin/stores/<int:store_id>/feedback", methods=["GET"])
    @login_required
    def store_feedback(store_id: int):
        if not can_manage_store_staff(session['user_id'], store_id):
            flash("You can only access your assigned store.", "danger")
            return redirect(url_for("stores_management"))
        # Handle marking a specific notification as read if requested
        mark_read_id = request.args.get('mark_read')
        if mark_read_id:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("UPDATE responses SET is_read = TRUE WHERE id = %s", (int(mark_read_id),))
                conn.commit()
                conn.close()
            except Exception as e:
                logger.error(f"Error marking specific notification as read: {e}")

        store = fetch_store_by_id(store_id=store_id)
        if not store:
            flash("Store not found.", "danger")
            return redirect(url_for("admin_dashboard"))

        status = request.args.get('status', 'all')
        responses = fetch_responses_for_store(store_id=store_id, limit=50, status=status)
        answers_by_response_id = fetch_answers_for_responses([int(r["id"]) for r in responses])
        
        # Fetch staff commendations for these responses
        commendations_by_response_id = {}
        if responses:
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)
            response_ids = [int(r["id"]) for r in responses]
            placeholders = ','.join(['%s'] * len(response_ids))
            cursor.execute(f"""
                SELECT sc.*, s.first_name, s.last_name, s.position, s.role
                FROM staff_commendations sc
                JOIN staff s ON sc.staff_id = s.id
                WHERE sc.response_id IN ({placeholders})
                ORDER BY sc.created_at DESC
            """, response_ids)
            commendations = cursor.fetchall()
            cursor.close()
            conn.close()
            
            for commendation in commendations:
                response_id = commendation['response_id']
                if response_id not in commendations_by_response_id:
                    commendations_by_response_id[response_id] = []
                commendations_by_response_id[response_id].append(commendation)

        # Calculate staff count and average rating
        staff_count = get_staff_count_for_store(store_id)
        staff_performance = get_staff_performance_for_store(store_id)
        
        # Calculate average rating from responses
        avg_ratings = []
        for response in responses:
            response_answers = answers_by_response_id.get(response["id"], [])
            rating_answers = [a for a in response_answers if a.get("question_type") == "rating"]
            if rating_answers:
                avg_rating = sum(a.get("rating_value", 0) for a in rating_answers) / len(rating_answers)
                avg_ratings.append(avg_rating)
        
        average_rating = round(sum(avg_ratings) / len(avg_ratings), 1) if avg_ratings else 0.0

        return render_template(
            "manage_stores/feedback.html",
            store=store,
            responses=responses,
            answers_by_response_id=answers_by_response_id,
            commendations_by_response_id=commendations_by_response_id,
            current_status=status,
            staff_count=staff_count,
            staff_performance=staff_performance,
            average_rating=average_rating,
        )

    @app.route("/admin/responses/<int:response_id>/status", methods=["POST"])
    def update_response_status(response_id: int):
        try:
            data = request.get_json()
            new_status = data.get('status')
            
            if new_status not in ['resolved', 'unresolved']:
                return {"success": False, "error": "Invalid status"}, 400
            
            conn = get_db_connection()
            try:
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE responses SET status = %s WHERE id = %s",
                    (new_status, response_id)
                )
                conn.commit()
                flash(f"Feedback marked as {new_status}", "success")
                return {"success": True}
            finally:
                conn.close()
                
        except Exception as e:
            return {"success": False, "error": str(e)}, 500

    @app.route("/api/notifications/unread")
    @login_required
    def get_unread_notifications():
        """Fetch feedback notifications for the bell icon, combining feedback and system notifications.

        Query params:
          - status: 'unseen' (default returns both for initial load) | 'seen' to fetch a page of seen
          - seen_offset: int, offset into the seen list (default 0)
          - seen_limit: int, page size for seen (default 20, max 50)
        """
        try:
            status = request.args.get('status', 'all')
            seen_offset = max(int(request.args.get('seen_offset', 0)), 0)
            seen_limit = min(max(int(request.args.get('seen_limit', 20)), 1), 50)
        except ValueError:
            status = 'all'
            seen_offset = 0
            seen_limit = 20

        def _format(rows):
            for n in rows:
                if n.get('created_at'):
                    n['created_at'] = n['created_at'].strftime('%b %d, %H:%M')
                else:
                    n['created_at'] = 'N/A'
                if n.get('notification_type') == 'system':
                    n['system_id'] = n['id']
                    n['id'] = None
            return rows

        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)

            unseen = []
            seen = []
            seen_total = 0

            # Get user's assigned stores (applies to all users except superadmin)
            user_id = session.get('user_id')
            user_role = session.get('role')
            assigned_store_ids = []
            if user_id and user_role != 'superadmin':
                # Get stores assigned via user_stores table
                cursor.execute(
                    "SELECT store_id FROM user_stores WHERE user_id = %s",
                    (user_id,)
                )
                assigned_store_ids = [row['store_id'] for row in cursor.fetchall()]

                # Also include stores where the user is the owner (stores.user_id)
                cursor.execute(
                    "SELECT id FROM stores WHERE user_id = %s",
                    (user_id,)
                )
                owned_store_ids = [row['id'] for row in cursor.fetchall()]
                # Combine and deduplicate
                assigned_store_ids = list(set(assigned_store_ids + owned_store_ids))

            # Build WHERE clause for store filtering
            store_filter = ""
            if assigned_store_ids:
                placeholders = ','.join(['%s'] * len(assigned_store_ids))
                store_filter = f"AND r.store_id IN ({placeholders})"
            elif user_role != 'superadmin':
                # Non-superadmin with no assigned stores - show no feedback notifications
                store_filter = "AND 1=0"

            if status in ('all', 'unseen'):
                # Fetch ALL unseen feedback responses (no cap so user always sees them)
                query = f"""
                    SELECT r.id, r.user_email, r.submitted_at as created_at, s.store_name, s.id as store_id, r.is_read, 'feedback' as notification_type, NULL as message, NULL as type,
                           (SELECT AVG(a.rating_value) FROM answers a WHERE a.response_id = r.id AND a.rating_value IS NOT NULL) as avg_rating
                    FROM responses r
                    JOIN stores s ON r.store_id = s.id
                    WHERE s.store_name IS NOT NULL AND r.is_read = FALSE {store_filter}
                    ORDER BY r.submitted_at DESC
                    """
                if assigned_store_ids:
                    cursor.execute(query, assigned_store_ids)
                else:
                    cursor.execute(query)
                unseen_feedback = cursor.fetchall()

                cursor.execute(
                    """
                    SELECT id, message, type, created_at, is_read, 'system' as notification_type, NULL as user_email, NULL as store_name, NULL as store_id
                    FROM system_notifications
                    WHERE is_read = FALSE
                    ORDER BY created_at DESC
                    """
                )
                unseen_system = cursor.fetchall()

                unseen = sorted(
                    unseen_feedback + unseen_system,
                    key=lambda x: x['created_at'] or datetime.min,
                    reverse=True,
                )
                _format(unseen)

            if status in ('all', 'seen'):
                # Optional cutoff: hide seen notifications submitted at or before this time
                cleared_at = session.get('notifications_cleared_at')
                cleared_dt = None
                if cleared_at:
                    try:
                        cleared_dt = datetime.fromisoformat(cleared_at)
                    except (ValueError, TypeError):
                        cleared_dt = None

                # Total seen count (for "load more" UI), respecting cleared cutoff
                if cleared_dt:
                    if assigned_store_ids:
                        placeholders = ','.join(['%s'] * len(assigned_store_ids))
                        cursor.execute(
                            f"SELECT COUNT(*) as count FROM responses r JOIN stores s ON r.store_id = s.id WHERE s.store_name IS NOT NULL AND r.is_read = TRUE AND r.submitted_at > %s AND r.store_id IN ({placeholders})",
                            [cleared_dt] + assigned_store_ids
                        )
                    else:
                        cursor.execute(
                            "SELECT COUNT(*) as count FROM responses r JOIN stores s ON r.store_id = s.id WHERE s.store_name IS NOT NULL AND r.is_read = TRUE AND r.submitted_at > %s" + (" AND 1=0" if user_role != 'superadmin' else ""),
                            (cleared_dt,)
                        )
                else:
                    if assigned_store_ids:
                        placeholders = ','.join(['%s'] * len(assigned_store_ids))
                        cursor.execute(
                            f"SELECT COUNT(*) as count FROM responses r JOIN stores s ON r.store_id = s.id WHERE s.store_name IS NOT NULL AND r.is_read = TRUE AND r.store_id IN ({placeholders})",
                            assigned_store_ids
                        )
                    else:
                        cursor.execute(
                            "SELECT COUNT(*) as count FROM responses r JOIN stores s ON r.store_id = s.id WHERE s.store_name IS NOT NULL AND r.is_read = TRUE" + (" AND 1=0" if user_role != 'superadmin' else "")
                        )
                seen_feedback_total = cursor.fetchone()['count']

                if cleared_dt:
                    cursor.execute(
                        "SELECT COUNT(*) as count FROM system_notifications WHERE is_read = TRUE AND created_at > %s",
                        (cleared_dt,)
                    )
                else:
                    cursor.execute("SELECT COUNT(*) as count FROM system_notifications WHERE is_read = TRUE")
                seen_system_total = cursor.fetchone()['count']
                seen_total = seen_feedback_total + seen_system_total

                # Fetch a window of seen notifications. Over-fetch from each table then merge,
                # so the merged page reflects true chronological order across both sources.
                window = seen_limit + 1  # fetch one extra to detect if there's more

                # Fetch seen feedback responses with store filter
                if cleared_dt:
                    try:
                        cleared_dt_parsed = datetime.fromisoformat(cleared_at)
                        query = f"""
                            SELECT r.id, r.user_email, r.submitted_at as created_at, s.store_name, s.id as store_id, r.is_read, 'feedback' as notification_type, NULL as message, NULL as type,
                                   (SELECT AVG(a.rating_value) FROM answers a WHERE a.response_id = r.id AND a.rating_value IS NOT NULL) as avg_rating
                            FROM responses r
                            JOIN stores s ON r.store_id = s.id
                            WHERE s.store_name IS NOT NULL AND r.is_read = TRUE AND r.submitted_at > %s {store_filter}
                            ORDER BY r.submitted_at DESC
                            LIMIT %s
                            """
                        if assigned_store_ids:
                            cursor.execute(query, [cleared_dt_parsed] + assigned_store_ids + [window])
                        else:
                            cursor.execute(query, [cleared_dt_parsed, window])
                    except ValueError:
                        cleared_dt = None
                        query = f"""
                            SELECT r.id, r.user_email, r.submitted_at as created_at, s.store_name, s.id as store_id, r.is_read, 'feedback' as notification_type, NULL as message, NULL as type,
                                   (SELECT AVG(a.rating_value) FROM answers a WHERE a.response_id = r.id AND a.rating_value IS NOT NULL) as avg_rating
                            FROM responses r
                            JOIN stores s ON r.store_id = s.id
                            WHERE s.store_name IS NOT NULL AND r.is_read = TRUE {store_filter}
                            ORDER BY r.submitted_at DESC
                            LIMIT %s
                            """
                        if assigned_store_ids:
                            cursor.execute(query, assigned_store_ids + [window])
                        else:
                            cursor.execute(query, [window])
                else:
                    query = f"""
                        SELECT r.id, r.user_email, r.submitted_at as created_at, s.store_name, s.id as store_id, r.is_read, 'feedback' as notification_type, NULL as message, NULL as type,
                               (SELECT AVG(a.rating_value) FROM answers a WHERE a.response_id = r.id AND a.rating_value IS NOT NULL) as avg_rating
                        FROM responses r
                        JOIN stores s ON r.store_id = s.id
                        WHERE s.store_name IS NOT NULL AND r.is_read = TRUE {store_filter}
                        ORDER BY r.submitted_at DESC
                        LIMIT %s
                        """
                    if assigned_store_ids:
                        cursor.execute(query, assigned_store_ids + [window])
                    else:
                        cursor.execute(query, [window])
                seen_feedback = cursor.fetchall()

                cursor.execute(
                    """
                    SELECT id, message, type, created_at, is_read, 'system' as notification_type, NULL as user_email, NULL as store_name, NULL as store_id
                    FROM system_notifications
                    WHERE is_read = TRUE
                    ORDER BY created_at DESC
                    LIMIT %s
                    """,
                    (window,)
                )
                seen_system = cursor.fetchall()

                merged_seen = sorted(
                    seen_feedback + seen_system,
                    key=lambda x: x['created_at'] or datetime.min,
                    reverse=True,
                )
                seen = merged_seen[seen_offset:seen_offset + seen_limit]
                _format(seen)

            # Total unread count (always returned) - filter by assigned stores for view-only users
            if assigned_store_ids:
                placeholders = ','.join(['%s'] * len(assigned_store_ids))
                cursor.execute(
                    f"SELECT COUNT(*) as count FROM responses r JOIN stores s ON r.store_id = s.id WHERE s.store_name IS NOT NULL AND r.is_read = FALSE AND r.store_id IN ({placeholders})",
                    assigned_store_ids
                )
            elif user_role != 'superadmin':
                cursor.execute(
                    "SELECT COUNT(*) as count FROM responses r JOIN stores s ON r.store_id = s.id WHERE s.store_name IS NOT NULL AND r.is_read = FALSE AND 1=0"
                )
            else:
                cursor.execute(
                    "SELECT COUNT(*) as count FROM responses r JOIN stores s ON r.store_id = s.id WHERE s.store_name IS NOT NULL AND r.is_read = FALSE"
                )
            unread_feedback_count = cursor.fetchone()['count']
            cursor.execute("SELECT COUNT(*) as count FROM system_notifications WHERE is_read = FALSE")
            unread_system_count = cursor.fetchone()['count']
            total_unread = unread_feedback_count + unread_system_count

            seen_has_more = (seen_offset + len(seen)) < seen_total

            return jsonify({
                "success": True,
                "unseen": unseen,
                "seen": seen,
                "seen_total": seen_total,
                "seen_has_more": seen_has_more,
                "total_unread": total_unread,
                # Backwards-compat: combined list (unseen first, then seen page)
                "notifications": unseen + seen,
            })
        except Exception as e:
            logger.error(f"Error fetching notifications: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
        finally:
            conn.close()

    @app.route("/api/notifications/clear-seen", methods=["POST"])
    def clear_seen_notifications():
        """Clear (hide) all currently seen notifications for this session.
        Stores a cutoff timestamp; any seen notifications with created_at <= cutoff
        are excluded from the seen list returned by /api/notifications/unread."""
        try:
            session['notifications_cleared_at'] = datetime.now().isoformat()
            return jsonify({"success": True})
        except Exception as e:
            logger.error(f"Error clearing seen notifications: {e}")
            return jsonify({"success": False, "error": str(e)}), 500

    @app.route("/admin/responses/<int:response_id>/reply", methods=["POST"])
    def reply_to_feedback(response_id: int):
        try:
            data = request.get_json()
            reply_message = data.get('message', '').strip()
            template_type = data.get('template_type', 'standard')
            cc_emails = data.get('cc_emails', [])
            bcc_emails = data.get('bcc_emails', [])
            
            if not reply_message:
                return {"success": False, "error": "Reply message cannot be empty"}, 400
            
            if template_type not in ['standard', 'apology', 'appreciation', 'follow_up']:
                template_type = 'standard'
            
            conn = get_db_connection()
            try:
                cursor = conn.cursor(dictionary=True)
                
                # Get response details including customer email and store info
                cursor.execute("""
                    SELECT r.user_email, r.submitted_at, s.store_name,
                           (SELECT GROUP_CONCAT(a.answer_text SEPARATOR ' ') 
                            FROM answers a 
                            WHERE a.response_id = r.id 
                            AND a.answer_text IS NOT NULL 
                            LIMIT 3) as feedback_summary,
                           (SELECT AVG(a.rating_value) 
                            FROM answers a 
                            WHERE a.response_id = r.id 
                            AND a.rating_value IS NOT NULL) as avg_rating
                    FROM responses r
                    JOIN stores s ON r.store_id = s.id
                    WHERE r.id = %s
                """, (response_id,))
                
                response = cursor.fetchone()
                
                if not response:
                    return {"success": False, "error": "Response not found"}, 404
                
                if not response['user_email']:
                    return {"success": False, "error": "No email address found for this feedback"}, 400
                
                # Extract customer name from email
                customer_name = response['user_email'].split('@')[0].replace('.', ' ').title()
                
                # Auto-select template based on rating if not specified
                if template_type == 'standard' and response['avg_rating'] is not None:
                    try:
                        avg_rating = float(response['avg_rating'])
                        if avg_rating <= 2:
                            template_type = 'apology'
                        elif avg_rating >= 4:
                            template_type = 'appreciation'
                        else:
                            template_type = 'follow_up'
                    except (ValueError, TypeError):
                        template_type = 'standard'
                
                # Send email using API or SMTP
                try:
                    success, message = email_config.send_feedback_reply(
                        to_email=response['user_email'],
                        customer_name=customer_name,
                        reply_message=reply_message,
                        store_name=response['store_name'],
                        feedback_summary=response['feedback_summary'],
                        template_type=template_type
                    )
                    if success:
                        return {"success": True, "message": "Reply sent successfully", "template_used": template_type}
                    else:
                        return {"success": False, "error": message}, 500
                except Exception as e:
                    logger.error(f"Email sending failed: {str(e)}")
                    return {"success": False, "error": str(e)}, 500
                    
            finally:
                conn.close()
                
        except Exception as e:
            return {"success": False, "error": str(e)}, 500
    
    @app.route("/admin/email/statistics", methods=["GET"])
    def email_statistics():
        """Get email sending statistics"""
        try:
            stats = email_config.get_email_statistics()
            return jsonify(stats)
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    @app.route("/api/system_notifications/<int:notification_id>/read", methods=["POST"])
    def mark_system_notification_read(notification_id: int):
        """Mark a single system notification as read."""
        conn = get_db_connection()
        try:
            cursor = conn.cursor()
            cursor.execute("UPDATE system_notifications SET is_read = TRUE WHERE id = %s", (notification_id,))
            conn.commit()
            return jsonify({"success": True})
        except Exception as e:
            logger.error(f"Error marking system notification as read: {e}")
            return jsonify({"success": False, "error": str(e)}), 500
        finally:
            conn.close()
    
    @app.route("/admin/email/bulk-reply", methods=["POST"])
    def bulk_reply_to_feedback():
        """Send bulk email replies to multiple feedback responses"""
        try:
            data = request.get_json()
            response_ids = data.get('response_ids', [])
            reply_message = data.get('message', '').strip()
            template_type = data.get('template_type', 'standard')
            
            if not response_ids:
                return {"success": False, "error": "No response IDs provided"}, 400
            
            if not reply_message:
                return {"success": False, "error": "Reply message cannot be empty"}, 400
            
            if template_type not in ['standard', 'apology', 'appreciation', 'follow_up']:
                template_type = 'standard'
            
            conn = get_db_connection()
            try:
                cursor = conn.cursor(dictionary=True)
                
                # Get all response details
                placeholders = ", ".join(["%s"] * len(response_ids))
                cursor.execute(f"""
                    SELECT r.id, r.user_email, s.store_name,
                           (SELECT GROUP_CONCAT(a.answer_text SEPARATOR ' ') 
                            FROM answers a 
                            WHERE a.response_id = r.id 
                            AND a.answer_text IS NOT NULL 
                            LIMIT 3) as feedback_summary
                    FROM responses r
                    JOIN stores s ON r.store_id = s.id
                    WHERE r.id IN ({placeholders}) AND r.user_email IS NOT NULL
                """, tuple(response_ids))
                
                responses = cursor.fetchall()
                
                if not responses:
                    return {"success": False, "error": "No valid responses found"}, 404
                
                # Prepare data for bulk email
                email_list = [r['user_email'] for r in responses]
                customer_names = [r['user_email'].split('@')[0].replace('.', ' ').title() for r in responses]
                feedback_summaries = [r['feedback_summary'] or "No text feedback provided" for r in responses]
                store_name = responses[0]['store_name']  # Use first store name (assuming same store)
                
                # Send bulk emails
                results = email_config.send_bulk_feedback_reply(
                    email_list=email_list,
                    customer_names=customer_names,
                    reply_message=reply_message,
                    store_name=store_name,
                    feedback_summaries=feedback_summaries,
                    template_type=template_type
                )
                
                # Mark responses as resolved
                successful_emails = [r['email'] for r in results if r['success']]
                if successful_emails:
                    placeholders = ", ".join(["%s"] * len(successful_emails))
                    cursor.execute(f"""
                        UPDATE responses 
                        SET status = 'resolved' 
                        WHERE user_email IN ({placeholders})
                    """, tuple(successful_emails))
                    conn.commit()
                
                return {
                    "success": True,
                    "message": f"Bulk reply completed. {len(successful_emails)} of {len(results)} emails sent successfully.",
                    "results": results
                }
                
            finally:
                conn.close()
                
        except Exception as e:
            return {"success": False, "error": str(e)}, 500

    return app


if __name__ == "__main__":
    app = create_app()
    # Railway provides the port via the PORT environment variable
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=False)
else:
    # This is used by gunicorn (web: gunicorn "app:create_app()")
    app = create_app()
